#!/usr/bin/env python3
"""
DNAAS Path Auto-Discovery Script

Automatically discovers DNAAS paths by:
1. Accepting device serial numbers as input
2. SSHing to discover device info and DNAAS-connected interfaces
3. Tracing through the DNAAS fabric via LLDP
4. Outputting results in multiple formats (JSON, Text, Excel, Terminal)

Usage:
    python3 dnaas_path_discovery.py <serial1> [serial2] [--use-cache]
    
Example:
    python3 dnaas_path_discovery.py wk31d7vv00023 100.64.0.220
"""

import paramiko
import time
import re
import json
import os
import sys
import argparse
import threading
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, field, asdict
from pathlib import Path

# Try to import openpyxl for Excel output
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel output will be skipped.")
    print("Install with: pip install openpyxl")

# ============================================================================
# CONFIGURATION
# ============================================================================

# Default credentials
DEFAULT_USER = "dnroot"
DEFAULT_PASS = "dnroot"

# DNAAS credentials (different from PE credentials)
DNAAS_USER = "sisaev"
DNAAS_PASS = "Drive1234!"

# Paths
SCRIPT_DIR = Path(__file__).parent.resolve()
INVENTORY_FILE = SCRIPT_DIR / "device_inventory.json"
CREDENTIALS_FILE = SCRIPT_DIR / "device_credentials.json"
OUTPUT_DIR = SCRIPT_DIR / "output"
DNAAS_TABLE_FILE = SCRIPT_DIR / "dnaas_table.xlsx"

# DNAAS detection keywords
DNAAS_KEYWORDS = ['DNAAS', 'LEAF', 'SPINE', 'FABRIC', 'TOR', 'AGGREGATION', 'SUPERSPINE']

# Terminal colors
class Colors:
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    DIM = '\033[2m'

# ============================================================================
# CREDENTIAL MANAGER
# ============================================================================

class CredentialManager:
    """
    Manages device credentials for fast reuse.
    Stores credentials in device_credentials.json with format:
    {
        "serial_or_hostname": {
            "hostname": "...",
            "mgmt_ip": "...",
            "user": "...",
            "password": "...",
            "device_type": "DUT|DNAAS",
            "last_used": "2025-12-30T..."
        }
    }
    """
    
    def __init__(self, filepath: Path = CREDENTIALS_FILE):
        self.filepath = filepath
        self.credentials: Dict[str, Dict[str, Any]] = {}
        self._load()
    
    def _load(self):
        """Load credentials from file"""
        if self.filepath.exists():
            try:
                with open(self.filepath, 'r') as f:
                    data = json.load(f)
                    self.credentials = data.get('devices', {})
                print(f"{Colors.CYAN}Loaded credentials for {len(self.credentials)} devices{Colors.ENDC}")
            except Exception as e:
                print(f"{Colors.YELLOW}Warning: Could not load credentials: {e}{Colors.ENDC}")
                self.credentials = {}
    
    def _save(self):
        """Save credentials to file"""
        try:
            data = {
                'devices': self.credentials,
                'last_updated': datetime.now().isoformat()
            }
            with open(self.filepath, 'w') as f:
                json.dump(data, f, indent=2)
        except Exception as e:
            print(f"{Colors.RED}Error saving credentials: {e}{Colors.ENDC}")
    
    def get(self, device_id: str) -> Optional[Dict[str, Any]]:
        """
        Get credentials for a device by serial, hostname, or IP.
        Returns dict with hostname, mgmt_ip, user, password or None.
        """
        device_id_lower = device_id.lower()
        
        # Direct lookup by key (serial, hostname, or IP)
        if device_id_lower in self.credentials:
            return self.credentials[device_id_lower]
        
        # Search by hostname or mgmt_ip
        for key, creds in self.credentials.items():
            if creds.get('hostname', '').lower() == device_id_lower:
                return creds
            if creds.get('mgmt_ip', '') == device_id:
                return creds
        
        return None
    
    def save_credentials(self, device_id: str, hostname: str, mgmt_ip: str,
                        user: str, password: str, device_type: str = "DUT"):
        """
        Save credentials for a device.
        device_type: "DUT" for user-owned devices, "DNAAS" for fabric devices
        """
        self.credentials[device_id.lower()] = {
            'hostname': hostname,
            'mgmt_ip': mgmt_ip,
            'user': user,
            'password': password,
            'device_type': device_type,
            'last_used': datetime.now().isoformat()
        }
        self._save()
        print(f"{Colors.GREEN}Saved credentials for {hostname} ({device_type}){Colors.ENDC}")
    
    def update_last_used(self, device_id: str):
        """Update the last_used timestamp for a device"""
        device_id_lower = device_id.lower()
        if device_id_lower in self.credentials:
            self.credentials[device_id_lower]['last_used'] = datetime.now().isoformat()
            self._save()
    
    def list_devices(self) -> List[Dict[str, Any]]:
        """List all stored devices with their credentials (passwords masked)"""
        devices = []
        for device_id, creds in self.credentials.items():
            devices.append({
                'id': device_id,
                'hostname': creds.get('hostname', ''),
                'mgmt_ip': creds.get('mgmt_ip', ''),
                'device_type': creds.get('device_type', 'DUT'),
                'last_used': creds.get('last_used', '')
            })
        return devices
    
    def remove(self, device_id: str) -> bool:
        """Remove credentials for a device"""
        device_id_lower = device_id.lower()
        if device_id_lower in self.credentials:
            del self.credentials[device_id_lower]
            self._save()
            return True
        return False

# Global credential manager instance
credential_manager = CredentialManager()

# ============================================================================
# DUT CONFIGURATOR
# ============================================================================

class DUTConfigurator:
    """
    Handles configuration changes on DUT (user-owned) devices only.
    IMPORTANT: This class should NEVER be used on DNAAS devices.
    
    Capabilities:
    - Enable physical interfaces
    - Add interfaces to LLDP
    - Commit configuration changes
    """
    
    def __init__(self, ssh_client: paramiko.SSHClient, hostname: str):
        self.ssh = ssh_client
        self.hostname = hostname
        self.pending_changes: List[str] = []
        self.config_log: List[str] = []
    
    def _run_command(self, command: str, timeout: int = 30) -> Tuple[str, bool]:
        """Run a command and return output and success status"""
        try:
            stdin, stdout, stderr = self.ssh.exec_command(command, timeout=timeout)
            output = stdout.read().decode('utf-8', errors='ignore')
            error = stderr.read().decode('utf-8', errors='ignore')
            
            # Check for errors
            if 'error' in error.lower() or 'failed' in error.lower():
                return error, False
            if 'error' in output.lower() and 'commit' in command:
                return output, False
                
            return output, True
        except Exception as e:
            return str(e), False
    
    def get_physical_interfaces(self) -> List[Dict[str, Any]]:
        """
        Get list of physical interfaces (ge100, ge400, etc.) with their status.
        Returns list of dicts with: name, admin_state, oper_state
        """
        output, success = self._run_command("show interfaces | no more")
        if not success:
            print(f"{Colors.RED}Failed to get interfaces{Colors.ENDC}")
            return []
        
        interfaces = []
        current_if = None
        
        for line in output.split('\n'):
            # Match physical interface lines (ge100-0/0/1, ge400-0/0/4, etc.)
            # Skip sub-interfaces (those with dots like ge100-0/0/1.100)
            if_match = re.match(r'^(ge\d+-\d+/\d+/\d+)\s', line)
            if if_match:
                if_name = if_match.group(1)
                # Only include if it's a physical interface (no dot = no sub-interface)
                if '.' not in if_name:
                    current_if = {'name': if_name, 'admin_state': '', 'oper_state': ''}
                    interfaces.append(current_if)
            
            # Extract admin/oper state
            if current_if:
                if 'admin-state:' in line.lower():
                    admin_match = re.search(r'admin-state:\s*(\w+)', line, re.I)
                    if admin_match:
                        current_if['admin_state'] = admin_match.group(1).lower()
                elif 'oper-state:' in line.lower():
                    oper_match = re.search(r'oper-state:\s*(\w+)', line, re.I)
                    if oper_match:
                        current_if['oper_state'] = oper_match.group(1).lower()
        
        return interfaces
    
    def get_disabled_interfaces(self) -> List[str]:
        """Get list of physical interfaces that are admin-disabled"""
        interfaces = self.get_physical_interfaces()
        disabled = [iface['name'] for iface in interfaces 
                   if iface.get('admin_state', '').lower() in ['disabled', 'down', '']]
        return disabled
    
    def get_lldp_configured_interfaces(self) -> List[str]:
        """Get list of interfaces configured in LLDP"""
        output, success = self._run_command("show config protocols lldp | no more")
        if not success:
            return []
        
        interfaces = []
        for line in output.split('\n'):
            # Look for interface lines in LLDP config
            match = re.search(r'interface\s+(ge\d+-\d+/\d+/\d+)', line)
            if match:
                interfaces.append(match.group(1))
        
        return interfaces
    
    def get_unconfigured_lldp_interfaces(self) -> List[str]:
        """Get physical interfaces NOT configured in LLDP"""
        all_physical = self.get_physical_interfaces()
        lldp_configured = self.get_lldp_configured_interfaces()
        
        unconfigured = [iface['name'] for iface in all_physical 
                       if iface['name'] not in lldp_configured]
        return unconfigured
    
    def get_interface_details(self, interface: str) -> Dict[str, Any]:
        """
        Get comprehensive interface details for Link Table.
        Uses multiple DNOS commands:
        - show interfaces <interface>
        - show interfaces <interface> transceiver
        - show config interfaces <interface>
        
        Returns dict with: admin_state, oper_state, speed, mtu, fec, transceiver, encapsulation, etc.
        """
        details = {
            'name': interface,
            'admin_state': '',
            'oper_state': '',
            'speed': '',
            'mtu': '',
            'fec': '',
            'transceiver': '',
            'transceiver_vendor': '',
            'transceiver_type': '',
            'encapsulation': '',
            'vlan_id': '',
            'inner_vlan': '',
            'l2_service': False,
            'bundle_parent': '',
            'sub_interfaces': []
        }
        
        # Get basic interface state
        base_iface = interface.split('.')[0]  # Strip sub-interface suffix
        output, success = self._run_command(f"show interfaces {base_iface} | no-more", timeout=15)
        if success:
            for line in output.split('\n'):
                line_lower = line.lower().strip()
                
                # Admin/oper state
                if 'admin-state:' in line_lower:
                    match = re.search(r'admin-state:\s*(\w+)', line, re.I)
                    if match:
                        details['admin_state'] = match.group(1)
                elif 'oper-state:' in line_lower:
                    match = re.search(r'oper-state:\s*(\w+)', line, re.I)
                    if match:
                        details['oper_state'] = match.group(1)
                elif 'speed:' in line_lower:
                    match = re.search(r'speed:\s*(\S+)', line, re.I)
                    if match:
                        details['speed'] = match.group(1)
                elif 'mtu:' in line_lower:
                    match = re.search(r'mtu:\s*(\d+)', line, re.I)
                    if match:
                        details['mtu'] = match.group(1)
        
        # Get transceiver info
        trans_output, trans_success = self._run_command(f"show interfaces {base_iface} transceiver | no-more", timeout=15)
        if trans_success:
            for line in trans_output.split('\n'):
                line_lower = line.lower()
                if 'vendor:' in line_lower or 'manufacturer:' in line_lower:
                    match = re.search(r'(?:vendor|manufacturer):\s*(.+)', line, re.I)
                    if match:
                        details['transceiver_vendor'] = match.group(1).strip()
                elif 'type:' in line_lower or 'module type:' in line_lower:
                    match = re.search(r'(?:type|module type):\s*(\S+)', line, re.I)
                    if match:
                        details['transceiver_type'] = match.group(1)
                elif 'serial' in line_lower:
                    match = re.search(r'serial[^:]*:\s*(\S+)', line, re.I)
                    if match:
                        if not details['transceiver']:
                            details['transceiver'] = match.group(1)
            
            # Combine vendor + type for display
            if details['transceiver_vendor'] or details['transceiver_type']:
                details['transceiver'] = f"{details['transceiver_vendor']} {details['transceiver_type']}".strip()
        
        # Get config for VLAN/encapsulation info (if sub-interface)
        if '.' in interface:
            cfg_output, cfg_success = self._run_command(f"show config interfaces {interface} | no-more", timeout=15)
            if cfg_success:
                for line in cfg_output.split('\n'):
                    line_lower = line.lower()
                    if 'l2-service' in line_lower and 'enabled' in line_lower:
                        details['l2_service'] = True
                    if 'encapsulation' in line_lower:
                        # Parse: encapsulation dot1q 210 or encapsulation dot1q outer-vlan 210 inner-vlan 100
                        vlan_match = re.search(r'dot1q\s+(?:outer-vlan\s+)?(\d+)', line, re.I)
                        if vlan_match:
                            details['vlan_id'] = vlan_match.group(1)
                            details['encapsulation'] = f"dot1q {vlan_match.group(1)}"
                        inner_match = re.search(r'inner-vlan\s+(\d+)', line, re.I)
                        if inner_match:
                            details['inner_vlan'] = inner_match.group(1)
                            details['encapsulation'] += f" inner {inner_match.group(1)}"
        
        # Get FEC configuration
        fec_output, fec_success = self._run_command(f"show config interfaces {base_iface} | include fec | no-more", timeout=10)
        if fec_success:
            fec_match = re.search(r'fec\s+(\S+)', fec_output)
            if fec_match:
                details['fec'] = fec_match.group(1)
        
        return details
    
    def get_link_table_data(self, interface_a: str, interface_b: str, device_a: str = "", device_b: str = "") -> Dict[str, Any]:
        """
        Get all data needed to populate the Link Table for a specific link.
        Returns comprehensive dict with both sides' interface details.
        """
        return {
            'device_a': device_a,
            'device_b': device_b,
            'interface_a': self.get_interface_details(interface_a),
            'interface_b': self.get_interface_details(interface_b) if interface_b else {}
        }

    def enable_interface(self, interface: str) -> bool:
        """
        Enable a physical interface using 'admin-state enabled'.
        Returns True if command was added to pending changes.
        """
        # Validate interface name
        if not re.match(r'^ge\d+-\d+/\d+/\d+$', interface):
            print(f"{Colors.RED}Invalid interface name: {interface}{Colors.ENDC}")
            return False
        
        config_cmd = f"config; interfaces interface {interface} admin-state enabled"
        self.pending_changes.append(config_cmd)
        self.config_log.append(f"[{datetime.now().isoformat()}] Enable interface: {interface}")
        print(f"{Colors.CYAN}Queued: Enable interface {interface}{Colors.ENDC}")
        return True
    
    def add_interface_to_lldp(self, interface: str) -> bool:
        """
        Add interface to LLDP configuration.
        Returns True if command was added to pending changes.
        """
        # Validate interface name
        if not re.match(r'^ge\d+-\d+/\d+/\d+$', interface):
            print(f"{Colors.RED}Invalid interface name: {interface}{Colors.ENDC}")
            return False
        
        config_cmd = f"config; protocols lldp interface {interface}"
        self.pending_changes.append(config_cmd)
        self.config_log.append(f"[{datetime.now().isoformat()}] Add to LLDP: {interface}")
        print(f"{Colors.CYAN}Queued: Add {interface} to LLDP{Colors.ENDC}")
        return True
    
    def commit_changes(self) -> Tuple[bool, str]:
        """
        Commit all pending configuration changes.
        Returns (success, message).
        """
        if not self.pending_changes:
            return True, "No changes to commit"
        
        print(f"\n{Colors.YELLOW}Committing {len(self.pending_changes)} configuration changes on {self.hostname}...{Colors.ENDC}")
        
        # Build combined config command
        # Join all config commands and add final commit
        config_commands = "; ".join(self.pending_changes) + "; commit"
        
        output, success = self._run_command(config_commands, timeout=60)
        
        if success:
            self.config_log.append(f"[{datetime.now().isoformat()}] COMMIT SUCCESS")
            self.pending_changes = []
            print(f"{Colors.GREEN}Configuration committed successfully!{Colors.ENDC}")
            return True, "Committed successfully"
        else:
            self.config_log.append(f"[{datetime.now().isoformat()}] COMMIT FAILED: {output}")
            print(f"{Colors.RED}Commit failed: {output}{Colors.ENDC}")
            return False, output
    
    def auto_configure_for_discovery(self) -> Tuple[bool, List[str]]:
        """
        Automatically configure the device for DNAAS discovery:
        1. Enable all disabled physical interfaces
        2. Add unconfigured interfaces to LLDP
        3. Commit changes
        
        Returns (success, list of configured interfaces)
        """
        configured_interfaces = []
        
        # Step 1: Enable disabled physical interfaces
        disabled = self.get_disabled_interfaces()
        if disabled:
            print(f"\n{Colors.YELLOW}Found {len(disabled)} disabled interfaces to enable:{Colors.ENDC}")
            for iface in disabled:
                print(f"  - {iface}")
                self.enable_interface(iface)
                configured_interfaces.append(iface)
        
        # Step 2: Add interfaces to LLDP
        unconfigured_lldp = self.get_unconfigured_lldp_interfaces()
        if unconfigured_lldp:
            print(f"\n{Colors.YELLOW}Found {len(unconfigured_lldp)} interfaces not in LLDP:{Colors.ENDC}")
            for iface in unconfigured_lldp:
                print(f"  - {iface}")
                self.add_interface_to_lldp(iface)
                if iface not in configured_interfaces:
                    configured_interfaces.append(iface)
        
        # Step 3: Commit if there are changes
        if self.pending_changes:
            success, msg = self.commit_changes()
            if not success:
                return False, configured_interfaces
        
        return True, configured_interfaces
    
    def get_config_log(self) -> List[str]:
        """Get the log of all configuration changes made"""
        return self.config_log.copy()

# ============================================================================
# BRIDGE DOMAIN DISCOVERY (DNAAS DEVICES ONLY - SHOW COMMANDS)
# ============================================================================

@dataclass
class BridgeDomain:
    """Bridge Domain information from DNAAS device"""
    bd_id: int
    name: str = ""
    vlan_id: Optional[int] = None
    interfaces: List[str] = field(default_factory=list)
    attachment_circuits: List[Dict[str, Any]] = field(default_factory=list)
    state: str = "unknown"

@dataclass
class LACPBundle:
    """LACP Bundle information"""
    bundle_id: str
    name: str = ""
    members: List[str] = field(default_factory=list)
    state: str = "unknown"
    partner_system_id: str = ""

class BridgeDomainDiscovery:
    """
    Handles Bridge Domain discovery on DNAAS devices.
    IMPORTANT: This class uses ONLY show commands - no configuration changes.
    
    Capabilities:
    - Parse 'show network-services bridge-domain'
    - Resolve LACP bundle members
    - Map BD interfaces to LLDP neighbors
    
    NOTE: Uses interactive shell because DNAAS devices require it.
    """
    
    def __init__(self, ssh_connection, hostname: str):
        """
        Initialize with an SSHConnection object (which has interactive shell)
        or a paramiko.SSHClient (fallback for compatibility)
        """
        self.hostname = hostname
        self._bd_cache: Optional[List[BridgeDomain]] = None
        self._lacp_cache: Dict[str, LACPBundle] = {}
        
        # Handle both SSHConnection and raw paramiko client
        if hasattr(ssh_connection, 'send_command'):
            # It's an SSHConnection - use its shell method
            self._ssh_conn = ssh_connection
            self._use_shell = True
        else:
            # It's a raw paramiko client - create shell
            self._ssh_client = ssh_connection
            self._shell = None
            self._use_shell = False
            self._init_shell()
    
    def _init_shell(self):
        """Initialize interactive shell for paramiko client"""
        if self._use_shell:
            return
        try:
            self._shell = self._ssh_client.invoke_shell(width=250, height=50)
            self._shell.settimeout(60)
            time.sleep(1)
            # Clear initial banner
            if self._shell.recv_ready():
                self._shell.recv(65535)
        except Exception as e:
            print(f"{Colors.RED}Failed to create shell on {self.hostname}: {e}{Colors.ENDC}")
    
    def _run_command(self, command: str, timeout: int = 30) -> str:
        """Run a show command and return output using interactive shell"""
        try:
            if self._use_shell:
                # Use SSHConnection's send_command
                output = self._ssh_conn.send_command(command, timeout=timeout)
            else:
                # Use our own shell
                if not self._shell:
                    return ""
                
                self._shell.send(command + "\n")
                output = ""
                end_time = time.time() + timeout
                
                while time.time() < end_time:
                    if self._shell.recv_ready():
                        chunk = self._shell.recv(65535).decode('utf-8', errors='ignore')
                        output += chunk
                        # Check for prompt (ends with # or >)
                        if output.rstrip().endswith('#') or output.rstrip().endswith('>'):
                            break
                    else:
                        time.sleep(0.1)
            
            # Strip ANSI color codes from output (| include highlights matches)
            # ANSI codes: \x1b[...m or \033[...m
            ansi_escape = re.compile(r'\x1b\[[0-9;]*m|\033\[[0-9;]*m|\[\d+m')
            output = ansi_escape.sub('', output)
            
            return output
        except Exception as e:
            print(f"{Colors.RED}Command failed on {self.hostname}: {e}{Colors.ENDC}")
            return ""
    
    def get_bridge_domains(self, force_refresh: bool = False) -> List[BridgeDomain]:
        """
        Get all bridge domains from the DNAAS device.
        
        ENHANCED: Captures ALL bridge domains with various naming patterns:
        - Simple numeric: 210, 213, 500
        - Prefixed: v210, v213, BD-100, vlan-500
        - Complex: g_yor_v210, yor_210, customer_bd_100
        
        Two-step process:
        1. 'show network-services bridge-domain' -> list of BD names
        2. For each BD, 'show network-services bridge-domain <name>' -> details
        """
        if self._bd_cache is not None and not force_refresh:
            return self._bd_cache
        
        print(f"{Colors.CYAN}Querying ALL bridge domains on {self.hostname}...{Colors.ENDC}")
        
        # Step 1: Get list of BD names - ENHANCED to capture all patterns
        output = self._run_command("show network-services bridge-domain | no-more")
        
        bd_names = []
        seen_names = set()  # Avoid duplicates
        
        for line in output.split('\n'):
            line_stripped = line.strip()
            
            # Skip empty lines and CLI prompts (hostname# or hostname>)
            if not line_stripped:
                continue
            if re.match(r'^[\w\-\.]+[#>]\s*$', line_stripped):
                continue
            
            # Skip header lines (usually contain "Name", "State", etc.)
            if line_stripped.lower().startswith('name') or line_stripped.startswith('----'):
                continue
            
            # Parse table format: "| bd_name | state | ..." or "| bd_name    |"
            if '|' in line_stripped:
                parts = [p.strip() for p in line_stripped.split('|') if p.strip()]
                if parts:
                    bd_name = parts[0]
                    # Skip if it looks like a header row
                    if bd_name.lower() in ['name', 'bridge-domain', 'bridge domain', 'bd']:
                        continue
                    if bd_name and bd_name not in seen_names:
                        bd_names.append(bd_name)
                        seen_names.add(bd_name)
            else:
                # Non-table format: BD names might be listed directly
                # Match common BD naming patterns
                bd_match = re.match(r'^([\w\-_]+)(?:\s|$)', line_stripped)
                if bd_match:
                    bd_name = bd_match.group(1)
                    # Filter out common non-BD words
                    if bd_name.lower() not in ['name', 'state', 'type', 'admin', 'oper', 'associated', 'interfaces']:
                        if bd_name and bd_name not in seen_names:
                            bd_names.append(bd_name)
                            seen_names.add(bd_name)
        
        print(f"  Found {len(bd_names)} BD names: {', '.join(bd_names[:10])}{'...' if len(bd_names) > 10 else ''}")
        
        bridge_domains = []
        
        # Step 2: Get details for each BD
        for bd_name in bd_names:
            bd = self._get_bd_details(bd_name)
            if bd:
                bridge_domains.append(bd)
        
        self._bd_cache = bridge_domains
        print(f"{Colors.GREEN}Found {len(bridge_domains)} bridge domains on {self.hostname}{Colors.ENDC}")
        return bridge_domains
    
    def _get_bd_details(self, bd_name: str) -> Optional[BridgeDomain]:
        """
        Get details of a specific bridge domain by name.
        
        ENHANCED: Uses multiple commands to get comprehensive BD information:
        1. show network-services bridge-domain <bd> - General BD info
        2. show network-services bridge-domain <bd> interface - Interface details
        """
        # First get general BD info
        output = self._run_command(f"show network-services bridge-domain {bd_name} | no-more")
        
        bd = BridgeDomain(bd_id=0)  # We'll use name instead of numeric ID
        bd.name = bd_name
        
        # Try to extract VLAN from name (multiple patterns)
        # Patterns: g_yor_v210, v213, BD-100, vlan-500, 210
        vlan_match = re.search(r'[_\-]?v?(\d{2,4})$', bd_name, re.I)
        if vlan_match:
            bd.vlan_id = int(vlan_match.group(1))
        elif re.match(r'^\d+$', bd_name):  # Pure numeric BD name
            bd.vlan_id = int(bd_name)
        
        parsing_interfaces = False
        for line in output.split('\n'):
            line_stripped = line.strip()
            
            # Parse Admin State
            if 'Admin State:' in line or 'admin-state:' in line.lower():
                state_match = re.search(r'(?:Admin State|admin-state):\s*(\w+)', line, re.I)
                if state_match:
                    bd.state = state_match.group(1).lower()
            
            # Detect start of interface list
            if 'Associated Interfaces:' in line or 'Interfaces:' in line:
                parsing_interfaces = True
                continue
            
            # Parse interfaces (indented lines after "Associated Interfaces:")
            if parsing_interfaces and line_stripped:
                # Stop if we hit CLI prompt or another section header
                if re.match(r'^[\w\-\.]+[#>]\s*$', line_stripped) or line_stripped.startswith('Bridge-Domain'):
                    parsing_interfaces = False
                    continue
                
                # Extract interface name (may be on its own line or prefixed)
                interface = line_stripped
                _IF_PREFIXES = ('ge', 'hu', 'ce', 'qsfp', 'bundle', 'ae', 'eth',
                                'ten', 'hundred', 'forty', 'fo', 'lag', 'port-channel')
                if interface and (interface.startswith(_IF_PREFIXES) or '.' in interface):
                    if interface not in bd.interfaces:
                        bd.interfaces.append(interface)
                        
                        # Parse sub-interface VLAN if present
                        subif_match = re.match(r'([\w\-/]+)\.(\d+)', interface)
                        if subif_match:
                            bd.attachment_circuits.append({
                                'interface': subif_match.group(1),
                                'sub_interface': interface,
                                'vlan': int(subif_match.group(2))
                            })
        
        # ENHANCED: Also try dedicated interface command for more details
        if not bd.interfaces:
            iface_output = self._run_command(f"show network-services bridge-domain {bd_name} interface | no-more")
            for line in iface_output.split('\n'):
                line_stripped = line.strip()
                if not line_stripped or re.match(r'^[\w\-\.]+[#>]\s*$', line_stripped):
                    continue
                
                # Parse interface lines from table format
                if '|' in line_stripped:
                    parts = [p.strip() for p in line_stripped.split('|') if p.strip()]
                    if parts and not parts[0].lower().startswith('interface'):
                        iface = parts[0]
                        if iface and iface not in bd.interfaces:
                            bd.interfaces.append(iface)
                            # Parse sub-interface VLAN
                            subif_match = re.match(r'([\w\-/]+)\.(\d+)', iface)
                            if subif_match:
                                bd.attachment_circuits.append({
                                    'interface': subif_match.group(1),
                                    'sub_interface': iface,
                                    'vlan': int(subif_match.group(2))
                                })
                else:
                    # Non-table format
                    _IF_PREFIXES2 = ('ge', 'hu', 'ce', 'qsfp', 'bundle', 'ae', 'eth',
                                     'ten', 'hundred', 'forty', 'fo', 'lag', 'port-channel')
                    if line_stripped.startswith(_IF_PREFIXES2):
                        if line_stripped not in bd.interfaces:
                            bd.interfaces.append(line_stripped)
        
        if bd.interfaces:
            print(f"    BD {bd_name}: {len(bd.interfaces)} interfaces - {', '.join(bd.interfaces[:3])}{'...' if len(bd.interfaces) > 3 else ''}")
        else:
            print(f"    BD {bd_name}: no interfaces parsed (will try config search)")
        return bd
    
    def get_bd_by_id(self, bd_id: int) -> Optional[BridgeDomain]:
        """Get a specific bridge domain by ID"""
        bds = self.get_bridge_domains()
        for bd in bds:
            if bd.bd_id == bd_id:
                return bd
        return None
    
    def find_bds_for_interface(self, interface: str) -> List[BridgeDomain]:
        """
        Find bridge domains containing a specific interface.
        DEMAND-DRIVEN: Only queries BD details when needed, not all BDs.
        
        Handles:
        - Direct interface matches
        - Sub-interface matches  
        - Bundle member matches (if interface is a member of a bundle in a BD)
        """
        matching_bds = []
        base_interface = interface.split('.')[0]
        
        # Get list of BD names (quick - just table output)
        bd_names = self._get_bd_names()
        print(f"  Checking {len(bd_names)} BDs for interface {interface}...")
        
        # Query each BD's details until we find matching ones
        for bd_name in bd_names:
            bd = self._get_bd_details(bd_name)
            if not bd:
                continue
                
            # Check if our interface is in this BD
            for bd_if in bd.interfaces:
                # Direct or sub-interface match
                if bd_if == interface or bd_if.startswith(base_interface + '.') or interface.startswith(bd_if.split('.')[0] + '.'):
                    matching_bds.append(bd)
                    break
                
                # Check if interface is a bundle member
                bundle_base = bd_if.split('.')[0]
                if self.is_bundle_interface(bundle_base):
                    bundle = self.get_lacp_bundle(bundle_base)
                    if bundle and interface in bundle.members:
                        matching_bds.append(bd)
                        break
        
        return matching_bds
    
    def find_bd_from_config(self, interface: str) -> List[str]:
        """
        Find bridge domains containing an interface using config search.
        
        Uses the CORRECT command for DNAAS:
        show config network-services bridge-domain | include <interface> leading 10 trailing 10
        
        This is more reliable than parsing operational show commands because:
        1. It finds the EXACT BD that contains the interface
        2. Works even if the interface is in a bundle
        3. Shows context around the interface reference
        
        Config format on DNAAS:
            instance g_yor_v210
              interface bundle-60000.210
              !
              interface ge100-0/0/4
              !
            !
        
        BD Naming Convention: g_yor_v210_xxx
        - g = global
        - yor = username/customer
        - v210 = VLAN 210 (provider global VLAN)
        - xxx = optional suffix (WAN, etc.)
        
        Returns list of BD names found for this interface.
        """
        bd_names = []
        seen = set()
        
        # Strip sub-interface to search base interface too
        base_interface = interface.split('.')[0]
        interfaces_to_search = [interface]
        if interface != base_interface:
            interfaces_to_search.append(base_interface)
        
        for search_if in interfaces_to_search:
            print(f"  Searching config for BD containing {search_if}...")
            cmd = f"show config network-services bridge-domain | include {search_if} leading 10 trailing 10 | no-more"
            output = self._run_command(cmd, timeout=45)
            
            if not output or 'No match' in output:
                continue
            
            # Parse output to find BD name
            # DNAAS Config format:
            #   instance g_yor_v210        <-- BD name is after "instance"
            #     interface ge100-0/0/4    <-- interface reference
            #   !
            current_bd = None
            for line in output.split('\n'):
                line_stripped = line.strip()
                
                # Look for "instance <bd_name>" - this is the BD definition
                instance_match = re.match(r'instance\s+(\S+)', line_stripped)
                if instance_match:
                    current_bd = instance_match.group(1)
                    continue
                
                # If we found an interface line matching our search, and we have a current BD
                # Use word-boundary match to prevent ge100-0/0/3 matching ge100-0/0/30
                if current_bd and 'interface' in line_stripped and re.search(
                        r'(?:^|\s)interface\s+' + re.escape(search_if) + r'(?:\s|$|\.)', line_stripped):
                    # Skip mgmt BDs - they span entire fabric and cause infinite loops
                    if 'mgmt' in current_bd.lower():
                        print(f"    {Colors.YELLOW}Skipping mgmt BD: {current_bd}{Colors.ENDC}")
                        current_bd = None
                        continue
                        
                    if current_bd not in seen:
                        bd_names.append(current_bd)
                        seen.add(current_bd)
                        print(f"    {Colors.GREEN}Found BD: {current_bd} (contains {search_if}){Colors.ENDC}")
                        # Reset current_bd to find next instance if any
                        current_bd = None
        
        return bd_names
    
    def get_bd_details_by_name(self, bd_name: str) -> Optional[BridgeDomain]:
        """
        Get full details of a BD by name (wrapper for _get_bd_details).
        Public method for use by tracing logic.
        """
        return self._get_bd_details(bd_name)
    
    def _get_bd_names(self) -> List[str]:
        """Get list of BD names (fast - just parses table output)"""
        output = self._run_command("show network-services bridge-domain | no-more")
        
        bd_names = []
        for line in output.split('\n'):
            line = line.strip()
            if line.startswith('|'):
                parts = [p.strip() for p in line.split('|') if p.strip()]
                if parts and parts[0].lower() != 'name' and not parts[0].startswith('-'):
                    bd_names.append(parts[0])
        
        return bd_names
    
    def get_lacp_bundle(self, bundle_id: str, force_refresh: bool = False) -> Optional[LACPBundle]:
        """
        Get LACP bundle members.
        Parses 'show lacp interface bundle-<ID>' output.
        
        Output format:
        | Interface    | Role    | Port State   | Protocol State   | ...
        | ge100-0/0/36 | actor   | active       | ascd             | ...
        """
        # Strip sub-interface if present (bundle-60000.210 -> bundle-60000)
        bundle_base = bundle_id.split('.')[0]
        
        if bundle_base in self._lacp_cache and not force_refresh:
            return self._lacp_cache[bundle_base]
        
        # Normalize bundle ID
        if not bundle_base.startswith('bundle-'):
            bundle_base = f"bundle-{bundle_base}"
        
        print(f"{Colors.CYAN}Querying LACP bundle {bundle_base} on {self.hostname}...{Colors.ENDC}")
        output = self._run_command(f"show lacp interface {bundle_base} | no-more")
        
        bundle = LACPBundle(bundle_id=bundle_base)
        seen_interfaces = set()  # Avoid duplicates (actor and partner rows)
        
        for line in output.split('\n'):
            line_stripped = line.strip()
            
            # Parse table rows: | ge100-0/0/36 | actor | ...
            _LACP_PREFIXES = ('ge', 'hu', 'ce', 'eth', 'qsfp', 'ten', 'hundred', 'forty', 'fo')
            if line_stripped.startswith('|'):
                parts = [p.strip() for p in line_stripped.split('|') if p.strip()]
                if len(parts) >= 2:
                    interface = parts[0]
                    role = parts[1].lower() if len(parts) > 1 else ''
                    
                    if interface.startswith(_LACP_PREFIXES) and role == 'actor':
                        if interface not in seen_interfaces:
                            bundle.members.append(interface)
                            seen_interfaces.add(interface)
            
            # Match partner system ID from the header section
            partner_match = re.search(r'System-id:\s*([\w:]+)', line_stripped)
            if partner_match and 'Peer:' in output[:output.index(line) if line in output else len(output)]:
                bundle.partner_system_id = partner_match.group(1)
        
        self._lacp_cache[bundle_base] = bundle
        print(f"{Colors.GREEN}Bundle {bundle_base} has {len(bundle.members)} members{Colors.ENDC}")
        return bundle
    
    def resolve_bundle_members(self, bundle_id: str) -> List[str]:
        """Get list of member interfaces for a bundle"""
        bundle = self.get_lacp_bundle(bundle_id)
        return bundle.members if bundle else []
    
    def is_bundle_interface(self, interface: str) -> bool:
        """Check if an interface is a bundle"""
        return 'bundle' in interface.lower() or interface.lower().startswith('ae')
    
    def get_physical_for_bundle(self, bundle_interface: str) -> List[str]:
        """
        If interface is a bundle, return its member interfaces.
        Otherwise, return the interface itself.
        """
        if self.is_bundle_interface(bundle_interface):
            return self.resolve_bundle_members(bundle_interface)
        return [bundle_interface]
    
    def get_lldp_neighbors(self) -> List[Dict[str, str]]:
        """Get LLDP neighbors from the DNAAS device"""
        output = self._run_command("show lldp neighbors | no-more")
        
        neighbors = []
        _IFACE_PREFIXES = ('ge', 'hu', 'ce', 'qsfp', 'bundle', 'ae')
        in_table = False
        for line in output.split('\n'):
            if 'Interface' in line and 'Neighbor' in line:
                in_table = True
                continue
            if '---' in line or '|-' in line or '-|' in line:
                continue
            if re.match(r'^[A-Za-z0-9_.-]+[#>]', line.strip()):
                in_table = False
                continue

            # Pipe-delimited table
            match = re.match(
                r'\|\s*([\w\-/\.]+)\s*\|\s*(\S+)\s*\|\s*([\w\-/\.]+)\s*\|\s*(\d+)',
                line
            )
            if match:
                local_if, neighbor, remote_if, ttl = match.groups()
                if local_if.lower() != 'interface':
                    neighbors.append({
                        'local_interface': local_if,
                        'neighbor_name': neighbor,
                        'neighbor_interface': remote_if,
                        'ttl': int(ttl)
                    })
                continue

            # Pipe-delimited without TTL (some DNOS versions)
            if '|' in line:
                parts = [p.strip() for p in line.split('|') if p.strip()]
                if len(parts) >= 3 and parts[1] not in ('Neighbor', 'Neighbor System Name', '-', ''):
                    if not parts[0].lower().startswith('interface'):
                        neighbors.append({
                            'local_interface': parts[0],
                            'neighbor_name': parts[1],
                            'neighbor_interface': parts[2],
                            'ttl': int(parts[3]) if len(parts) > 3 and parts[3].isdigit() else 120
                        })
                continue

            # Space-aligned table fallback
            if in_table and line.strip():
                parts = re.split(r'\s{2,}', line.strip())
                if len(parts) >= 3 and parts[0].startswith(_IFACE_PREFIXES):
                    neighbors.append({
                        'local_interface': parts[0],
                        'neighbor_name': parts[1],
                        'neighbor_interface': parts[2],
                        'ttl': int(parts[3]) if len(parts) > 3 and parts[3].isdigit() else 120
                    })
                continue

            # Last resort: simple whitespace split
            parts = line.split()
            if len(parts) >= 3 and parts[0].startswith(_IFACE_PREFIXES):
                neighbors.append({
                    'local_interface': parts[0],
                    'neighbor_name': parts[1],
                    'neighbor_interface': parts[2]
                })
        
        if not neighbors and output.strip():
            print(f"    {Colors.YELLOW}[LLDP DEBUG] Parser returned 0 neighbors. Raw output ({len(output)} chars):{Colors.ENDC}")
            for dbg_line in output.split('\n')[:15]:
                print(f"      {dbg_line}")
        
        return neighbors
    
    def get_lldp_neighbors_detail(self) -> List[Dict[str, Any]]:
        """
        Get detailed LLDP neighbor information using 'show lldp neighbors detail'.
        
        Returns enhanced neighbor info including:
        - System capabilities
        - Management addresses
        - Port descriptions
        """
        output = self._run_command("show lldp neighbors detail | no-more")
        
        neighbors = []
        current_neighbor = {}
        
        for line in output.split('\n'):
            line_stripped = line.strip()
            
            # New neighbor section starts with "Interface:"
            if line_stripped.startswith('Interface:'):
                if current_neighbor:
                    neighbors.append(current_neighbor)
                current_neighbor = {
                    'local_interface': line_stripped.split(':', 1)[1].strip() if ':' in line_stripped else ''
                }
            elif current_neighbor:
                # Parse neighbor details
                if 'Neighbor:' in line_stripped or 'System Name:' in line_stripped:
                    val = line_stripped.split(':', 1)[1].strip() if ':' in line_stripped else ''
                    current_neighbor['neighbor_name'] = val
                elif 'Neighbor Interface:' in line_stripped or 'Port ID:' in line_stripped:
                    val = line_stripped.split(':', 1)[1].strip() if ':' in line_stripped else ''
                    current_neighbor['neighbor_interface'] = val
                elif 'System Capabilities:' in line_stripped:
                    val = line_stripped.split(':', 1)[1].strip() if ':' in line_stripped else ''
                    current_neighbor['capabilities'] = val
                elif 'Management Address:' in line_stripped or 'Mgmt Address:' in line_stripped:
                    val = line_stripped.split(':', 1)[1].strip() if ':' in line_stripped else ''
                    current_neighbor['mgmt_address'] = val
                elif 'Port Description:' in line_stripped:
                    val = line_stripped.split(':', 1)[1].strip() if ':' in line_stripped else ''
                    current_neighbor['port_description'] = val
        
        # Add last neighbor
        if current_neighbor:
            neighbors.append(current_neighbor)
        
        return neighbors
    
    def get_interface_state(self, interface: str) -> Dict[str, Any]:
        """
        Get operational state of an interface using 'show interface <if>'.
        
        Returns:
        - admin_state: enabled/disabled
        - oper_state: up/down
        - speed: interface speed
        - duplex: full/half
        - mtu: configured MTU
        """
        result = {
            'interface': interface,
            'admin_state': None,
            'oper_state': None,
            'speed': None,
            'duplex': None,
            'mtu': None
        }
        
        output = self._run_command(f"show interface {interface} | no-more")
        
        for line in output.split('\n'):
            line_lower = line.lower()
            
            if 'admin-state:' in line_lower or 'admin state:' in line_lower:
                match = re.search(r'(?:admin[- ]state):\s*(\w+)', line, re.I)
                if match:
                    result['admin_state'] = match.group(1).lower()
            elif 'oper-state:' in line_lower or 'oper state:' in line_lower:
                match = re.search(r'(?:oper[- ]state):\s*(\w+)', line, re.I)
                if match:
                    result['oper_state'] = match.group(1).lower()
            elif 'speed:' in line_lower:
                match = re.search(r'speed:\s*([\w\d]+)', line, re.I)
                if match:
                    result['speed'] = match.group(1)
            elif 'mtu:' in line_lower:
                match = re.search(r'mtu:\s*(\d+)', line, re.I)
                if match:
                    result['mtu'] = int(match.group(1))
        
        return result
    
    def match_bd_interface_to_lldp(self, bd_interface: str) -> Optional[Dict[str, str]]:
        """
        Match a BD interface to an LLDP neighbor.
        Handles bundle resolution if needed.
        """
        # Get physical interfaces (resolve bundle if needed)
        physical_interfaces = self.get_physical_for_bundle(bd_interface.split('.')[0])
        
        # Get LLDP neighbors
        neighbors = self.get_lldp_neighbors()
        
        # Match
        for neighbor in neighbors:
            if neighbor['local_interface'] in physical_interfaces:
                return neighbor
        
        return None
    
    def get_connected_dut_info(self, dut_lldp_interface: str) -> Optional[Dict[str, Any]]:
        """
        Given an LLDP interface that connects to a DUT, find:
        - All bridge domains on that interface
        - VLAN mappings
        - Bundle info if applicable
        """
        bds = self.find_bds_for_interface(dut_lldp_interface)
        
        if not bds:
            return None
        
        return {
            'interface': dut_lldp_interface,
            'bridge_domains': [
                {
                    'bd_id': bd.bd_id,
                    'name': bd.name,
                    'vlan_id': bd.vlan_id,
                    'state': bd.state,
                    'all_interfaces': bd.interfaces
                }
                for bd in bds
            ]
        }

    def get_interface_config_details(self, interface: str) -> Dict[str, Any]:
        """
        Fetch detailed configuration for an interface.
        Runs 'show configure interface <interface>' and parses VLAN/encapsulation info.
        
        Returns dict with:
        - encapsulation: dot1q config
        - vlan_id: outer VLAN
        - inner_vlan: inner VLAN (QinQ)
        - rewrite_ingress/egress: VLAN manipulation rules
        - l2_service: True if interface is L2
        - admin_state: up/down
        - raw_config: full config output
        """
        result = {
            'interface': interface,
            'encapsulation': None,
            'vlan_id': None,
            'inner_vlan': None,
            'rewrite_ingress': None,
            'rewrite_egress': None,
            'l2_service': False,
            'admin_state': None,
            'oper_state': None,
            'raw_config': None
        }
        
        try:
            # Get interface configuration
            config_output = self._run_command(f"show configure interface {interface} | no-more")
            result['raw_config'] = config_output
            
            # Parse encapsulation
            encap_match = re.search(r'encapsulation\s+dot1q\s+(\d+)(?:\s+second-dot1q\s+(\d+))?', config_output)
            if encap_match:
                result['encapsulation'] = encap_match.group(0)
                result['vlan_id'] = int(encap_match.group(1))
                if encap_match.group(2):
                    result['inner_vlan'] = int(encap_match.group(2))
            
            # Parse vlan-id (sub-interface style)
            vlan_match = re.search(r'vlan-id\s+(\d+)', config_output)
            if vlan_match and not result['vlan_id']:
                result['vlan_id'] = int(vlan_match.group(1))
            
            # Parse rewrite rules
            rewrite_in = re.search(r'rewrite\s+ingress\s+tag\s+(push|pop|translate)\s+[^\n]+', config_output)
            if rewrite_in:
                result['rewrite_ingress'] = rewrite_in.group(0)
            
            rewrite_out = re.search(r'rewrite\s+egress\s+tag\s+(push|pop|translate)\s+[^\n]+', config_output)
            if rewrite_out:
                result['rewrite_egress'] = rewrite_out.group(0)
            
            # Check for L2 service
            if 'l2-service' in config_output.lower():
                result['l2_service'] = True
            
            # Get interface state
            state_output = self._run_command(f"show interface {interface} | no-more")
            admin_match = re.search(r'admin[- ]?state[:\s]+(\w+)', state_output, re.IGNORECASE)
            if admin_match:
                result['admin_state'] = admin_match.group(1).lower()
            oper_match = re.search(r'oper[- ]?(?:state|status)[:\s]+(\w+)', state_output, re.IGNORECASE)
            if oper_match:
                result['oper_state'] = oper_match.group(1).lower()
                
        except Exception as e:
            print(f"{Colors.YELLOW}Warning: Could not get config for {interface}: {e}{Colors.ENDC}")
        
        return result
    
    def get_all_bd_interface_configs(self) -> Dict[str, Dict[str, Any]]:
        """
        Fetch configuration details for ALL interfaces across all Bridge Domains.
        Returns dict mapping interface name to config details.
        """
        all_configs = {}
        bds = self.get_bridge_domains()
        
        print(f"{Colors.CYAN}Fetching interface configs for {len(bds)} bridge domains...{Colors.ENDC}")
        
        for bd in bds:
            for iface in bd.interfaces:
                if iface not in all_configs:
                    print(f"    Getting config for {iface}...", end=' ', flush=True)
                    config = self.get_interface_config_details(iface)
                    all_configs[iface] = config
                    vlan_str = f"VLAN {config['vlan_id']}" if config['vlan_id'] else "no VLAN"
                    print(f"{Colors.GREEN}{vlan_str}{Colors.ENDC}")
        
        return all_configs
    
    def get_downlink_interface_configs(self, dut_interfaces: List[str]) -> Dict[str, Dict[str, Any]]:
        """
        Fetch configuration details for downlink interfaces towards DUTs.
        
        Args:
            dut_interfaces: List of interface names connected to DUTs (from LLDP)
            
        Returns dict mapping interface name to config details.
        """
        configs = {}
        
        print(f"{Colors.CYAN}Fetching configs for {len(dut_interfaces)} DUT downlinks...{Colors.ENDC}")
        
        for iface in dut_interfaces:
            # Get config for the physical interface
            base_iface = iface.split('.')[0]  # Remove sub-interface suffix
            if base_iface not in configs:
                print(f"    Getting config for {base_iface}...", end=' ', flush=True)
                config = self.get_interface_config_details(base_iface)
                configs[base_iface] = config
                state = config.get('admin_state', 'unknown')
                print(f"{Colors.GREEN}state: {state}{Colors.ENDC}")
            
            # Also get sub-interface config if present
            if '.' in iface and iface not in configs:
                print(f"    Getting config for {iface}...", end=' ', flush=True)
                config = self.get_interface_config_details(iface)
                configs[iface] = config
                vlan_str = f"VLAN {config['vlan_id']}" if config['vlan_id'] else "no VLAN"
                print(f"{Colors.GREEN}{vlan_str}{Colors.ENDC}")
        
        return configs

# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class ManagementInterface:
    """Management interface information"""
    name: str
    admin_state: str = ""
    oper_state: str = ""
    ipv4: str = ""
    ipv6: str = ""
    mtu: str = ""
    vrf: str = ""

@dataclass
class LLDPNeighbor:
    """LLDP neighbor information"""
    local_interface: str
    neighbor_name: str
    neighbor_interface: str
    ttl: int = 120
    is_dnaas: bool = False

@dataclass  
class InterfaceDetails:
    """Detailed interface information"""
    name: str
    description: str = ""
    vlan_id: Optional[int] = None
    bundle: str = ""
    sub_interfaces: List[str] = field(default_factory=list)
    admin_state: str = ""
    oper_state: str = ""

@dataclass
class DeviceInfo:
    """Device information cached in inventory"""
    serial: str
    hostname: str = ""
    mgmt_ip: str = ""
    mgmt_ipv6: str = ""
    system_type: str = ""
    dnos_version: str = ""
    last_seen: str = ""
    dnaas_interfaces: List[str] = field(default_factory=list)
    lldp_neighbors: List[Dict] = field(default_factory=list)
    all_mgmt_interfaces: List[Dict] = field(default_factory=list)
    interface_details: Dict[str, Dict] = field(default_factory=dict)  # Details per DNAAS interface

@dataclass
class PathHop:
    """A single hop in the DNAAS path"""
    device_name: str
    device_type: str  # PE, LEAF, SPINE
    device_serial: str = ""
    device_ip: str = ""
    ingress_interface: str = ""
    egress_interface: str = ""
    vlan_id: Optional[int] = None
    bundle: str = ""
    
@dataclass
class DNAASPath:
    """Complete DNAAS path between two endpoints"""
    source_device: str
    destination_device: str
    hops: List[PathHop] = field(default_factory=list)
    discovered_at: str = ""
    success: bool = False
    error_message: str = ""

# ============================================================================
# DNAAS TABLE LOADER
# ============================================================================

class DNAASTable:
    """Loads and queries the DNAAS device table from Excel"""
    
    def __init__(self, filepath: Path = DNAAS_TABLE_FILE):
        self.filepath = filepath
        self.devices: Dict[str, Dict] = {}
        self._load()
    
    def _load(self):
        """Load DNAAS table from Excel"""
        if not self.filepath.exists():
            print(f"{Colors.YELLOW}Warning: DNAAS table not found at {self.filepath}{Colors.ENDC}")
            return
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.filepath, read_only=True)
            ws = wb.active
            
            # Expected columns: Index, Rack, Name, Serial, IP Address, Status, TACACS, Accessible
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[2]:  # Name column
                    name = str(row[2]).strip()
                    # Normalize name (replace underscores with dashes for matching)
                    normalized_name = name.replace('_', '-').upper()
                    
                    self.devices[normalized_name] = {
                        'name': name,
                        'serial': str(row[3]) if row[3] else '',
                        'ip': str(row[4]) if row[4] else '',
                        'rack': str(row[1]) if row[1] else '',
                        'status': str(row[5]) if row[5] else '',
                        'accessible': str(row[7]).lower() == 'yes' if row[7] else False
                    }
            
            print(f"{Colors.GREEN}Loaded {len(self.devices)} DNAAS devices from table{Colors.ENDC}")
            wb.close()
            
        except Exception as e:
            print(f"{Colors.RED}Error loading DNAAS table: {e}{Colors.ENDC}")
    
    def lookup(self, name: str) -> Optional[Dict]:
        """Look up DNAAS device by name (fuzzy matching)"""
        # Normalize input name
        normalized = name.replace('_', '-').upper()
        
        # Direct match
        if normalized in self.devices:
            return self.devices[normalized]
        
        # Fuzzy match - find devices containing the search term
        for key, device in self.devices.items():
            if normalized in key or key in normalized:
                return device
            # Also try without DNAAS prefix
            short_name = normalized.replace('DNAAS-', '').replace('DNAAS_', '')
            if short_name in key:
                return device
        
        return None
    
    def get_ip(self, name: str) -> Optional[str]:
        """Get IP address for a DNAAS device"""
        device = self.lookup(name)
        return device['ip'] if device else None
    
    def list_all(self) -> List[Dict]:
        """List all DNAAS devices"""
        return list(self.devices.values())

# ============================================================================
# DEVICE INVENTORY DATABASE
# ============================================================================

class DeviceInventory:
    """Manages the device inventory database"""
    
    def __init__(self, filepath: Path = INVENTORY_FILE):
        self.filepath = filepath
        self.data = self._load()
    
    def _load(self) -> Dict:
        """Load inventory from file"""
        if self.filepath.exists():
            try:
                with open(self.filepath, 'r') as f:
                    return json.load(f)
            except (json.JSONDecodeError, IOError) as e:
                print(f"{Colors.YELLOW}Warning: Could not load inventory: {e}{Colors.ENDC}")
        return {"devices": {}, "last_updated": ""}
    
    def save(self):
        """Save inventory to file"""
        self.data["last_updated"] = datetime.now().isoformat()
        with open(self.filepath, 'w') as f:
            json.dump(self.data, f, indent=2, default=str)
        print(f"{Colors.GREEN}Inventory saved to {self.filepath}{Colors.ENDC}")
    
    def get_device(self, serial: str) -> Optional[DeviceInfo]:
        """Get device info by serial (case-insensitive, partial match)"""
        devices = self.data.get("devices", {})
        
        # Exact match first
        if serial in devices:
            return DeviceInfo(**devices[serial])
        
        # Case-insensitive match
        serial_lower = serial.lower()
        for key, d in devices.items():
            if key.lower() == serial_lower:
                return DeviceInfo(**d)
        
        # Partial match (serial contained in key or key in serial)
        for key, d in devices.items():
            dev_hostname = (d.get('hostname') or '').lower()
            if serial_lower in key.lower() or key.lower() in serial_lower or \
               serial_lower in dev_hostname or dev_hostname in serial_lower:
                return DeviceInfo(**d)
        
        return None
    
    def get_device_by_hostname(self, hostname: str) -> Optional[DeviceInfo]:
        """Get device info by hostname (with or without date suffix)"""
        if not hostname:
            return None
        
        # Normalize hostname - strip date suffix like (05-Jan-2026-18:48:06)
        clean_hostname = hostname.split('(')[0].strip().lower()
        
        for serial, device_data in self.data.get("devices", {}).items():
            device_hostname = (device_data.get("hostname") or "").split('(')[0].strip().lower()
            if device_hostname == clean_hostname or device_hostname == hostname.lower():
                return DeviceInfo(**device_data)
        return None
    
    def update_device(self, device: DeviceInfo):
        """Update or add device to inventory"""
        device.last_seen = datetime.now().isoformat()
        self.data["devices"][device.serial] = asdict(device)
        self.save()
    
    def list_devices(self) -> List[str]:
        """List all known device serials"""
        return list(self.data["devices"].keys())

# ============================================================================
# SSH CONNECTION MANAGER
# ============================================================================

class SSHConnection:
    """Manages SSH connections to devices"""
    
    def __init__(self, host: str, username: str = DEFAULT_USER, 
                 password: str = DEFAULT_PASS, timeout: int = 30):
        self.host = host
        self.username = username
        self.password = password
        self.timeout = timeout
        self.client: Optional[paramiko.SSHClient] = None
        self.shell = None
        self.connected = False
    
    def connect(self) -> bool:
        """Establish SSH connection"""
        try:
            print(f"  {Colors.CYAN}Connecting to {self.host}...{Colors.ENDC}", flush=True)
            self.client = paramiko.SSHClient()
            self.client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            self.client.connect(
                self.host, 
                username=self.username, 
                password=self.password, 
                timeout=self.timeout,
                look_for_keys=False,
                allow_agent=False
            )
            self.shell = self.client.invoke_shell(width=250, height=50)
            self.shell.settimeout(60)
            time.sleep(1)
            # Clear initial banner
            if self.shell.recv_ready():
                self.shell.recv(65535)
            self.connected = True
            print(f"  {Colors.GREEN}Connected!{Colors.ENDC}", flush=True)
            return True
        except Exception as e:
            print(f"  {Colors.RED}Connection failed: {e}{Colors.ENDC}", flush=True)
            return False
    
    def send_command(self, cmd: str, timeout: int = 30) -> str:
        """Send command and return output"""
        if not self.shell:
            return ""
        
        self.shell.send(cmd + "\n")
        output = ""
        end_time = time.time() + timeout
        
        while time.time() < end_time:
            if self.shell.recv_ready():
                chunk = self.shell.recv(65535).decode('utf-8', errors='ignore')
                output += chunk
                # Check for prompt (ends with # or >)
                if output.rstrip().endswith('#') or output.rstrip().endswith('>'):
                    break
            else:
                time.sleep(0.1)
        
        # Strip ANSI escape codes (| include command highlights matches)
        # Patterns: \x1b[91m (red), \x1b[0m (reset), [F, etc.
        ansi_pattern = re.compile(r'\x1b\[[0-9;]*[a-zA-Z]|\[F')
        output = ansi_pattern.sub('', output)
        
        return output
    
    def close(self):
        """Close SSH connection"""
        if self.client:
            self.client.close()
            self.connected = False

# ============================================================================
# DEVICE DISCOVERY
# ============================================================================

class DeviceDiscovery:
    """Discovers device information via SSH"""
    
    def __init__(self, ssh: SSHConnection):
        self.ssh = ssh
    
    def get_hostname(self) -> str:
        """Get device hostname from show system.
        Strips NCC standby date suffixes like (01-Mar-2026-15:06:20)."""
        output = self.ssh.send_command("show system | no-more", timeout=15)
        hostname = ""
        for line in output.split('\n'):
            if 'hostname' in line.lower() or 'system name' in line.lower():
                match = re.search(r':\s*(\S+)', line)
                if match:
                    hostname = match.group(1)
                    break
            match = re.match(r'^(\S+)[#>]', line.strip())
            if match and match.group(1) not in ['show', 'no-more']:
                hostname = match.group(1)
                break
        # Strip NCC standby date suffix: "YOR_CL_PE-4(01-Mar-2026-15:06:20)" -> "YOR_CL_PE-4"
        if hostname and re.search(r'\(\d{2}-\w{3}-\d{4}', hostname):
            clean = hostname.split('(')[0].strip()
            if clean:
                print(f"  {Colors.YELLOW}[Hostname] Stripped date suffix: {hostname} -> {clean}{Colors.ENDC}")
                hostname = clean
        return hostname
    
    def get_system_info(self) -> Dict[str, str]:
        """Get system type and version"""
        info = {"system_type": "", "dnos_version": ""}
        output = self.ssh.send_command("show system | no-more", timeout=15)
        
        for line in output.split('\n'):
            if 'system type' in line.lower() or 'platform' in line.lower():
                match = re.search(r':\s*(.+)', line)
                if match:
                    info["system_type"] = match.group(1).strip()
            if 'version' in line.lower() and 'dnos' in line.lower():
                match = re.search(r':\s*(.+)', line)
                if match:
                    info["dnos_version"] = match.group(1).strip()
        
        return info
    
    def get_management_interfaces(self) -> List[ManagementInterface]:
        """Get all management interfaces"""
        interfaces = []
        output = self.ssh.send_command("show interfaces management | no-more", timeout=30)
        
        # Parse table format
        # | Interface | Admin | Operational | IPv4 Address | IPv6 Address | MTU | VRF |
        for line in output.split('\n'):
            match = re.match(
                r'\|\s*([\w\-/\.]+)\s*\|\s*(\w+)\s*\|\s*(\S*)\s*\|\s*([^\|]*)\s*\|\s*([^\|]*)\s*\|\s*(\d*)\s*\|\s*(\S*)\s*\|',
                line
            )
            if match:
                name, admin, oper, ipv4, ipv6, mtu, vrf = match.groups()
                if name.lower() != 'interface':  # Skip header
                    interfaces.append(ManagementInterface(
                        name=name.strip(),
                        admin_state=admin.strip(),
                        oper_state=oper.strip(),
                        ipv4=ipv4.strip(),
                        ipv6=ipv6.strip(),
                        mtu=mtu.strip(),
                        vrf=vrf.strip()
                    ))
        
        return interfaces
    
    def get_primary_mgmt_ip(self, interfaces: List[ManagementInterface]) -> Tuple[str, str]:
        """Extract primary management IPv4 and IPv6"""
        ipv4, ipv6 = "", ""
        
        # Prefer mgmt0 or first interface with IP
        for iface in interfaces:
            if iface.name == 'mgmt0' and iface.ipv4:
                ipv4 = iface.ipv4.split('/')[0]
                ipv6 = iface.ipv6.split('/')[0] if iface.ipv6 else ""
                break
        
        # Fallback to any interface with IP
        if not ipv4:
            for iface in interfaces:
                if iface.ipv4:
                    ipv4 = iface.ipv4.split('/')[0]
                    ipv6 = iface.ipv6.split('/')[0] if iface.ipv6 else ""
                    break
        
        return ipv4, ipv6
    
    def get_lldp_neighbors(self) -> List[LLDPNeighbor]:
        """Get all LLDP neighbors"""
        neighbors = []
        output = self.ssh.send_command("show lldp neighbors | no-more", timeout=30)
        _IFACE_PREFIXES = ('ge', 'hu', 'ce', 'qsfp', 'bundle', 'ae')
        in_table = False

        for line in output.split('\n'):
            if 'Interface' in line and 'Neighbor' in line:
                in_table = True
                continue
            if '---' in line or '|-' in line or '-|' in line:
                continue
            if re.match(r'^[A-Za-z0-9_.-]+[#>]', line.strip()):
                in_table = False
                continue

            local_if = neighbor = remote_if = ''
            ttl = 120

            # Pipe-delimited table
            match = re.match(
                r'\|\s*([\w\-/\.]+)\s*\|\s*(\S+)\s*\|\s*([\w\-/\.]+)\s*\|\s*(\d+)',
                line
            )
            if match:
                local_if, neighbor, remote_if, ttl = match.group(1), match.group(2), match.group(3), int(match.group(4))
            elif '|' in line:
                parts = [p.strip() for p in line.split('|') if p.strip()]
                if len(parts) >= 3 and parts[1] not in ('Neighbor', 'Neighbor System Name', '-', ''):
                    if not parts[0].lower().startswith('interface'):
                        local_if, neighbor, remote_if = parts[0], parts[1], parts[2]
                        ttl = int(parts[3]) if len(parts) > 3 and parts[3].isdigit() else 120
            elif in_table and line.strip():
                parts = re.split(r'\s{2,}', line.strip())
                if len(parts) >= 3 and parts[0].startswith(_IFACE_PREFIXES):
                    local_if, neighbor, remote_if = parts[0], parts[1], parts[2]
                    ttl = int(parts[3]) if len(parts) > 3 and parts[3].isdigit() else 120
            else:
                parts = line.split()
                if len(parts) >= 3 and parts[0].startswith(_IFACE_PREFIXES):
                    local_if, neighbor, remote_if = parts[0], parts[1], parts[2]

            if local_if and neighbor and local_if.lower() != 'interface':
                is_dnaas = any(kw in neighbor.upper() for kw in DNAAS_KEYWORDS)
                neighbors.append(LLDPNeighbor(
                    local_interface=local_if,
                    neighbor_name=neighbor,
                    neighbor_interface=remote_if,
                    ttl=ttl,
                    is_dnaas=is_dnaas
                ))

        if not neighbors and output.strip():
            print(f"    {Colors.YELLOW}[LLDP DEBUG] DNAASDiscovery parser returned 0 neighbors. Raw output ({len(output)} chars):{Colors.ENDC}")
            for dbg_line in output.split('\n')[:15]:
                print(f"      {dbg_line}")

        return neighbors
    
    def get_interface_descriptions(self) -> Dict[str, str]:
        """Get interface descriptions from running config"""
        descriptions = {}
        output = self.ssh.send_command("show running-config interfaces | no-more", timeout=60)
        
        current_interface = None
        for line in output.split('\n'):
            # Match interface definitions
            iface_match = re.match(r'\s*(ge\d+-\d/\d/\d|bundle-ether\d+)', line)
            if iface_match:
                current_interface = iface_match.group(1)
            
            # Match descriptions
            if current_interface:
                desc_match = re.match(r'\s*description\s+"?([^"]+)"?', line)
                if desc_match:
                    descriptions[current_interface] = desc_match.group(1).strip()
        
        return descriptions
    
    def get_interface_details(self, interface_name: str) -> Dict[str, Any]:
        """Get detailed information about a specific interface including sub-interfaces"""
        details = {
            'interface': interface_name,
            'sub_interfaces': [],
            'speed': '',
            'bundle': '',
            'description': '',
            'encapsulation': ''
        }
        
        # Get main interface info
        output = self.ssh.send_command(f"show interfaces {interface_name} | no-more", timeout=30)
        
        for line in output.split('\n'):
            if 'Speed:' in line:
                match = re.search(r'Speed:\s*(\S+)', line)
                if match:
                    details['speed'] = match.group(1)
            if 'Bundle-id:' in line:
                match = re.search(r'Bundle-id:\s*(\S+)', line)
                if match and match.group(1) != 'N/A':
                    details['bundle'] = match.group(1)
            if 'Description:' in line:
                match = re.search(r'Description:\s*(.+)', line)
                if match:
                    details['description'] = match.group(1).strip()
            if 'Encapsulation:' in line:
                match = re.search(r'Encapsulation:\s*(\S+)', line)
                if match:
                    details['encapsulation'] = match.group(1)
        
        # Get sub-interfaces
        output2 = self.ssh.send_command(f"show interfaces | include {interface_name}. | no-more", timeout=30)
        
        # Remove ANSI color codes
        ansi_escape = re.compile(r'\x1b\[[0-9;]*m')
        clean_output = ansi_escape.sub('', output2)
        
        for line in clean_output.split('\n'):
            # Parse table format: | interface | admin | oper | ipv4 | ipv6 | vlan | mtu | service |
            # Example: | ge400-0/0/4.12 | enabled | up | 12.12.12.1/30 | | 12 | 9220 | VRF (default) |
            match = re.search(
                rf'\|\s*({re.escape(interface_name)}\.\d+)(?:\s*\(L2\))?\s*\|\s*(\w+)\s*\|\s*(\w+)\s*\|\s*([^\|]*)\|\s*([^\|]*)\|\s*(\d*)\s*\|\s*(\d*)\s*\|\s*([^\|]*)\|',
                line
            )
            if match:
                sub_iface, admin, oper, ipv4, ipv6, vlan, mtu, service = match.groups()
                details['sub_interfaces'].append({
                    'name': sub_iface.strip(),
                    'admin': admin.strip(),
                    'oper': oper.strip(),
                    'ipv4': ipv4.strip(),
                    'vlan': int(vlan) if vlan.strip() else None,
                    'mtu': int(mtu) if mtu.strip() else None,
                    'service': service.strip()
                })
        
        return details
    
    def detect_dnaas_interfaces(self, neighbors: List[LLDPNeighbor], 
                                 descriptions: Dict[str, str],
                                 config_text: str = "") -> List[str]:
        """Detect DNAAS-connected interfaces using all methods"""
        dnaas_interfaces = set()
        
        # Method 1: LLDP neighbor name contains DNAAS keywords
        for neighbor in neighbors:
            if neighbor.is_dnaas:
                dnaas_interfaces.add(neighbor.local_interface)
                print(f"    {Colors.GREEN}[LLDP]{Colors.ENDC} {neighbor.local_interface} -> {neighbor.neighbor_name}")
        
        # Method 2: Interface description contains DNAAS keywords
        for iface, desc in descriptions.items():
            if any(kw in desc.upper() for kw in DNAAS_KEYWORDS):
                dnaas_interfaces.add(iface)
                print(f"    {Colors.GREEN}[DESC]{Colors.ENDC} {iface}: {desc}")
        
        # Method 3: Config patterns (bundle-ether with DNAAS in description)
        if config_text:
            # Look for bundle-ether interfaces connected to DNAAS
            bundle_pattern = re.compile(
                r'(bundle-ether\d+).*?description\s+["\']?([^"\']+)["\']?',
                re.DOTALL
            )
            for match in bundle_pattern.finditer(config_text):
                bundle, desc = match.groups()
                if any(kw in desc.upper() for kw in DNAAS_KEYWORDS):
                    dnaas_interfaces.add(bundle)
                    print(f"    {Colors.GREEN}[CONFIG]{Colors.ENDC} {bundle}: {desc}")
        
        return list(dnaas_interfaces)

# ============================================================================
# PE/TERMINATION DEVICE INFO FETCHER
# ============================================================================

def fetch_pe_device_info(hostname_or_ip: str, user: str = DEFAULT_USER, 
                          password: str = DEFAULT_PASS, timeout: int = 20) -> Dict[str, str]:
    """
    Quickly connect to a PE/termination device and extract:
    - Serial number (from show system)
    - Management IP (from show interfaces management)
    
    Returns dict with 'serial', 'mgmt_ip', 'hostname' or empty values on failure.
    Used to auto-populate SSH credentials for discovered termination devices.
    """
    result = {
        'serial': '',
        'mgmt_ip': '',
        'hostname': '',
        'success': False
    }
    
    print(f"    {Colors.CYAN}Fetching info from PE: {hostname_or_ip}...{Colors.ENDC}", flush=True)
    
    try:
        ssh = SSHConnection(hostname_or_ip, user, password)
        if not ssh.connect():
            print(f"      {Colors.YELLOW}Could not connect to {hostname_or_ip}{Colors.ENDC}")
            return result
        
        try:
            discovery = DeviceDiscovery(ssh)
            
            # Get hostname
            hostname = discovery.get_hostname()
            if hostname:
                result['hostname'] = hostname
            
            # Get system info (contains serial number)
            output = ssh.send_command("show system | no-more", timeout=15)
            for line in output.split('\n'):
                line_lower = line.lower()
                if 'serial' in line_lower and 'number' in line_lower:
                    match = re.search(r':\s*(\S+)', line)
                    if match:
                        result['serial'] = match.group(1).strip()
                        break
            
            # Get management IP
            mgmt_interfaces = discovery.get_management_interfaces()
            mgmt_ip, _ = discovery.get_primary_mgmt_ip(mgmt_interfaces)
            if mgmt_ip:
                result['mgmt_ip'] = mgmt_ip
            
            result['success'] = True
            print(f"      {Colors.GREEN}Got: SN={result['serial'] or 'N/A'}, IP={result['mgmt_ip'] or 'N/A'}{Colors.ENDC}")
            
        finally:
            ssh.close()
            
    except Exception as e:
        print(f"      {Colors.RED}Error fetching PE info: {e}{Colors.ENDC}")
    
    return result


# ============================================================================
# PATH TRACING ENGINE
# ============================================================================

class PathTracer:
    """
    Traces paths through DNAAS fabric with Bridge Domain awareness.
    
    Enhanced capabilities:
    - Auto-configures DUT interfaces for LLDP discovery
    - Discovers Bridge Domains on DNAAS devices
    - Maps full DUT-to-DUT paths through BD
    - Resolves LACP bundles to physical interfaces
    """
    
    def __init__(self, inventory: DeviceInventory, dnaas_table: DNAASTable = None):
        self.inventory = inventory
        self.dnaas_table = dnaas_table or DNAASTable()
        self.visited_devices = set()
        self.max_depth = 15  # Limit depth to prevent overly long traces (increased for multi-BD discovery)
        self.found_target = False
        
        # Bridge Domain tracking
        self.discovered_bds: Dict[str, List[BridgeDomain]] = {}  # device -> BDs
        self.bd_path: List[Dict[str, Any]] = []  # BD path through fabric
        
        # LACP tracking  
        self.discovered_bundles: Dict[str, LACPBundle] = {}  # bundle_id -> bundle
        
        # Configuration changes made (for DUT devices only)
        self.config_changes: List[Dict[str, Any]] = []
        
        # Connection tracking for topology generation
        # Each entry: {"from": device_name, "to": device_name, "from_if": iface, "to_if": iface}
        self.discovered_connections: List[Dict[str, str]] = []
    
    def _get_lldp_from_dnaas(self, device_name: str, device_ip: str) -> List[LLDPNeighbor]:
        """SSH to DNAAS device and get LLDP neighbors"""
        neighbors = []
        
        print(f"    {Colors.CYAN}SSHing to {device_name} ({device_ip})...{Colors.ENDC}")
        
        ssh = SSHConnection(device_ip, DNAAS_USER, DNAAS_PASS)
        if not ssh.connect():
            print(f"    {Colors.RED}Failed to connect to {device_name}{Colors.ENDC}")
            return neighbors
        
        try:
            discovery = DeviceDiscovery(ssh)
            neighbors = discovery.get_lldp_neighbors()
            print(f"    {Colors.GREEN}Found {len(neighbors)} LLDP neighbors{Colors.ENDC}")
        except Exception as e:
            print(f"    {Colors.RED}Error getting LLDP: {e}{Colors.ENDC}")
        finally:
            ssh.close()
        
        return neighbors
    
    def _classify_device(self, name: str) -> str:
        """Classify device type based on name"""
        upper_name = name.upper()
        
        # Test equipment detection (Spirent, Ixia, etc.)
        if any(test in upper_name for test in ['SPIRENT', 'IXIA', 'KEYSIGHT', 'TESTER', 'TRAFFIC']):
            return "TEST_EQUIPMENT"
        
        # Server/Host detection
        if any(srv in upper_name for srv in ['SERVER', 'HOST', 'VM', 'WORKSTATION']):
            return "SERVER"
        
        # DNAAS fabric devices
        if 'SUPERSPINE' in upper_name:
            return "SUPERSPINE"
        elif 'SPINE' in upper_name:
            return "SPINE"
        elif 'LEAF' in upper_name:
            return "LEAF"
        elif 'DNAAS' in upper_name:
            return "DNAAS"
        else:
            return "PE"
    
    def _trace_through_dnaas(self, current_device: str, current_ip: str, 
                              target_devices: List[str], depth: int = 0) -> List[PathHop]:
        """Recursively trace through DNAAS fabric with optimized path finding"""
        hops = []
        
        if self.found_target:
            return hops
        
        print(f"  {Colors.CYAN}[TRACE depth={depth}] {current_device}{Colors.ENDC}")
        
        if depth >= self.max_depth:
            print(f"    {Colors.YELLOW}Max depth ({self.max_depth}) reached{Colors.ENDC}")
            return hops
        
        if current_device in self.visited_devices:
            return hops
        
        self.visited_devices.add(current_device)
        
        # Get LLDP neighbors from this DNAAS device
        neighbors = self._get_lldp_from_dnaas(current_device, current_ip)
        
        # First pass: Check if any neighbor is our target
        for neighbor in neighbors:
            neighbor_name = neighbor.neighbor_name
            for target in target_devices:
                if target.upper() in neighbor_name.upper() or neighbor_name.upper() in target.upper():
                    print(f"    {Colors.GREEN}*** FOUND TARGET: {neighbor_name} ***{Colors.ENDC}")
                    hops.append(PathHop(
                        device_name=neighbor_name,
                        device_type="PE",
                        ingress_interface=neighbor.neighbor_interface
                    ))
                    self.found_target = True
                    return hops
        
        # Second pass: Prioritize SPINEs and SuperSPINEs (they're more central)
        spine_neighbors = []
        leaf_neighbors = []
        
        for neighbor in neighbors:
            if neighbor.is_dnaas and neighbor.neighbor_name not in self.visited_devices:
                device_type = self._classify_device(neighbor.neighbor_name)
                if device_type in ['SPINE', 'SUPERSPINE']:
                    spine_neighbors.append((neighbor, device_type))
                elif device_type == 'LEAF':
                    leaf_neighbors.append((neighbor, device_type))
        
        # Trace through SPINEs first (more efficient path), limit to 2
        for neighbor, device_type in spine_neighbors[:2]:
            if self.found_target:
                return hops
                
            neighbor_name = neighbor.neighbor_name
            dnaas_info = self.dnaas_table.lookup(neighbor_name)
            
            if dnaas_info and dnaas_info.get('ip'):
                print(f"    {Colors.CYAN}-> {neighbor_name} ({device_type}){Colors.ENDC}")
                
                hop = PathHop(
                    device_name=neighbor_name,
                    device_type=device_type,
                    device_ip=dnaas_info['ip'],
                    ingress_interface=neighbor.neighbor_interface
                )
                hops.append(hop)
                
                deeper_hops = self._trace_through_dnaas(
                    neighbor_name, 
                    dnaas_info['ip'],
                    target_devices,
                    depth + 1
                )
                
                if deeper_hops:
                    hops.extend(deeper_hops)
                    return hops
                else:
                    hops.pop()  # Remove hop if it didn't lead to target
        
        # If no SPINEs or target not found, try LEAFs (limit to 2)
        for neighbor, device_type in leaf_neighbors[:2]:
            if self.found_target:
                return hops
                
            neighbor_name = neighbor.neighbor_name
            dnaas_info = self.dnaas_table.lookup(neighbor_name)
            
            if dnaas_info and dnaas_info.get('ip'):
                print(f"    {Colors.CYAN}-> {neighbor_name} ({device_type}){Colors.ENDC}")
                
                hop = PathHop(
                    device_name=neighbor_name,
                    device_type=device_type,
                    device_ip=dnaas_info['ip'],
                    ingress_interface=neighbor.neighbor_interface
                )
                hops.append(hop)
                
                deeper_hops = self._trace_through_dnaas(
                    neighbor_name, 
                    dnaas_info['ip'],
                    target_devices,
                    depth + 1
                )
                
                if deeper_hops:
                    hops.extend(deeper_hops)
                    return hops
                else:
                    hops.pop()  # Remove hop if it didn't lead to target
        
        return hops
    
    def trace_path(self, source_serial: str, dest_serial: str = "") -> DNAASPath:
        """Trace path between two devices through DNAAS"""
        path = DNAASPath(
            source_device=source_serial,
            destination_device=dest_serial,
            discovered_at=datetime.now().isoformat()
        )
        
        self.visited_devices.clear()
        self.found_target = False
        
        # Get source device info
        source_info = self.inventory.get_device(source_serial)
        if not source_info:
            path.error_message = f"Source device {source_serial} not in inventory"
            return path
        
        # Get destination device info if specified
        dest_info = None
        target_devices = []
        if dest_serial:
            dest_info = self.inventory.get_device(dest_serial)
            if dest_info:
                target_devices.append(dest_info.hostname or dest_serial)
        
        print(f"\n{Colors.HEADER}Tracing DNAAS path from {source_info.hostname or source_serial}...{Colors.ENDC}")
        
        # Add source as first hop
        path.hops.append(PathHop(
            device_name=source_info.hostname or source_serial,
            device_type="PE",
            device_serial=source_serial,
            device_ip=source_info.mgmt_ip
        ))
        
        # Find DNAAS neighbors from source
        dnaas_neighbors = [n for n in source_info.lldp_neighbors if n.get('is_dnaas', False)]
        
        if not dnaas_neighbors:
            path.error_message = "No DNAAS neighbors found on source device"
            return path
        
        # Trace through each DNAAS neighbor (typically LEAFs)
        for neighbor in dnaas_neighbors:
            leaf_name = neighbor.get('neighbor_name', '')
            local_if = neighbor.get('local_interface', '')
            remote_if = neighbor.get('neighbor_interface', '')
            
            print(f"  {Colors.CYAN}Found DNAAS LEAF: {leaf_name} via {local_if}{Colors.ENDC}")
            
            # Update source hop with egress interface
            path.hops[0].egress_interface = local_if
            
            # Look up LEAF IP from DNAAS table
            leaf_info = self.dnaas_table.lookup(leaf_name)
            leaf_ip = leaf_info['ip'] if leaf_info else None
            
            # Add LEAF hop
            leaf_hop = PathHop(
                device_name=leaf_name,
                device_type="LEAF",
                device_ip=leaf_ip or "",
                ingress_interface=remote_if
            )
            path.hops.append(leaf_hop)
            
            # If we have the LEAF IP, trace deeper into DNAAS
            if leaf_ip:
                # Don't add to visited here - let _trace_through_dnaas handle it
                deeper_hops = self._trace_through_dnaas(
                    leaf_name, 
                    leaf_ip, 
                    target_devices
                )
                
                if deeper_hops:
                    path.hops.extend(deeper_hops)
                    if self.found_target:
                        path.success = True
                    break
                else:
                    print(f"    {Colors.YELLOW}Target not found through {leaf_name}{Colors.ENDC}")
                    # Still add the explored path for visibility
                    if not target_devices:
                        path.success = True  # No specific target = exploration mode
        
        # If destination specified but path not complete, add destination from its LLDP
        if dest_info and not path.success:
            dest_dnaas = [n for n in dest_info.lldp_neighbors if n.get('is_dnaas', False)]
            if dest_dnaas:
                # Add the connecting LEAF if not already in path
                for neighbor in dest_dnaas:
                    leaf_name = neighbor.get('neighbor_name', '')
                    if not any(h.device_name == leaf_name for h in path.hops):
                        leaf_info = self.dnaas_table.lookup(leaf_name)
                        path.hops.append(PathHop(
                            device_name=leaf_name,
                            device_type="LEAF",
                            device_ip=leaf_info['ip'] if leaf_info else "",
                            egress_interface=neighbor.get('neighbor_interface', '')
                        ))
                
                # Add destination PE
                path.hops.append(PathHop(
                    device_name=dest_info.hostname or dest_serial,
                    device_type="PE",
                    device_serial=dest_serial,
                    device_ip=dest_info.mgmt_ip,
                    ingress_interface=dest_dnaas[0].get('local_interface', '')
                ))
                path.success = True
        
        if not path.success:
            if len(path.hops) > 1:
                path.success = True  # Partial success if we found at least one DNAAS device
                if dest_info and not dest_info.lldp_neighbors:
                    path.error_message = f"Partial path: {dest_info.hostname or dest_serial} has no LLDP neighbors (not connected to DNAAS?)"
            elif dest_info and not dest_info.lldp_neighbors:
                path.error_message = f"Target {dest_info.hostname or dest_serial} has 0 LLDP neighbors - not connected to DNAAS fabric"
        
        return path
    
    # =========================================================================
    # BRIDGE DOMAIN AWARE TRACING
    # =========================================================================
    
    def configure_dut_for_discovery(self, ssh_client: paramiko.SSHClient, 
                                    hostname: str) -> Tuple[bool, List[str]]:
        """
        Configure DUT device for DNAAS discovery.
        Enables interfaces and adds them to LLDP.
        
        Returns (success, list of configured interfaces)
        """
        print(f"\n{Colors.HEADER}Configuring {hostname} for discovery...{Colors.ENDC}")
        
        configurator = DUTConfigurator(ssh_client, hostname)
        success, interfaces = configurator.auto_configure_for_discovery()
        
        if interfaces:
            self.config_changes.append({
                'device': hostname,
                'timestamp': datetime.now().isoformat(),
                'interfaces': interfaces,
                'config_log': configurator.get_config_log()
            })
        
        return success, interfaces
    
    def discover_bridge_domains_on_dnaas(self, device_name: str, device_ip: str,
                                          dut_interface: str,
                                          target_bd_name: str = "") -> List[BridgeDomain]:
        """
        Connect to DNAAS device and find the Bridge Domain for the specified interface.
        
        DEMAND-DRIVEN approach:
        - If target_bd_name provided, only look for that specific BD
        - Otherwise, find all BDs for the interface (first hop only)
        
        Args:
            device_name: DNAAS device hostname
            device_ip: IP address to connect
            dut_interface: Interface connecting to DUT
            target_bd_name: Specific BD name to look for (speeds up traversal)
        
        Returns:
            List of BridgeDomain objects found
        """
        if target_bd_name:
            print(f"\n{Colors.HEADER}Looking for BD '{target_bd_name}' on {device_name}...{Colors.ENDC}")
        else:
            print(f"\n{Colors.HEADER}Finding BD for interface {dut_interface} on {device_name}...{Colors.ENDC}")
        
        ssh = SSHConnection(device_ip, DNAAS_USER, DNAAS_PASS)
        if not ssh.connect():
            print(f"{Colors.RED}Failed to connect to {device_name}{Colors.ENDC}")
            return []
        
        try:
            bd_discovery = BridgeDomainDiscovery(ssh, device_name)
            
            relevant_bds = []
            
            if target_bd_name:
                # FAST PATH: Direct lookup by BD name
                # Try exact match first, then partial match for cross-device names
                # e.g., g_yor_v210 on LEAF might be g_yor_v210_WAN on SPINE
                bd = bd_discovery._get_bd_details(target_bd_name)
                if bd:
                    relevant_bds = [bd]
                else:
                    # Try partial match (BD name might differ slightly between devices)
                    bd_names = bd_discovery._get_bd_names()
                    for name in bd_names:
                        # Match by base name or VLAN suffix
                        if target_bd_name.split('_')[0:2] == name.split('_')[0:2]:
                            bd = bd_discovery._get_bd_details(name)
                            if bd:
                                relevant_bds.append(bd)
                                break
            else:
                # FIRST HOP: Find BDs for this interface
                relevant_bds = bd_discovery.find_bds_for_interface(dut_interface)
            
            if not relevant_bds:
                print(f"{Colors.YELLOW}No BDs found{Colors.ENDC}")
                return []
            
            print(f"{Colors.GREEN}Found {len(relevant_bds)} BD(s){Colors.ENDC}")
            
            # For each relevant BD, get the other interfaces (these are the paths through DNAAS)
            for bd in relevant_bds:
                print(f"\n  {Colors.CYAN}BD: {bd.name or f'BD-{bd.bd_id}'}{Colors.ENDC}")
                print(f"    Interfaces: {', '.join(bd.interfaces)}")
                
                # Check for bundles and resolve members (needed for LLDP matching)
                for iface in bd.interfaces:
                    bundle_base = iface.split('.')[0]
                    if bd_discovery.is_bundle_interface(bundle_base):
                        bundle = bd_discovery.get_lacp_bundle(bundle_base)
                        if bundle and bundle.members:
                            self.discovered_bundles[bundle_base] = bundle
                            print(f"    Bundle {bundle_base} -> {', '.join(bundle.members)}")
                
                # Fetch interface configs for all BD interfaces
                interface_configs = {}
                print(f"    {Colors.CYAN}Fetching interface configs...{Colors.ENDC}")
                for iface in bd.interfaces:
                    config = bd_discovery.get_interface_config_details(iface)
                    interface_configs[iface] = config
                    vlan_str = f"VLAN {config['vlan_id']}" if config.get('vlan_id') else ""
                    rewrite_str = "rewrite" if config.get('rewrite_ingress') else ""
                    if vlan_str or rewrite_str:
                        print(f"      {iface}: {vlan_str} {rewrite_str}")
                
                # Add to BD path tracking with interface configs
                self.bd_path.append({
                    'device': device_name,
                    'bd_id': bd.bd_id,
                    'bd_name': bd.name,
                    'vlan_id': bd.vlan_id,
                    'interfaces': bd.interfaces,
                    'state': bd.state,
                    'interface_configs': interface_configs  # NEW: Store interface configs
                })
            
            return relevant_bds
            
        except Exception as e:
            print(f"{Colors.RED}Error discovering BDs on {device_name}: {e}{Colors.ENDC}")
            import traceback
            traceback.print_exc()
            return []
        finally:
            ssh.close()
    
    def trace_bd_path(self, source_serial: str, dest_serial: str = "",
                      auto_configure: bool = True) -> DNAASPath:
        """
        Trace path between devices with Bridge Domain awareness.
        
        This enhanced trace:
        1. Optionally configures DUT interfaces for LLDP
        2. Waits for LLDP to populate
        3. Discovers Bridge Domains on DNAAS devices
        4. Maps full path through BD topology
        
        Args:
            source_serial: Source device serial/hostname
            dest_serial: Destination device serial/hostname (optional)
            auto_configure: If True, auto-configure DUT interfaces
        
        Returns:
            DNAASPath with BD information
        """
        # Start with standard path trace
        path = DNAASPath(
            source_device=source_serial,
            destination_device=dest_serial,
            discovered_at=datetime.now().isoformat()
        )
        
        self.visited_devices.clear()
        self.found_target = False
        self.discovered_bds.clear()
        self.bd_path.clear()
        
        # Get source device info
        source_info = self.inventory.get_device(source_serial)
        if not source_info:
            path.error_message = f"Source device {source_serial} not in inventory"
            return path
        
        print(f"\n{Colors.HEADER}{'='*60}{Colors.ENDC}")
        print(f"{Colors.HEADER}BD-Aware Path Trace: {source_info.hostname or source_serial}{Colors.ENDC}")
        print(f"{Colors.HEADER}{'='*60}{Colors.ENDC}")
        
        # Step 1: Configure DUT if needed
        if auto_configure and source_info.mgmt_ip:
            # Get credentials
            creds = credential_manager.get(source_serial)
            user = creds.get('user', DEFAULT_USER) if creds else DEFAULT_USER
            passwd = creds.get('password', DEFAULT_PASS) if creds else DEFAULT_PASS
            
            ssh = SSHConnection(source_info.mgmt_ip, user, passwd)
            if ssh.connect():
                success, interfaces = self.configure_dut_for_discovery(
                    ssh.client, source_info.hostname
                )
                ssh.close()
                
                if interfaces:
                    print(f"\n{Colors.YELLOW}Waiting 30 seconds for LLDP to populate...{Colors.ENDC}")
                    time.sleep(30)
                    
                    # Re-discover LLDP neighbors after configuration
                    ssh = SSHConnection(source_info.mgmt_ip, user, passwd)
                    if ssh.connect():
                        discovery = DeviceDiscovery(ssh.client)
                        source_info.lldp_neighbors = [
                            asdict(n) for n in discovery.get_lldp_neighbors()
                        ]
                        ssh.close()
        
        # Step 2: Add source as first hop
        path.hops.append(PathHop(
            device_name=source_info.hostname or source_serial,
            device_type="PE",
            device_serial=source_serial,
            device_ip=source_info.mgmt_ip
        ))
        
        # Step 3: Find DNAAS neighbors and discover BDs
        dnaas_neighbors = [n for n in source_info.lldp_neighbors if n.get('is_dnaas', False)]
        
        if not dnaas_neighbors:
            path.error_message = "No DNAAS neighbors found. Try running with auto_configure=True"
            return path
        
        for neighbor in dnaas_neighbors:
            leaf_name = neighbor.get('neighbor_name', '')
            local_if = neighbor.get('local_interface', '')
            remote_if = neighbor.get('neighbor_interface', '')
            
            print(f"\n{Colors.CYAN}DNAAS connection: {local_if} -> {leaf_name}:{remote_if}{Colors.ENDC}")
            
            # Update source hop
            path.hops[0].egress_interface = local_if
            
            # Look up LEAF IP
            leaf_info = self.dnaas_table.lookup(leaf_name)
            leaf_ip = leaf_info['ip'] if leaf_info else None
            
            if leaf_ip:
                # Add LEAF hop
                leaf_hop = PathHop(
                    device_name=leaf_name,
                    device_type="LEAF",
                    device_ip=leaf_ip,
                    ingress_interface=remote_if
                )
                path.hops.append(leaf_hop)
                
                # Track this PE-to-LEAF connection for topology generation
                source_name = source_info.hostname or source_serial
                self.discovered_connections.append({
                    "from": source_name,
                    "to": leaf_name,
                    "from_if": local_if,
                    "to_if": remote_if
                })
                
                # Discover Bridge Domains on LEAF
                bds = self.discover_bridge_domains_on_dnaas(leaf_name, leaf_ip, remote_if)
                
                # Continue tracing through BDs
                if bds:
                    target_devices = []
                    
                    # If destination specified, look for it
                    if dest_serial:
                        target_info = self.inventory.get_device(dest_serial)
                        if target_info:
                            target_devices = [target_info.hostname or dest_serial]
                    
                    # Get the BD name for targeted lookup on subsequent devices
                    source_bd_name = bds[0].name if bds else ""
                    
                    # Use BD info to guide deeper tracing (exploration mode or targeted)
                    deeper_hops = self._trace_through_dnaas_with_bd(
                        leaf_name, leaf_ip, target_devices, bds,
                        source_bd_name=source_bd_name
                    )
                    if deeper_hops:
                        path.hops.extend(deeper_hops)
                        if self.found_target:
                            path.success = True
                    else:
                        path.success = True  # Exploration complete
                else:
                    path.success = True  # No BDs to trace
        
        return path
    
    def _trace_through_dnaas_with_bd(self, current_device: str, current_ip: str,
                                      target_devices: List[str], 
                                      current_bds: List[BridgeDomain],
                                      depth: int = 0,
                                      source_bd_name: str = "") -> List[PathHop]:
        """
        Trace through DNAAS fabric following a specific Bridge Domain.
        Uses BD interfaces to guide path rather than just LLDP.
        
        For transit nodes (like SuperSpine) that don't have the BD configured,
        falls back to LLDP-based tracing through ALL DNAAS neighbors.
        
        source_bd_name: The BD we're tracing (used for targeted lookup on next hops)
        """
        hops = []
        
        if self.found_target or depth >= self.max_depth:
            return hops
        
        if current_device in self.visited_devices:
            return hops
        
        self.visited_devices.add(current_device)
        
        device_type = self._classify_device(current_device)
        is_superspine = 'SUPERSPINE' in current_device.upper()
        
        print(f"\n  {Colors.CYAN}[BD-TRACE depth={depth}] {current_device} ({device_type}){Colors.ENDC}")
        
        # Get all interfaces from relevant BDs
        bd_interfaces = set()
        for bd in current_bds:
            bd_interfaces.update(bd.interfaces)
        
        # Determine if this is a transit node (no BD configured here)
        is_transit_node = len(bd_interfaces) == 0
        
        if is_transit_node:
            print(f"    {Colors.YELLOW}Transit node - no BD interfaces, using LLDP to continue{Colors.ENDC}")
        else:
            print(f"    BD interfaces: {', '.join(bd_interfaces)}")
        
        # Connect to get LLDP and match with BD interfaces
        ssh = SSHConnection(current_ip, DNAAS_USER, DNAAS_PASS)
        if not ssh.connect():
            return hops
        
        try:
            bd_discovery = BridgeDomainDiscovery(ssh, current_device)
            neighbors = bd_discovery.get_lldp_neighbors()
            
            print(f"    LLDP neighbors: {len(neighbors)} total")
            
            # Expand BD interfaces to include bundle members
            # e.g., if BD has bundle-60000.210, expand to include ge100-0/0/36, 37, 38
            expanded_bd_interfaces = set()
            for bd_if in bd_interfaces:
                expanded_bd_interfaces.add(bd_if)
                base_bd_if = bd_if.split('.')[0]  # bundle-60000.210 -> bundle-60000
                expanded_bd_interfaces.add(base_bd_if)
                
                # If it's a bundle, get its member interfaces
                if bd_discovery.is_bundle_interface(base_bd_if):
                    members = bd_discovery.resolve_bundle_members(base_bd_if)
                    expanded_bd_interfaces.update(members)
                    print(f"      Expanded {base_bd_if} -> {', '.join(members)}")
            
            # Find neighbors connected via BD interfaces (or ALL DNAAS neighbors for transit nodes)
            exploration_mode = len(target_devices) == 0
            found_dnaas_devices = []
            found_pe_devices = []
            
            # Check if current device is a LEAF (edge devices connect to LEAFs)
            is_leaf_node = 'LEAF' in current_device.upper()
            
            for neighbor in neighbors:
                local_if = neighbor.get('local_interface', '')
                remote_device = neighbor.get('neighbor_name', '')
                remote_if = neighbor.get('neighbor_interface', '')
                
                if not remote_device:
                    continue
                
                # Check if this is a DNAAS device
                is_dnaas_neighbor = any(kw in remote_device.upper() for kw in DNAAS_KEYWORDS)
                
                # For transit nodes (SuperSpine), follow ALL DNAAS neighbors
                # For nodes with BD, only follow neighbors on BD interfaces
                base_if = local_if.split('.')[0]
                interface_matches_bd = (local_if in expanded_bd_interfaces or 
                                       base_if in expanded_bd_interfaces)
                
                should_follow_dnaas = is_transit_node or interface_matches_bd
                
                # For LEAF nodes, discover ALL non-DNAAS neighbors as edge devices
                # This finds all PEs/edge devices connected to the fabric
                should_discover_edge = is_leaf_node and not is_dnaas_neighbor
                
                if should_follow_dnaas:
                    # Check if this is our target (only in targeted mode)
                    if not exploration_mode:
                        for target in target_devices:
                            if target.upper() in remote_device.upper():
                                print(f"    {Colors.GREEN}*** FOUND TARGET: {remote_device} ***{Colors.ENDC}")
                                hops.append(PathHop(
                                    device_name=remote_device,
                                    device_type="PE",
                                    ingress_interface=remote_if
                                ))
                                self.found_target = True
                                return hops
                    
                    # Check if this is another DNAAS device
                    if is_dnaas_neighbor:
                        if remote_device not in self.visited_devices:
                            found_dnaas_devices.append({
                                'name': remote_device,
                                'interface': remote_if,
                                'local_if': local_if
                            })
                
                # CRITICAL FIX: Only add edge devices that are on BD TERMINATION interfaces
                # This prevents listing all LLDP neighbors - only those actually in the BD
                if is_leaf_node and interface_matches_bd and not is_dnaas_neighbor:
                    if remote_device not in self.visited_devices:
                        if not any(d['name'] == remote_device for d in found_pe_devices):
                            print(f"    {Colors.BLUE}BD Edge device: {remote_device}{Colors.ENDC} via BD interface {local_if}")
                            found_pe_devices.append({
                                'name': remote_device,
                                'interface': remote_if,
                                'local_if': local_if
                            })
            
            # Deduplicate DNAAS devices (same device may appear on multiple bundle members)
            seen_devices = set()
            unique_dnaas_devices = []
            for dev in found_dnaas_devices:
                if dev['name'] not in seen_devices:
                    seen_devices.add(dev['name'])
                    unique_dnaas_devices.append(dev)
            
            # Also deduplicate PE devices
            seen_pe_devices = set()
            unique_pe_devices = []
            for dev in found_pe_devices:
                if dev['name'] not in seen_pe_devices:
                    seen_pe_devices.add(dev['name'])
                    unique_pe_devices.append(dev)
            
            print(f"    Found {len(unique_dnaas_devices)} DNAAS devices, {len(unique_pe_devices)} potential PEs")
            
            # Sort DNAAS devices: prioritize by type for better topology layout
            # SuperSpine -> SPINE -> LEAF order for going "up" then "down"
            def sort_key(x):
                name = x['name'].upper()
                if 'SUPERSPINE' in name:
                    return 0
                elif 'SPINE' in name:
                    return 1
                elif 'LEAF' in name:
                    return 2
                return 3
            
            # Process ALL discovered DNAAS devices (not just first one!)
            for dnaas_dev in sorted(unique_dnaas_devices, key=sort_key):
                remote_device = dnaas_dev['name']
                remote_if = dnaas_dev['interface']
                local_if = dnaas_dev['local_if']
                
                dnaas_info = self.dnaas_table.lookup(remote_device)
                if dnaas_info and dnaas_info.get('ip'):
                    device_type = self._classify_device(remote_device)
                    print(f"      -> {Colors.GREEN}{remote_device}{Colors.ENDC} ({device_type}) via {local_if}")
                    
                    # Track this connection for topology generation
                    self.discovered_connections.append({
                        "from": current_device,
                        "to": remote_device,
                        "from_if": local_if,
                        "to_if": remote_if
                    })
                    
                    hop = PathHop(
                        device_name=remote_device,
                        device_type=device_type,
                        device_ip=dnaas_info['ip'],
                        ingress_interface=remote_if
                    )
                    hops.append(hop)
                    
                    # Get BDs on next device - use source BD name for targeted lookup
                    next_bds = self.discover_bridge_domains_on_dnaas(
                        remote_device, dnaas_info['ip'], remote_if,
                        target_bd_name=source_bd_name
                    )
                    
                    # Continue tracing deeper
                    deeper = self._trace_through_dnaas_with_bd(
                        remote_device, dnaas_info['ip'],
                        target_devices, next_bds, depth + 1,
                        source_bd_name=source_bd_name
                    )
                    if deeper:
                        hops.extend(deeper)
                    
                    # In targeted mode, return after finding target
                    if self.found_target:
                        return hops
                    
                    # In exploration mode, KEEP all discovered hops (don't pop!)
                    # This allows us to build the full topology
                else:
                    print(f"      {Colors.YELLOW}No IP for {remote_device} in DNAAS table{Colors.ENDC}")
            
            # Also add discovered PE devices (these are endpoints!)
            # For each PE, try to fetch their SSH credentials (SN, mgmt IP)
            for pe_dev in unique_pe_devices:
                pe_name = pe_dev['name']
                local_if = pe_dev['local_if']
                remote_if = pe_dev['interface']
                print(f"      -> {Colors.BLUE}PE: {pe_name}{Colors.ENDC} via {local_if}")
                
                # Try to fetch PE device info for SSH credentials
                # The PE name from LLDP is often the hostname, which is resolvable
                pe_info = fetch_pe_device_info(pe_name)
                pe_serial = pe_info.get('serial', '')
                pe_mgmt_ip = pe_info.get('mgmt_ip', '')
                
                # Track this PE connection
                self.discovered_connections.append({
                    "from": current_device,
                    "to": pe_name,
                    "from_if": local_if,
                    "to_if": remote_if
                })
                
                hops.append(PathHop(
                    device_name=pe_name,
                    device_type="PE",
                    device_serial=pe_serial,
                    device_ip=pe_mgmt_ip,
                    ingress_interface=remote_if
                ))
        
        except Exception as e:
            print(f"    {Colors.RED}Error in BD trace: {e}{Colors.ENDC}")
            import traceback
            traceback.print_exc()
        finally:
            ssh.close()
        
        return hops
    
    def get_bd_summary(self) -> Dict[str, Any]:
        """Get summary of discovered Bridge Domains"""
        return {
            'devices_with_bds': list(self.discovered_bds.keys()),
            'total_bds_discovered': sum(len(bds) for bds in self.discovered_bds.values()),
            'bd_path': self.bd_path,
            'bundles_discovered': {k: asdict(v) for k, v in self.discovered_bundles.items()},
            'config_changes': self.config_changes
        }


# ============================================================================
# MULTI-BD PATH TRACER - Maps ALL Bridge Domains from a device
# ============================================================================

# Color palette for BD visualization (10 distinct colors)
BD_COLOR_PALETTE = [
    '#3498db',  # Blue
    '#e74c3c',  # Red
    '#2ecc71',  # Green
    '#9b59b6',  # Purple
    '#f39c12',  # Orange
    '#1abc9c',  # Teal
    '#e91e63',  # Pink
    '#00bcd4',  # Cyan
    '#ff5722',  # Deep Orange
    '#8bc34a'   # Light Green
]


def classify_bd_type(bd_name: str) -> Tuple[str, Optional[int]]:
    """
    Classify Bridge Domain type and extract global VLAN.
    
    Returns (type, global_vlan) where:
    - type: "global" if starts with "g_", "local" if starts with "l_", else "unknown"
    - global_vlan: extracted from "_vXXX" pattern (e.g., "_v210" -> 210)
    """
    bd_name_lower = bd_name.lower()
    
    # Determine BD type from prefix
    if bd_name_lower.startswith("g_"):
        bd_type = "global"
    elif bd_name_lower.startswith("l_"):
        bd_type = "local"
    else:
        bd_type = "unknown"
    
    # Extract VLAN from pattern like "_v210" or "v210"
    vlan_match = re.search(r'_v(\d+)', bd_name, re.IGNORECASE)
    if not vlan_match:
        vlan_match = re.search(r'v(\d+)$', bd_name, re.IGNORECASE)
    
    global_vlan = int(vlan_match.group(1)) if vlan_match else None
    
    return bd_type, global_vlan


def cleanup_old_discovery_files(output_dir: Path, keep_count: int = 3, prefix: str = "multi_bd_"):
    """
    Clean up old discovery files, keeping only the most recent N files.
    
    Args:
        output_dir: Directory containing discovery files
        keep_count: Number of most recent files to keep (default: 3)
        prefix: File prefix to match (default: "multi_bd_")
    """
    try:
        # Find all discovery files matching the pattern
        discovery_files = list(output_dir.glob(f"{prefix}*.json"))
        
        if len(discovery_files) <= keep_count:
            return  # Nothing to clean up
        
        # Sort by modification time (newest first)
        discovery_files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
        
        # Delete older files
        files_to_delete = discovery_files[keep_count:]
        for old_file in files_to_delete:
            try:
                old_file.unlink()
                print(f"  {Colors.DIM}Cleaned up old file: {old_file.name}{Colors.ENDC}")
            except Exception as e:
                print(f"  {Colors.YELLOW}Warning: Could not delete {old_file.name}: {e}{Colors.ENDC}")
        
        if files_to_delete:
            print(f"  {Colors.GREEN}Kept last {keep_count} discovery files{Colors.ENDC}")
    
    except Exception as e:
        print(f"  {Colors.YELLOW}Warning: Cleanup failed: {e}{Colors.ENDC}")


@dataclass
class BDPathInfo:
    """Information about a Bridge Domain's path through DNAAS fabric"""
    bd_name: str
    bd_type: str  # "global", "local", "unknown"
    global_vlan: Optional[int]
    color: str  # Assigned color from palette
    path: List[Dict[str, Any]] = field(default_factory=list)  # Devices in path
    interfaces: Dict[str, str] = field(default_factory=dict)  # device -> interface mapping
    dut_subinterface: str = ""  # Sub-interface on DUT side
    vlan_config: Dict[str, Any] = field(default_factory=dict)  # VLAN config details


class MultiBDPathTracer:
    """
    Traces paths for ALL Bridge Domains from a device through DNAAS fabric.
    
    Enhanced capabilities:
    - Discovers all BDs on connected DNAAS LEAF
    - Traces each BD's path through Spine/SuperSpine
    - Matches DUT sub-interfaces by VLAN configuration
    - Assigns unique colors for visualization
    """
    
    def __init__(self, inventory: DeviceInventory, dnaas_table: DNAASTable = None):
        self.inventory = inventory
        self.dnaas_table = dnaas_table or DNAASTable()
        self.visited_devices: Dict[str, set] = {}  # BD-specific visited tracking
        self.max_depth = 15
        
        # Results
        self.bd_paths: Dict[str, BDPathInfo] = {}  # bd_name -> BDPathInfo
        self.discovered_connections: List[Dict[str, Any]] = []
        self.discovered_bundles: Dict[str, LACPBundle] = {}
        
        # CONNECTION CACHING - Reuse SSH connections to avoid repeated reconnects
        self._ssh_cache: Dict[str, SSHConnection] = {}  # ip -> SSHConnection
        self._bd_discovery_cache: Dict[str, 'BridgeDomainDiscovery'] = {}  # device_name -> BridgeDomainDiscovery
        
        # TRACE CACHING - Skip BDs that have already been fully traced
        self._traced_bds: set = set()  # BD names that are fully traced
        
        # Thread safety for parallel LEAF discovery
        self._bd_paths_lock = threading.Lock()
    
    def _get_cached_ssh(self, ip: str, user: str, password: str) -> Optional[SSHConnection]:
        """Get or create cached SSH connection."""
        if ip in self._ssh_cache:
            ssh = self._ssh_cache[ip]
            # Check if connection is still alive
            if ssh.channel and not ssh.channel.closed:
                return ssh
            # Connection dead, remove from cache
            del self._ssh_cache[ip]
        
        # Create new connection
        ssh = SSHConnection(ip, user, password, timeout=60)
        if ssh.connect():
            self._ssh_cache[ip] = ssh
            return ssh
        return None
    
    def _close_all_connections(self):
        """Close all cached SSH connections."""
        for ip, ssh in self._ssh_cache.items():
            try:
                ssh.close()
            except:
                pass
        self._ssh_cache.clear()
        self._bd_discovery_cache.clear()
        
    def discover_all_bds(self, source_serial: str) -> Dict[str, BDPathInfo]:
        """
        Discover ALL Bridge Domains from a source device.
        
        Args:
            source_serial: Serial number or hostname of the source DUT
            
        Returns:
            Dict mapping BD name to BDPathInfo with full path details
        """
        print(f"\n{Colors.HEADER}{'='*70}{Colors.ENDC}")
        print(f"{Colors.HEADER}MULTI-BD DISCOVERY from {source_serial}{Colors.ENDC}")
        print(f"{Colors.HEADER}{'='*70}{Colors.ENDC}")
        
        # Get source device info
        source_info = self.inventory.get_device(source_serial)
        if not source_info:
            print(f"{Colors.RED}Source device {source_serial} not in inventory{Colors.ENDC}")
            return {}
        
        # Find DNAAS neighbors (LEAFs)
        dnaas_neighbors = [n for n in source_info.lldp_neighbors if n.get('is_dnaas', False)]
        
        if not dnaas_neighbors:
            print(f"{Colors.YELLOW}No DNAAS neighbors found on {source_serial} - attempting auto-configuration...{Colors.ENDC}")
            
            # Try to auto-configure the DUT
            configured = self._auto_configure_dut_for_discovery(source_info)
            
            if configured:
                # Wait for LLDP to discover neighbors
                import time
                print(f"{Colors.CYAN}Waiting 10 seconds for LLDP discovery...{Colors.ENDC}")
                time.sleep(10)
                
                # Re-discover the device to get updated LLDP neighbors
                print(f"{Colors.CYAN}Re-discovering device LLDP neighbors...{Colors.ENDC}")
                source_info = self._refresh_device_lldp(source_info)
                
                # Check again for DNAAS neighbors
                dnaas_neighbors = [n for n in source_info.lldp_neighbors if n.get('is_dnaas', False)]
                
                if not dnaas_neighbors:
                    print(f"{Colors.RED}Still no DNAAS neighbors after auto-configuration.{Colors.ENDC}")
                    print(f"{Colors.YELLOW}Please check physical connectivity to DNAAS fabric.{Colors.ENDC}")
                    return {}
            else:
                print(f"{Colors.RED}Auto-configuration failed. No DNAAS neighbors available.{Colors.ENDC}")
                return {}
        
        print(f"\n{Colors.CYAN}Found {len(dnaas_neighbors)} DNAAS neighbor(s){Colors.ENDC}")
        
        # Group ALL neighbors by LEAF name (collect every interface per LEAF)
        leafs_all = {}  # leaf_name -> list of (local_if, remote_if)
        for neighbor in dnaas_neighbors:
            leaf_name = neighbor.get('neighbor_name', '')
            local_if = neighbor.get('local_interface', '')
            remote_if = neighbor.get('neighbor_interface', '')
            if leaf_name not in leafs_all:
                leafs_all[leaf_name] = []
            leafs_all[leaf_name].append((local_if, remote_if))
        
        for lname, pairs in leafs_all.items():
            print(f"  {Colors.CYAN}LEAF {lname}: {len(pairs)} interfaces ({', '.join(p[1] for p in pairs)}){Colors.ENDC}")
        
        # Resolve LEAF IPs and build work items (with ALL interfaces)
        leaf_work = []
        for leaf_name, if_pairs in leafs_all.items():
            leaf_info = self.dnaas_table.lookup(leaf_name)
            if not leaf_info or not leaf_info.get('ip'):
                print(f"  {Colors.YELLOW}No IP found for {leaf_name} in DNAAS table{Colors.ENDC}")
                continue
            dut_interfaces = [p[0] for p in if_pairs]
            leaf_interfaces = [p[1] for p in if_pairs]
            leaf_work.append((leaf_name, leaf_info['ip'], dut_interfaces, leaf_interfaces))
        
        # Process LEAFs in PARALLEL (up to 4 at once) for faster discovery
        if len(leaf_work) > 1:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            max_workers = min(4, len(leaf_work))
            print(f"\n{Colors.CYAN}Parallel discovery: {len(leaf_work)} LEAFs with {max_workers} workers{Colors.ENDC}")
            
            def _process_leaf(args):
                lname, lip, dut_ifs, leaf_ifs = args
                print(f"\n{Colors.BLUE}Processing LEAF: {lname} ({len(dut_ifs)} DUT interfaces){Colors.ENDC}")
                try:
                    self._discover_bds_on_leaf(source_info, lname, lip, dut_ifs, leaf_ifs)
                except Exception as e:
                    print(f"  {Colors.RED}LEAF {lname} failed: {e}{Colors.ENDC}")
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {executor.submit(_process_leaf, w): w[0] for w in leaf_work}
                for future in as_completed(futures):
                    leaf_name = futures[future]
                    try:
                        future.result()
                    except Exception as e:
                        print(f"  {Colors.RED}LEAF {leaf_name} error: {e}{Colors.ENDC}")
        else:
            for leaf_name, leaf_ip, dut_ifs, leaf_ifs in leaf_work:
                print(f"\n{Colors.BLUE}Processing LEAF: {leaf_name} ({len(dut_ifs)} DUT interfaces){Colors.ENDC}")
                self._discover_bds_on_leaf(source_info, leaf_name, leaf_ip, dut_ifs, leaf_ifs)
        
        # Assign colors to discovered BDs
        # RULE: mgmt BDs are ALWAYS RED, others get distinct colors
        # Color palette without red for non-mgmt BDs
        non_red_colors = ['#3498db', '#2ecc71', '#9b59b6', '#f39c12', '#1abc9c', '#e67e22', '#34495e', '#16a085']
        non_mgmt_idx = 0
        for bd_name, bd_info in self.bd_paths.items():
            if 'mgmt' in bd_name.lower():
                bd_info.color = '#e74c3c'  # RED for mgmt
            else:
                bd_info.color = non_red_colors[non_mgmt_idx % len(non_red_colors)]
                non_mgmt_idx += 1
        
        print(f"\n{Colors.GREEN}Discovered {len(self.bd_paths)} Bridge Domains{Colors.ENDC}")
        for bd_name, bd_info in self.bd_paths.items():
            print(f"  • {bd_name}: {bd_info.bd_type}, VLAN {bd_info.global_vlan or 'N/A'}")
        
        return self.bd_paths
    
    def _auto_configure_dut_for_discovery(self, source_info: DeviceInfo) -> bool:
        """
        Auto-configure the DUT device for DNAAS discovery:
        1. Enable all disabled physical interfaces
        2. Add all physical interfaces to LLDP
        3. Commit changes
        
        Returns True if configuration was successful.
        """
        # Get connection info (DeviceInfo uses 'serial' not 'device_serial')
        hostname = source_info.hostname or source_info.serial
        mgmt_ip = source_info.mgmt_ip or source_info.serial
        
        print(f"\n{Colors.HEADER}Auto-configuring {hostname} for DNAAS discovery...{Colors.ENDC}")
        
        # Try to get stored credentials
        creds = credential_manager.get(source_info.serial)
        user = creds.get('user', DEFAULT_USER) if creds else DEFAULT_USER
        password = creds.get('password', DEFAULT_PASS) if creds else DEFAULT_PASS
        
        ssh = SSHConnection(mgmt_ip, user, password)
        if not ssh.connect():
            print(f"  {Colors.RED}Failed to connect to {hostname} for configuration{Colors.ENDC}")
            return False
        
        try:
            configurator = DUTConfigurator(ssh, hostname)
            success, interfaces = configurator.auto_configure_for_discovery()
            
            if interfaces:
                print(f"  {Colors.GREEN}Configured {len(interfaces)} interfaces for LLDP discovery{Colors.ENDC}")
                for iface in interfaces[:10]:
                    print(f"    • {iface}")
                if len(interfaces) > 10:
                    print(f"    ... and {len(interfaces) - 10} more")
            else:
                print(f"  {Colors.YELLOW}No interfaces needed configuration{Colors.ENDC}")
            
            return success
            
        except Exception as e:
            print(f"  {Colors.RED}Auto-configuration failed: {e}{Colors.ENDC}")
            import traceback
            traceback.print_exc()
            return False
        finally:
            ssh.close()
    
    def _refresh_device_lldp(self, source_info: DeviceInfo) -> DeviceInfo:
        """
        Re-fetch LLDP neighbors for a device after auto-configuration.
        Updates the device info with new LLDP data.
        """
        hostname = source_info.hostname or source_info.serial
        mgmt_ip = source_info.mgmt_ip or source_info.serial
        
        # Try to get stored credentials
        creds = credential_manager.get(source_info.serial)
        user = creds.get('user', DEFAULT_USER) if creds else DEFAULT_USER
        password = creds.get('password', DEFAULT_PASS) if creds else DEFAULT_PASS
        
        ssh = SSHConnection(mgmt_ip, user, password)
        if not ssh.connect():
            print(f"  {Colors.RED}Failed to reconnect to {hostname}{Colors.ENDC}")
            return source_info
        
        try:
            discovery = DeviceDiscovery(ssh)
            new_neighbors = discovery.get_lldp_neighbors()
            
            print(f"  {Colors.GREEN}Found {len(new_neighbors)} LLDP neighbors after configuration{Colors.ENDC}")
            
            # Update source_info with new LLDP data
            source_info.lldp_neighbors = []
            for n in new_neighbors:
                neighbor_dict = {
                    'local_interface': n.local_interface,
                    'neighbor_name': n.neighbor_name,
                    'neighbor_interface': n.neighbor_interface,
                    'ttl': n.ttl,
                    'is_dnaas': any(kw in n.neighbor_name.upper() for kw in DNAAS_KEYWORDS)
                }
                source_info.lldp_neighbors.append(neighbor_dict)
                
                if neighbor_dict['is_dnaas']:
                    print(f"    {Colors.CYAN}DNAAS: {n.local_interface} -> {n.neighbor_name}{Colors.ENDC}")
                else:
                    print(f"    {n.local_interface} -> {n.neighbor_name}")
            
            # Save updated device info
            self.inventory.save_device(source_info)
            
            return source_info
            
        except Exception as e:
            print(f"  {Colors.RED}Failed to refresh LLDP: {e}{Colors.ENDC}")
            return source_info
        finally:
            ssh.close()
    
    def _discover_bds_on_leaf(self, source_info: DeviceInfo, leaf_name: str, 
                              leaf_ip: str, dut_interfaces: list, leaf_interfaces: list):
        """
        Connect to LEAF and discover all Bridge Domains relevant to the DUT.
        
        Args:
            dut_interfaces: ALL DUT-side interface names for this LEAF (from inventory LLDP)
            leaf_interfaces: ALL LEAF-side interface names for this LEAF (from inventory LLDP)
        """
        ssh = SSHConnection(leaf_ip, DNAAS_USER, DNAAS_PASS)
        if not ssh.connect():
            print(f"  {Colors.RED}Failed to connect to {leaf_name}{Colors.ENDC}")
            return
        
        try:
            bd_discovery = BridgeDomainDiscovery(ssh, leaf_name)
            
            # Use LEAF-side interfaces from inventory (already known from LLDP)
            dut_facing_interfaces = list(leaf_interfaces)
            print(f"  {Colors.GREEN}DUT-facing interfaces (from inventory): {', '.join(dut_facing_interfaces)}{Colors.ENDC}")
            
            # Also try LLDP on the LEAF to catch any additional interfaces
            lldp_neighbors = bd_discovery.get_lldp_neighbors()
            dut_hostname = source_info.hostname or source_info.serial
            for n in lldp_neighbors:
                neighbor_name = n.get('neighbor_name', '')
                local_if = n.get('local_interface', '')
                name_lower = neighbor_name.lower()
                host_lower = dut_hostname.lower()
                # Match by substring or by serial
                if (host_lower in name_lower or name_lower in host_lower or
                        source_info.serial.lower() in name_lower):
                    if local_if not in dut_facing_interfaces:
                        dut_facing_interfaces.append(local_if)
                        print(f"  {Colors.CYAN}[LLDP] Additional DUT interface: {local_if} -> {neighbor_name}{Colors.ENDC}")
            
            print(f"  {Colors.GREEN}Total DUT-facing interfaces: {len(dut_facing_interfaces)}{Colors.ENDC}")
            
            # Resolve bundles that contain DUT-facing interfaces so we can
            # match BDs that reference bundles instead of physical interfaces
            dut_bundles = set()
            for dut_if in dut_facing_interfaces:
                base_if = dut_if.split('.')[0]
                # Search all bundles for this member
                bundle_output = bd_discovery._run_command(f"show lacp | include {base_if} | no-more")
                for line in bundle_output.split('\n'):
                    bm = re.search(r'(bundle-\d+)', line)
                    if bm:
                        dut_bundles.add(bm.group(1))
            
            if dut_bundles:
                print(f"  {Colors.CYAN}DUT-facing bundles: {', '.join(dut_bundles)}{Colors.ENDC}")
            
            # Get ALL Bridge Domains from the LEAF
            all_bds = bd_discovery.get_bridge_domains()
            print(f"  {Colors.GREEN}Found {len(all_bds)} Bridge Domains on {leaf_name}{Colors.ENDC}")
            
            # Also get BD names directly from config (catches BDs that _get_bd_details
            # might have dropped due to interface parsing issues)
            all_bd_names_from_config = set(bd.name for bd in all_bds)
            
            # STEP A: Match BDs by interface (primary method)
            relevant_bds = []
            
            for bd in all_bds:
                if 'mgmt' in bd.name.lower():
                    continue
                
                for dut_if in dut_facing_interfaces:
                    if self._bd_contains_interface(bd, dut_if, bd_discovery):
                        if bd not in relevant_bds:
                            relevant_bds.append(bd)
                            print(f"    Found (interface match): {bd.name} via {dut_if}")
                        break
                else:
                    # Also check if BD contains any of the DUT-facing bundles
                    for bd_iface in bd.interfaces:
                        bd_base = bd_iface.split('.')[0]
                        if bd_base in dut_bundles:
                            if bd not in relevant_bds:
                                relevant_bds.append(bd)
                                print(f"    Found (bundle match): {bd.name} via {bd_base}")
                            break
            
            # STEP B: Config-based search (FALLBACK only when STEP A found nothing)
            if not relevant_bds:
                print(f"  {Colors.YELLOW}No BDs found via interface/bundle match, searching config...{Colors.ENDC}")
                search_terms = list(dut_facing_interfaces) + list(dut_bundles)
                config_found_names = set()
                
                for search_term in search_terms:
                    config_bd_names = bd_discovery.find_bd_from_config(search_term)
                    for bd_name in config_bd_names:
                        if 'mgmt' in bd_name.lower():
                            continue
                        config_found_names.add(bd_name)
                
                existing_names = set(bd.name for bd in relevant_bds)
                for bd_name in config_found_names:
                    if bd_name not in existing_names:
                        bd = bd_discovery._get_bd_details(bd_name)
                        if not bd:
                            bd = BridgeDomain(bd_id=0)
                            bd.name = bd_name
                        relevant_bds.append(bd)
                        print(f"    Found (config match): {bd.name}")
            
            print(f"  {Colors.CYAN}Relevant BDs for DUT: {len(relevant_bds)}{Colors.ENDC}")
            
            # Build LEAF->DUT interface pair map (for port-mode and sub-interface resolution)
            leaf_to_dut_map = {}
            for i, l_if in enumerate(leaf_interfaces):
                if i < len(dut_interfaces):
                    leaf_to_dut_map[l_if] = dut_interfaces[i]
            
            # Process each relevant BD
            for bd in relevant_bds:
                with self._bd_paths_lock:
                    already_traced = bd.name in self.bd_paths
                if already_traced:
                    print(f"\n  {Colors.CYAN}[BD] {bd.name} - already traced, skipping{Colors.ENDC}")
                    continue
                
                bd_type, global_vlan = classify_bd_type(bd.name)
                
                print(f"\n  {Colors.HEADER}[BD] {bd.name} ({bd_type}, VLAN {global_vlan or 'N/A'}){Colors.ENDC}")
                
                bd_path_info = BDPathInfo(
                    bd_name=bd.name,
                    bd_type=bd_type,
                    global_vlan=global_vlan,
                    color=""
                )
                
                # Find which LEAF interface carries this BD (needed before DUT lookup)
                leaf_if_for_bd = leaf_interfaces[0] if leaf_interfaces else ''
                bd_uses_bundle = None
                for l_if in leaf_interfaces:
                    if self._bd_contains_interface(bd, l_if, bd_discovery):
                        leaf_if_for_bd = l_if
                        break
                else:
                    # Check bundles
                    for b in dut_bundles:
                        for bd_iface in bd.interfaces:
                            if bd_iface.split('.')[0] == b:
                                leaf_if_for_bd = bd_iface
                                bd_uses_bundle = b
                                break
                
                # Determine the correct DUT interface for this BD
                primary_dut_if = dut_interfaces[0] if dut_interfaces else ''
                
                if bd_uses_bundle:
                    # BD uses a bundle - find which DUT interfaces are members
                    members = bd_discovery.resolve_bundle_members(bd_uses_bundle)
                    for member in members:
                        if member in leaf_to_dut_map:
                            primary_dut_if = leaf_to_dut_map[member]
                            break
                elif leaf_if_for_bd in leaf_to_dut_map:
                    primary_dut_if = leaf_to_dut_map[leaf_if_for_bd]
                
                # Find DUT sub-interface
                dut_subif = None
                if global_vlan:
                    dut_subif = self._find_dut_subinterface(source_info, primary_dut_if, global_vlan, bd.name)
                
                bd_path_info.dut_subinterface = dut_subif
                
                dut_if_with_vlan = dut_subif
                if not dut_if_with_vlan and global_vlan:
                    dut_if_with_vlan = f"{primary_dut_if}.{global_vlan}"
                elif not dut_if_with_vlan:
                    # Port-mode: use the bare DUT interface
                    dut_if_with_vlan = primary_dut_if
                
                bd_path_info.path.append({
                    'device': source_info.hostname or source_info.serial,
                    'device_type': 'PE',
                    'interface': dut_if_with_vlan,
                    'bd_interface': None
                })
                
                bd_path_info.path.append({
                    'device': leaf_name,
                    'device_type': 'LEAF',
                    'interface': leaf_if_for_bd,
                    'bd_interface': self._find_bd_interface(bd, leaf_if_for_bd, bd_discovery)
                })
                
                bd_path_info.interfaces[source_info.hostname or source_info.serial] = dut_if_with_vlan
                bd_path_info.interfaces[leaf_name] = leaf_if_for_bd
                
                self._trace_bd_path(bd, bd_path_info, leaf_name, leaf_ip, bd_discovery, lldp_neighbors)
                
                with self._bd_paths_lock:
                    self.bd_paths[bd.name] = bd_path_info
                
                from_interface = dut_subif
                if not from_interface and global_vlan:
                    from_interface = f"{primary_dut_if}.{global_vlan}"
                elif not from_interface:
                    from_interface = primary_dut_if
                    
                self.discovered_connections.append({
                    'from': source_info.hostname or source_info.serial,
                    'to': leaf_name,
                    'from_if': from_interface,
                    'to_if': leaf_if_for_bd,
                    'bd_name': bd.name,
                    'bd_type': bd_type,
                    'global_vlan': global_vlan
                })
            
        except Exception as e:
            print(f"  {Colors.RED}Error discovering BDs on {leaf_name}: {e}{Colors.ENDC}")
            import traceback
            traceback.print_exc()
        finally:
            ssh.close()
    
    def _bd_contains_interface(self, bd: BridgeDomain, interface: str, 
                               bd_discovery: BridgeDomainDiscovery) -> bool:
        """Check if a BD contains the specified interface or its bundle."""
        base_interface = interface.split('.')[0]
        
        for bd_iface in bd.interfaces:
            bd_base = bd_iface.split('.')[0]
            
            # Direct match
            if bd_base == base_interface or bd_iface == interface:
                return True
            
            # Bundle match - check if interface is a bundle member
            if bd_base.startswith('bundle-'):
                members = bd_discovery.resolve_bundle_members(bd_base)
                if base_interface in members or interface in members:
                    return True
        
        return False
    
    def _find_bd_interface(self, bd: BridgeDomain, leaf_interface: str,
                           bd_discovery: BridgeDomainDiscovery) -> Optional[str]:
        """Find the specific BD interface (with sub-interface) for a physical interface."""
        base_interface = leaf_interface.split('.')[0]
        
        for bd_iface in bd.interfaces:
            bd_base = bd_iface.split('.')[0]
            
            # Direct match
            if bd_base == base_interface:
                return bd_iface
            
            # Bundle match
            if bd_base.startswith('bundle-'):
                members = bd_discovery.resolve_bundle_members(bd_base)
                if base_interface in members:
                    return bd_iface
        
        return None
    
    def _find_dut_subinterface(self, source_info: DeviceInfo, lldp_interface: str,
                                target_vlan: Optional[int], bd_name: str) -> str:
        """
        Find the DUT sub-interface that matches the BD VLAN.
        Uses cached SSH connection to avoid reconnecting per-BD.
        """
        if not target_vlan:
            return lldp_interface
        
        mgmt_ip = source_info.mgmt_ip
        if not mgmt_ip:
            mgmt_ip = source_info.serial
        if not mgmt_ip:
            return f"{lldp_interface}.{target_vlan}"
        
        print(f"    {Colors.CYAN}Looking for sub-interface with VLAN {target_vlan}...{Colors.ENDC}")
        
        try:
            ssh = self._get_cached_ssh(mgmt_ip, DEFAULT_USER, DEFAULT_PASS)
            if not ssh:
                return f"{lldp_interface}.{target_vlan}"
            
            base_if = lldp_interface.split('.')[0]
            
            # Method 1: Check sub-interfaces of the LLDP interface
            cmd1 = f"show configuration interfaces {base_if} | include vlan-id | no-more"
            output1 = self._run_dut_command(ssh, cmd1)
            
            for line in output1.split('\n'):
                if f'vlan-id {target_vlan}' in line or f'vlan-id: {target_vlan}' in line:
                    subif_match = re.search(rf'{re.escape(base_if)}\.(\d+)', output1)
                    if subif_match:
                        subif = f"{base_if}.{subif_match.group(1)}"
                        print(f"    {Colors.GREEN}Found sub-interface: {subif}{Colors.ENDC}")
                        return subif
            
            # Method 2: Search all interfaces for the VLAN
            cmd2 = f"show configuration interfaces | include {target_vlan} | no-more"
            output2 = self._run_dut_command(ssh, cmd2)
            
            for line in output2.split('\n'):
                if base_if in line:
                    subif_match = re.search(rf'({re.escape(base_if)}\.\d+)', line)
                    if subif_match:
                        subif = subif_match.group(1)
                        print(f"    {Colors.GREEN}Found sub-interface: {subif}{Colors.ENDC}")
                        return subif
            
            # Method 3: Try common sub-interface pattern
            common_subif = f"{base_if}.{target_vlan}"
            cmd3 = f"show configuration interfaces {common_subif} | no-more"
            output3 = self._run_dut_command(ssh, cmd3)
            
            if 'Unknown' not in output3 and 'ERROR' not in output3 and output3.strip():
                print(f"    {Colors.GREEN}Found sub-interface: {common_subif}{Colors.ENDC}")
                return common_subif
                
        except Exception as e:
            print(f"    {Colors.YELLOW}Could not determine sub-interface: {e}{Colors.ENDC}")
        
        return f"{lldp_interface}.{target_vlan}"
    
    def _run_dut_command(self, ssh: SSHConnection, cmd: str) -> str:
        """Run a command on DUT and return output."""
        try:
            return ssh.execute(cmd)
        except Exception as e:
            return ""
    
    def _trace_bd_path(self, bd: BridgeDomain, bd_path_info: BDPathInfo,
                       current_device: str, current_ip: str,
                       bd_discovery: BridgeDomainDiscovery,
                       lldp_neighbors: List[Dict[str, str]]):
        """
        Trace the BD path from current device through DNAAS fabric.
        Follows BD interfaces to find next hops.
        
        NOTE: lldp_neighbors is a list of dicts with keys: 
              'local_interface', 'neighbor_name', 'neighbor_interface', 'ttl'
        """
        # Find BD interfaces that lead to other DNAAS devices (not DUT)
        uplink_interfaces = []
        
        for bd_iface in bd.interfaces:
            base_iface = bd_iface.split('.')[0]
            
            # Skip interfaces that connect to DUT (we already handled that)
            is_dut_facing = False
            for n in lldp_neighbors:
                if n['local_interface'].split('.')[0] == base_iface:
                    if not any(kw in n['neighbor_name'].upper() for kw in DNAAS_KEYWORDS):
                        is_dut_facing = True
                        break
            
            if is_dut_facing:
                continue
            
            # Find DNAAS neighbors on this interface
            for n in lldp_neighbors:
                if n['local_interface'].split('.')[0] == base_iface:
                    if any(kw in n['neighbor_name'].upper() for kw in DNAAS_KEYWORDS):
                        uplink_interfaces.append({
                            'bd_interface': bd_iface,
                            'local_if': n['local_interface'],
                            'neighbor': n['neighbor_name'],
                            'neighbor_if': n['neighbor_interface']
                        })
            
            # Also check if it's a bundle
            if base_iface.startswith('bundle-'):
                members = bd_discovery.resolve_bundle_members(base_iface)
                for member in members:
                    for n in lldp_neighbors:
                        if n['local_interface'] == member:
                            if any(kw in n['neighbor_name'].upper() for kw in DNAAS_KEYWORDS):
                                if n['neighbor_name'] not in [u['neighbor'] for u in uplink_interfaces]:
                                    uplink_interfaces.append({
                                        'bd_interface': bd_iface,
                                        'local_if': member,
                                        'neighbor': n['neighbor_name'],
                                        'neighbor_if': n['neighbor_interface']
                                    })
        
        # IMPORTANT: Include initiator DUT + all its aliases in visited
        # so it's NOT re-discovered as a termination on remote LEAFs
        initiator_name = bd_path_info.path[0]['device'] if bd_path_info.path else None
        visited = {current_device}
        if initiator_name:
            visited.add(initiator_name)
            # Add known aliases: serial, hostname, NCC serials
            inv_entry = self.inventory.get_device(initiator_name)
            if inv_entry:
                if inv_entry.hostname:
                    visited.add(inv_entry.hostname.rstrip(','))
                visited.add(inv_entry.serial)
            # For cluster devices, also add the main device name and NCC entries
            for key, dev in self.inventory.data.get("devices", {}).items():
                h = (dev.get('hostname') or '').split('(')[0].strip()
                if h and (h == initiator_name or key == initiator_name):
                    visited.add(key)
                    visited.add(h)
        
        for uplink in uplink_interfaces:
            neighbor = uplink['neighbor']
            
            if neighbor in visited:
                continue
            visited.add(neighbor)
            
            # Look up neighbor IP
            neighbor_info = self.dnaas_table.lookup(neighbor)
            if not neighbor_info or not neighbor_info.get('ip'):
                continue
            
            neighbor_ip = neighbor_info['ip']
            device_type = self._classify_device(neighbor)
            
            # Add to path
            bd_path_info.path.append({
                'device': neighbor,
                'device_type': device_type,
                'interface': uplink['neighbor_if'],
                'bd_interface': uplink['bd_interface']
            })
            
            bd_path_info.interfaces[neighbor] = uplink['neighbor_if']
            
            # Add connection
            self.discovered_connections.append({
                'from': current_device,
                'to': neighbor,
                'from_if': uplink['local_if'],
                'to_if': uplink['neighbor_if'],
                'bd_name': bd.name,
                'bd_type': bd_path_info.bd_type,
                'global_vlan': bd_path_info.global_vlan
            })
            
            # Recursively trace through fabric
            # ENHANCED: Also trace through LEAFs to find remote terminations
            if device_type in ['SPINE', 'SUPERSPINE']:
                self._trace_through_fabric(
                    bd, bd_path_info, neighbor, neighbor_ip, visited,
                    incoming_interface=uplink['neighbor_if']  # Pass incoming interface for config-based BD lookup
                )
            elif device_type == 'LEAF':
                # Found another LEAF - discover its connected DUTs
                self._discover_leaf_terminations(
                    bd, bd_path_info, neighbor, neighbor_ip, visited,
                    incoming_interface=uplink['neighbor_if']  # Pass incoming interface
                )
    
    def _trace_through_fabric(self, bd: BridgeDomain, bd_path_info: BDPathInfo,
                              current_device: str, current_ip: str, visited: set,
                              incoming_interface: str = None):
        """
        Continue tracing BD path through SPINE/SuperSpine.
        
        IMPROVED: Uses config-based BD discovery:
        show config network-services bridge-domain | include <interface> leading 10 trailing 10
        
        This finds the exact BD that contains the incoming interface.
        """
        if len(visited) > self.max_depth:
            return
        
        ssh = SSHConnection(current_ip, DNAAS_USER, DNAAS_PASS)
        if not ssh.connect():
            return
        
        try:
            bd_discovery = BridgeDomainDiscovery(ssh, current_device)
            
            # Get LLDP neighbors first - we'll need them for finding incoming interface
            lldp_neighbors = bd_discovery.get_lldp_neighbors()
            print(f"      LLDP on {current_device}: {len(lldp_neighbors)} neighbors")
            
            # Debug: Show LEAF neighbors
            leaf_neighbors = [n for n in lldp_neighbors if 'LEAF' in n['neighbor_name'].upper()]
            if leaf_neighbors and 'SPINE' in current_device.upper():
                leaf_strs = [n['local_interface'] + "->" + n['neighbor_name'] for n in leaf_neighbors]
                print(f"        LEAF neighbors: {leaf_strs}")
            
            matching_bd = None
            
            # IMPROVED: Use config-based BD discovery if we have incoming interface
            if incoming_interface:
                print(f"    Tracing BD via incoming interface {incoming_interface} on {current_device}...")
                bd_names = bd_discovery.find_bd_from_config(incoming_interface)
                if bd_names:
                    # Get details of the first matching BD
                    matching_bd = bd_discovery.get_bd_details_by_name(bd_names[0])
                    if matching_bd:
                        print(f"      Found BD: {matching_bd.name} with {len(matching_bd.interfaces)} interfaces")
            
            # Fallback: Try to find BD by VLAN pattern from original BD name
            if not matching_bd and bd_path_info.global_vlan:
                vlan_pattern = f"v{bd_path_info.global_vlan}"
                print(f"    Searching for BD with VLAN pattern {vlan_pattern} on {current_device}...")
                
                # Search config for this VLAN pattern
                # Config format: "instance g_yor_v210" not "bridge-domain g_yor_v210"
                cmd = f"show config network-services bridge-domain | include {vlan_pattern} leading 5 trailing 5 | no-more"
                output = ssh.send_command(cmd, timeout=45)
                
                # Parse output for BD names - format is "instance <bd_name>"
                # IMPORTANT: Use precise matching to avoid v214 matching v2146
                # Pattern must match vXXX followed by non-digit or end of word
                precise_pattern = r'instance\s+(\S*' + vlan_pattern + r'(?:\D|_|$)\S*)'
                for line in output.split('\n'):
                    line_stripped = line.strip()
                    instance_match = re.match(precise_pattern, line_stripped)
                    if instance_match:
                        bd_name = instance_match.group(1)
                        # Double-check: the BD name should contain exactly our VLAN pattern
                        # e.g., "g_yor_v214" should match v214, but "DHCP_v2146" should NOT
                        if re.search(vlan_pattern + r'(?:\D|_|$)', bd_name):
                            matching_bd = bd_discovery.get_bd_details_by_name(bd_name)
                            if matching_bd:
                                print(f"      {Colors.GREEN}Found BD by VLAN pattern: {matching_bd.name}{Colors.ENDC}")
                                break
            
            # Final fallback: Match by BD name substring
            if not matching_bd:
                for target_bd in bd_discovery.get_bridge_domains():
                    if bd.name in target_bd.name or target_bd.name in bd.name:
                        matching_bd = target_bd
                        break
            
            if not matching_bd:
                print(f"      {Colors.YELLOW}No matching BD found on {current_device}{Colors.ENDC}")
                return
            
            # SKIP mgmt BDs completely - they span entire fabric
            if 'mgmt' in matching_bd.name.lower():
                print(f"      {Colors.YELLOW}Skipping mgmt BD: {matching_bd.name}{Colors.ENDC}")
                return
            
            # Find uplinks to other DNAAS devices
            print(f"      Checking {len(matching_bd.interfaces)} BD interfaces against {len(lldp_neighbors)} LLDP neighbors")
            
            for bd_iface in matching_bd.interfaces:
                base_iface = bd_iface.split('.')[0]
                
                # Build list of interfaces to check against LLDP neighbors
                interfaces_to_check = []
                
                if base_iface.startswith('bundle-'):
                    # For bundles, check physical members via LLDP
                    members = bd_discovery.resolve_bundle_members(base_iface)
                    if members:
                        interfaces_to_check.extend(members)
                        print(f"        {base_iface} members: {members}")
                    else:
                        # Empty bundle — use VLAN-hop: probe unvisited DNAAS neighbors
                        # for the same VLAN BD on their side
                        print(f"      {Colors.YELLOW}Bundle {base_iface} has no members — trying VLAN-hop for v{bd_path_info.global_vlan}{Colors.ENDC}")
                        if bd_path_info.global_vlan:
                            self._vlan_hop_through_neighbors(
                                bd, bd_path_info, current_device,
                                lldp_neighbors, visited, bd_discovery
                            )
                else:
                    # Physical interface - check directly
                    interfaces_to_check.append(base_iface)
                
                for check_if in interfaces_to_check:
                    for n in lldp_neighbors:
                        # lldp_neighbors is a list of dicts from get_lldp_neighbors()
                        local_if_base = n['local_interface'].split('.')[0]
                        if local_if_base == check_if or n['local_interface'] == check_if:
                            neighbor = n['neighbor_name']
                            
                            # Debug: Show LEAF matches on SPINEs
                            if 'SPINE' in current_device.upper() and 'LEAF' in neighbor.upper():
                                print(f"        FOUND: {check_if} -> {neighbor} (visited={neighbor in visited})")
                            
                            if neighbor in visited:
                                continue
                            
                            if any(kw in neighbor.upper() for kw in DNAAS_KEYWORDS):
                                visited.add(neighbor)
                                
                                neighbor_info = self.dnaas_table.lookup(neighbor)
                                if not neighbor_info:
                                    # Create minimal info for DNAAS devices (they're SSH-able by hostname)
                                    neighbor_info = {'ip': neighbor, 'hostname': neighbor}
                                
                                device_type = self._classify_device(neighbor)
                                
                                # Add to path
                                bd_path_info.path.append({
                                    'device': neighbor,
                                    'device_type': device_type,
                                    'interface': n['neighbor_interface'],
                                    'bd_interface': bd_iface
                                })
                                
                                bd_path_info.interfaces[neighbor] = n['neighbor_interface']
                                
                                # Add connection
                                self.discovered_connections.append({
                                    'from': current_device,
                                    'to': neighbor,
                                    'from_if': n['local_interface'],
                                    'to_if': n['neighbor_interface'],
                                    'bd_name': bd.name,
                                    'bd_type': bd_path_info.bd_type,
                                    'global_vlan': bd_path_info.global_vlan
                                })
                                
                                # Continue tracing through fabric (up and down)
                                # ENHANCED: Trace through ALL DNAAS devices, not just SUPERSPINE
                                # This ensures we find termination points on remote LEAFs
                                neighbor_ip = neighbor_info.get('ip')
                                
                                # Try hostname directly if no IP (DNAAS devices are resolvable by name)
                                if not neighbor_ip and any(kw in neighbor.upper() for kw in ['LEAF', 'SPINE']):
                                    neighbor_ip = neighbor  # Use hostname directly for SSH
                                
                                if neighbor_ip:
                                    if device_type in ['SPINE', 'SUPERSPINE']:
                                        # Continue up/across through SPINE/SUPERSPINE
                                        self._trace_through_fabric(
                                            bd, bd_path_info, neighbor, 
                                            neighbor_ip, visited,
                                            incoming_interface=n['neighbor_interface']  # Config-based lookup
                                        )
                                    elif device_type == 'LEAF':
                                        # LEAF found - discover its connected DUTs/PEs
                                        print(f"        -> Tracing LEAF {neighbor} via {neighbor_ip}")
                                        self._discover_leaf_terminations(
                                            bd, bd_path_info, neighbor,
                                            neighbor_ip, visited,
                                            incoming_interface=n['neighbor_interface']  # Config-based lookup
                                        )
                                    else:
                                        print(f"        -> Skipping {neighbor} (type={device_type})")
                                else:
                                    print(f"          {Colors.YELLOW}Cannot trace to {neighbor} - no IP available{Colors.ENDC}")
        
        except Exception as e:
            print(f"      {Colors.RED}Error tracing through {current_device}: {e}{Colors.ENDC}")
        finally:
            ssh.close()
    
    def _vlan_hop_through_neighbors(self, bd: BridgeDomain, bd_path_info: BDPathInfo,
                                     current_device: str, lldp_neighbors: list,
                                     visited: set, bd_discovery: BridgeDomainDiscovery):
        """
        Workaround for empty bundles: probe all unvisited DNAAS LLDP neighbors
        in parallel and check if they carry a BD with the same VLAN.
        """
        from concurrent.futures import ThreadPoolExecutor, as_completed
        
        vlan = bd_path_info.global_vlan
        if not vlan:
            return
        
        vlan_pattern = f"v{vlan}"
        candidates = []
        seen = set()
        
        for n in lldp_neighbors:
            neighbor = n['neighbor_name']
            if neighbor in visited or neighbor in seen:
                continue
            if not any(kw in neighbor.upper() for kw in DNAAS_KEYWORDS):
                continue
            seen.add(neighbor)
            
            neighbor_info = self.dnaas_table.lookup(neighbor)
            neighbor_ip = neighbor_info.get('ip') if neighbor_info else None
            if not neighbor_ip:
                neighbor_ip = neighbor
            candidates.append((n, neighbor, neighbor_ip))
        
        if not candidates:
            return
        
        print(f"      [VLAN-hop] Probing {len(candidates)} neighbors in parallel for v{vlan}...")
        
        def _probe_one(entry):
            """SSH probe a single neighbor — returns match info or None."""
            n, neighbor, neighbor_ip = entry
            probe_ssh = SSHConnection(neighbor_ip, DNAAS_USER, DNAAS_PASS)
            if not probe_ssh.connect():
                return None
            try:
                cmd = f"show config network-services bridge-domain | include {vlan_pattern} leading 1 | no-more"
                output = probe_ssh.send_command(cmd, timeout=20)
                
                found_bd_name = None
                precise = r'instance\s+(\S*' + vlan_pattern + r'(?:\D|_|$)\S*)'
                for line in output.split('\n'):
                    m = re.match(precise, line.strip())
                    if m and re.search(vlan_pattern + r'(?:\D|_|$)', m.group(1)):
                        found_bd_name = m.group(1)
                        break
                
                if found_bd_name:
                    return (n, neighbor, neighbor_ip, found_bd_name)
                return None
            except Exception as e:
                print(f"      {Colors.YELLOW}[VLAN-hop] {neighbor} probe failed: {e}{Colors.ENDC}")
                return None
            finally:
                probe_ssh.close()
        
        hits = []
        with ThreadPoolExecutor(max_workers=min(len(candidates), 8)) as pool:
            futures = {pool.submit(_probe_one, c): c for c in candidates}
            for fut in as_completed(futures):
                result = fut.result()
                if result:
                    hits.append(result)
        
        for n, neighbor, neighbor_ip, found_bd_name in hits:
            if neighbor in visited:
                continue
            
            device_type = self._classify_device(neighbor)
            print(f"      {Colors.GREEN}[VLAN-hop] {neighbor} has BD {found_bd_name} (v{vlan}){Colors.ENDC}")
            visited.add(neighbor)
            
            bd_path_info.path.append({
                'device': neighbor,
                'device_type': device_type,
                'interface': n['neighbor_interface'],
                'bd_interface': found_bd_name
            })
            bd_path_info.interfaces[neighbor] = n['neighbor_interface']
            
            self.discovered_connections.append({
                'from': current_device,
                'to': neighbor,
                'from_if': n['local_interface'],
                'to_if': n['neighbor_interface'],
                'bd_name': bd.name,
                'bd_type': bd_path_info.bd_type,
                'global_vlan': vlan
            })
            
            if device_type in ['SPINE', 'SUPERSPINE']:
                self._trace_through_fabric(
                    bd, bd_path_info, neighbor,
                    neighbor_ip, visited,
                    incoming_interface=n['neighbor_interface']
                )
            elif device_type == 'LEAF':
                print(f"        -> VLAN-hop LEAF {neighbor} via {neighbor_ip}")
                self._discover_leaf_terminations(
                    bd, bd_path_info, neighbor,
                    neighbor_ip, visited,
                    incoming_interface=n['neighbor_interface']
                )
    
    def _discover_leaf_terminations(self, bd: BridgeDomain, bd_path_info: BDPathInfo,
                                      leaf_name: str, leaf_ip: str, visited: set,
                                      incoming_interface: str = None):
        """
        Discover DUTs/PEs connected to a LEAF that terminate the BD.
        This finds the remote termination points of the BD path.
        
        IMPROVED: Uses config-based BD discovery via incoming interface.
        """
        if len(visited) > self.max_depth:
            print(f"    {Colors.YELLOW}Max depth reached for {leaf_name}{Colors.ENDC}")
            return
        
        ssh = SSHConnection(leaf_ip, DNAAS_USER, DNAAS_PASS)
        if not ssh.connect():
            print(f"    {Colors.RED}SSH failed to {leaf_name} ({leaf_ip}){Colors.ENDC}")
            print(f"      {Colors.YELLOW}Could not connect to LEAF {leaf_name}{Colors.ENDC}")
            return
        
        try:
            bd_discovery = BridgeDomainDiscovery(ssh, leaf_name)
            
            # Get LLDP neighbors first
            lldp_neighbors = bd_discovery.get_lldp_neighbors()
            print(f"      LLDP on {leaf_name}: {len(lldp_neighbors)} neighbors")
            
            # Show non-DNAAS neighbors (potential terminations)
            pe_neighbors = [n for n in lldp_neighbors if not any(kw in n['neighbor_name'].upper() for kw in DNAAS_KEYWORDS)]
            if pe_neighbors:
                for pn in pe_neighbors[:5]:
                    print(f"        PE neighbor: {pn['local_interface']} -> {pn['neighbor_name']}")
            
            # IMPROVED: Use config-based BD discovery if we have incoming interface
            matching_bd = None
            
            if incoming_interface:
                print(f"    Discovering LEAF terminations via incoming interface {incoming_interface} on {leaf_name}...")
                bd_names = bd_discovery.find_bd_from_config(incoming_interface)
                if bd_names:
                    matching_bd = bd_discovery.get_bd_details_by_name(bd_names[0])
                    if matching_bd:
                        print(f"      Found BD: {matching_bd.name} with {len(matching_bd.interfaces)} interfaces")
            
            # Fallback: Search by VLAN pattern
            if not matching_bd and bd_path_info.global_vlan:
                vlan_pattern = f"v{bd_path_info.global_vlan}"
                print(f"    Searching for BD with VLAN pattern {vlan_pattern} on {leaf_name}...")
                
                cmd = f"show config network-services bridge-domain | include {vlan_pattern} leading 5 trailing 5 | no-more"
                output = ssh.send_command(cmd, timeout=45)
                
                # Parse output - format is "instance <bd_name>"
                # IMPORTANT: Use precise matching to avoid v214 matching v2146
                precise_pattern = r'instance\s+(\S*' + vlan_pattern + r'(?:\D|_|$)\S*)'
                for line in output.split('\n'):
                    line_stripped = line.strip()
                    instance_match = re.match(precise_pattern, line_stripped)
                    if instance_match:
                        bd_name = instance_match.group(1)
                        # Double-check the VLAN pattern is exact
                        if re.search(vlan_pattern + r'(?:\D|_|$)', bd_name):
                            matching_bd = bd_discovery.get_bd_details_by_name(bd_name)
                            if matching_bd:
                                print(f"      {Colors.GREEN}Found BD by VLAN pattern: {matching_bd.name}{Colors.ENDC}")
                                break
            
            # Final fallback: Match by BD name substring
            if not matching_bd:
                for target_bd in bd_discovery.get_bridge_domains():
                    if bd.name in target_bd.name or target_bd.name in bd.name:
                        matching_bd = target_bd
                        break
            
            if not matching_bd:
                print(f"      {Colors.YELLOW}BD {bd.name} not found on LEAF {leaf_name}{Colors.ENDC}")
                return
            
            # SKIP mgmt BDs completely - they span entire fabric
            if 'mgmt' in matching_bd.name.lower():
                print(f"      {Colors.YELLOW}Skipping mgmt BD: {matching_bd.name}{Colors.ENDC}")
                return
            
            # Find non-DNAAS neighbors (these are DUTs/PEs terminating the BD)
            print(f"      BD {matching_bd.name} has {len(matching_bd.interfaces)} interfaces: {matching_bd.interfaces[:5]}...")
            
            # ========================================
            # PHASE 1: Identify all potential DUT terminations from LLDP
            # Map: DUT name -> {physical_if, neighbor_interface, ...}
            # ========================================
            potential_duts = {}  # dut_name -> {'local_if': ..., 'neighbor_if': ...}
            
            for bd_iface in matching_bd.interfaces:
                base_iface = bd_iface.split('.')[0]
                
                # Build list of interfaces to check against LLDP neighbors
                interfaces_to_check = []
                
                if base_iface.startswith('bundle-'):
                    # For bundles, check physical members via LLDP
                    members = bd_discovery.resolve_bundle_members(base_iface)
                    if members:
                        interfaces_to_check.extend(members)
                    else:
                        print(f"        {Colors.YELLOW}Bundle {base_iface} has no members{Colors.ENDC}")
                else:
                    # Physical interface - check directly
                    interfaces_to_check.append(base_iface)
                
                for check_if in interfaces_to_check:
                    for n in lldp_neighbors:
                        local_if = n['local_interface'].split('.')[0]
                        if local_if == check_if or n['local_interface'] == check_if:
                            neighbor = n['neighbor_name']
                            
                            # Skip DNAAS devices
                            if any(kw in neighbor.upper() for kw in DNAAS_KEYWORDS):
                                continue
                            
                            # Skip already visited (exact match OR hostname suffix match)
                            # This prevents the initiator DUT from being discovered as a termination
                            # ENHANCED: Handle prefixed hostnames (e.g., YOR_PE-1 vs PE-1)
                            skip_neighbor = False
                            for v in visited:
                                if neighbor == v:
                                    skip_neighbor = True
                                    break
                                    
                                neighbor_upper = neighbor.upper()
                                v_upper = v.upper()
                                
                                # Direct substring match
                                if v_upper in neighbor_upper or neighbor_upper in v_upper:
                                    skip_neighbor = True
                                    print(f"        Skipping {neighbor} - matches initiator {v}")
                                    break
                                
                                # ENHANCED: Extract base hostname (after last underscore or hyphen prefix)
                                # Handle cases like: YOR_PE-1 vs PE-1, LAB_R1 vs R1
                                # Get base name: strip common prefixes like "YOR_", "LAB_", etc.
                                neighbor_base = re.sub(r'^[A-Z]+_', '', neighbor_upper)
                                v_base = re.sub(r'^[A-Z]+_', '', v_upper)
                                
                                if neighbor_base == v_base:
                                    skip_neighbor = True
                                    print(f"        Skipping {neighbor} - base name matches {v} ({neighbor_base} == {v_base})")
                                    break
                                
                                # Also check if one ends with the other (e.g., serial contains hostname)
                                if neighbor_base.endswith(v_base) or v_base.endswith(neighbor_base):
                                    skip_neighbor = True
                                    print(f"        Skipping {neighbor} - suffix matches {v}")
                                    break
                                    
                            if skip_neighbor:
                                continue
                            
                            # Skip mgmt BDs
                            if 'mgmt' in bd.name.lower():
                                continue
                            
                            if neighbor not in potential_duts:
                                potential_duts[neighbor] = {
                                    'local_if': n['local_interface'],
                                    'neighbor_if': n['neighbor_interface'],
                                    'bd_interfaces': []
                                }
                            potential_duts[neighbor]['bd_interfaces'].append(bd_iface)
            
            if not potential_duts:
                return
            
            print(f"      {Colors.CYAN}Found {len(potential_duts)} potential DUT terminations{Colors.ENDC}")
            
            # ========================================
            # PHASE 2: Query each DUT for its sub-interfaces (VLAN discovery)
            # ========================================
            dut_subif_cache = {}  # dut_name -> {vlan_map, all_subifs, ...}
            
            for dut_name, dut_info in potential_duts.items():
                print(f"      {Colors.CYAN}Querying DUT {dut_name} for sub-interfaces...{Colors.ENDC}")
                subif_info = self._get_dut_subinterfaces(dut_name, dut_info['neighbor_if'])
                if subif_info:
                    dut_subif_cache[dut_name] = subif_info
            
            # ========================================
            # PHASE 3: Match BD VLANs with DUT sub-interfaces
            # Only create connections where VLANs match
            # ========================================
            bd_vlan = bd_path_info.global_vlan
            
            for dut_name, dut_info in potential_duts.items():
                visited.add(dut_name)
                
                # Classify termination device
                device_type = self._classify_device(dut_name)
                if device_type in ['SPINE', 'LEAF', 'SUPERSPINE', 'DNAAS']:
                    device_type = 'PE'
                
                type_label = device_type if device_type in ['TEST_EQUIPMENT', 'SERVER'] else 'PE'
                
                # Check if DUT has sub-interface with BD's VLAN
                dut_subif_name = ""
                dut_has_matching_subif = False
                
                if bd_vlan and dut_name in dut_subif_cache:
                    cached_info = dut_subif_cache[dut_name]
                    vlan_map = cached_info.get('vlan_map', {})
                    
                    if bd_vlan in vlan_map:
                        dut_subif_name = vlan_map[bd_vlan]
                        dut_has_matching_subif = True
                        print(f"      {Colors.GREEN}✓ DUT {dut_name} has sub-if {dut_subif_name} matching VLAN {bd_vlan}{Colors.ENDC}")
                    else:
                        # DUT doesn't have this VLAN - skip
                        print(f"      {Colors.YELLOW}✗ DUT {dut_name} has NO sub-if with VLAN {bd_vlan} - skipping{Colors.ENDC}")
                        continue
                elif bd_vlan:
                    # Couldn't query DUT - try validation anyway
                    dut_validated = self._validate_dut_vlan_subinterface(
                        dut_name, dut_info['neighbor_if'], bd_vlan
                    )
                    if dut_validated:
                        dut_has_matching_subif = True
                        dut_subif_name = dut_validated.get('subif', '')
                        print(f"      {Colors.GREEN}✓ DUT {dut_name} validated with sub-if {dut_subif_name}{Colors.ENDC}")
                    else:
                        print(f"      {Colors.YELLOW}✗ DUT {dut_name} validation failed - skipping{Colors.ENDC}")
                        continue
                else:
                    # No VLAN to validate - accept based on LLDP
                    print(f"      {Colors.YELLOW}⚠ No VLAN to validate for BD {bd.name}{Colors.ENDC}")
                    dut_has_matching_subif = True
                
                print(f"      {Colors.GREEN}Found termination: {dut_name} ({type_label}) via {dut_info['local_if']}{Colors.ENDC}")
                
                # Add termination device to path
                bd_path_info.path.append({
                    'device': dut_name,
                    'device_type': device_type,
                    'interface': dut_subif_name or dut_info['neighbor_if'],
                    'bd_interface': dut_info['bd_interfaces'][0] if dut_info['bd_interfaces'] else '',
                    'is_termination': True,
                    'vlan_validated': dut_has_matching_subif
                })
                
                bd_path_info.interfaces[dut_name] = dut_subif_name or dut_info['neighbor_if']
                
                # Add connection (LEAF -> Termination PE)
                self.discovered_connections.append({
                    'from': leaf_name,
                    'to': dut_name,
                    'from_if': dut_info['local_if'],
                    'to_if': dut_subif_name or dut_info['neighbor_if'],
                                'bd_name': bd.name,
                                'bd_type': bd_path_info.bd_type,
                                'global_vlan': bd_path_info.global_vlan,
                    'is_termination': True,
                    'vlan_validated': dut_has_matching_subif
                            })
        
        except Exception as e:
            print(f"      {Colors.RED}Error discovering terminations on {leaf_name}: {e}{Colors.ENDC}")
        finally:
            ssh.close()
    
    def _classify_device(self, name: str) -> str:
        """Classify device type based on name."""
        upper_name = name.upper()
        if 'SUPERSPINE' in upper_name:
            return "SUPERSPINE"
        elif 'SPINE' in upper_name:
            return "SPINE"
        elif 'LEAF' in upper_name:
            return "LEAF"
        return "PE"
    
    def _get_dut_subinterfaces(self, dut_name: str, physical_if: str) -> Optional[Dict[str, Any]]:
        """
        Query a DUT for ALL its sub-interfaces on the uplink interface.
        Returns a dict mapping VLAN -> sub-interface name.
        
        This allows us to match DUT sub-interfaces with DNaaS BD VLANs.
        
        Args:
            dut_name: DUT hostname (may contain serial number)
            physical_if: Physical interface on DUT (from LLDP)
            
        Returns:
            Dict with 'vlan_map': {vlan: subif_name}, 'all_subifs': [list], or None if can't connect
        """
        # Extract serial number from DUT name if present
        dut_serial = None
        sn_patterns = [
            r'([A-Z]{2}\d{2}[A-Z0-9]{6,})',  # e.g., WK31D7VW00016
            r'([A-Z]{3}\d[A-Z]\d{2}[A-Z0-9]+)',  # e.g., YBW1F8VJ00002P1
            r'-([A-Z0-9]{8,})',  # After dash, 8+ chars
            r'_([A-Z]{2,3}\d+[A-Z0-9]+\d{4,})',  # After underscore with numbers
        ]
        for pattern in sn_patterns:
            match = re.search(pattern, dut_name, re.IGNORECASE)
            if match:
                dut_serial = match.group(1) if match.lastindex else match.group(0)
                break
        
        # Get DUT IP from inventory
        dut_ip = None
        if self.inventory:
            dut_info = self.inventory.get_device(dut_serial or dut_name)
            if dut_info:
                dut_ip = dut_info.mgmt_ip
        
        if not dut_ip:
            # Try DNAAS table as fallback (for test equipment)
            dnaas_info = self.dnaas_table.lookup(dut_name) if self.dnaas_table else None
            if dnaas_info:
                dut_ip = dnaas_info.get('ip', '')
        
        if not dut_ip:
            print(f"        {Colors.YELLOW}Cannot query {dut_name} - no IP found{Colors.ENDC}")
            return None
        
        # Connect to DUT and get all sub-interfaces
        try:
            ssh = SSHConnection(dut_ip, "dnroot", "dnroot", timeout=30)
            if not ssh.connect():
                print(f"        {Colors.YELLOW}Cannot connect to {dut_name} ({dut_ip}){Colors.ENDC}")
                return None
            
            # Get the base interface name from the physical interface (from LLDP)
            base_if = physical_if.split('.')[0]
            
            # Query for ALL sub-interfaces on this physical interface
            # DNOS format: interfaces section lists sub-interfaces like "ge100-0/0/5.2635"
            cmd = f"show config interfaces | include {base_if}\\. | no-more"
            output = ssh.send_command(cmd, timeout=30)
            
            # Parse sub-interfaces and extract VLANs
            vlan_map = {}  # VLAN -> sub-interface name
            all_subifs = []
            
            # Match sub-interface patterns like "ge100-0/0/5.2635"
            subif_pattern = rf'({re.escape(base_if)}\.(\d+))'
            for match in re.finditer(subif_pattern, output):
                subif_name = match.group(1)
                vlan = int(match.group(2))
                vlan_map[vlan] = subif_name
                all_subifs.append(subif_name)
            
            # Also check for bundle interfaces if the physical is part of a bundle
            # Query bundle membership
            cmd2 = f"show config interfaces {base_if} | include bundle | no-more"
            output2 = ssh.send_command(cmd2, timeout=20)
            
            bundle_match = re.search(r'(bundle-\d+)', output2)
            if bundle_match:
                bundle_name = bundle_match.group(1)
                # Get sub-interfaces on the bundle
                cmd3 = f"show config interfaces | include {bundle_name}\\. | no-more"
                output3 = ssh.send_command(cmd3, timeout=30)
                
                bundle_subif_pattern = rf'({re.escape(bundle_name)}\.(\d+))'
                for match in re.finditer(bundle_subif_pattern, output3):
                    subif_name = match.group(1)
                    vlan = int(match.group(2))
                    if vlan not in vlan_map:
                        vlan_map[vlan] = subif_name
                        all_subifs.append(subif_name)
            
            ssh.close()
            
            print(f"        {Colors.CYAN}DUT {dut_name} has {len(vlan_map)} sub-ifs on {base_if}: VLANs {sorted(vlan_map.keys())[:10]}...{Colors.ENDC}")
            
            return {
                'vlan_map': vlan_map,
                'all_subifs': all_subifs,
                'base_if': base_if,
                'dut_ip': dut_ip
            }
            
        except Exception as e:
            print(f"        {Colors.YELLOW}Error querying {dut_name}: {e}{Colors.ENDC}")
            return None
    
    def _validate_dut_vlan_subinterface(self, dut_name: str, physical_if: str, vlan: int, 
                                         dut_subif_cache: Dict[str, Any] = None) -> Optional[Dict[str, Any]]:
        """
        Validate that a DUT has a sub-interface with the expected VLAN.
        
        Uses cached DUT sub-interface info if available to avoid repeated SSH connections.
        
        Args:
            dut_name: DUT hostname (may contain serial number)
            physical_if: Physical interface on DUT (from LLDP)
            vlan: Expected VLAN from the BD
            dut_subif_cache: Cached sub-interface info from _get_dut_subinterfaces()
            
        Returns:
            Dict with 'subif' if found, None if not
        """
        # Use cached info if available
        if dut_subif_cache:
            vlan_map = dut_subif_cache.get('vlan_map', {})
            if vlan in vlan_map:
                return {'subif': vlan_map[vlan], 'validated': True}
            else:
                return None  # DUT doesn't have this VLAN
        
        # No cache - query DUT directly (fallback)
        dut_info = self._get_dut_subinterfaces(dut_name, physical_if)
        if not dut_info:
            # Can't connect to DUT - accept with warning
            base_if = physical_if.split('.')[0]
            return {'subif': f"{base_if}.{vlan}", 'validated': False}
        
        vlan_map = dut_info.get('vlan_map', {})
        if vlan in vlan_map:
            return {'subif': vlan_map[vlan], 'validated': True}
        
        return None  # DUT doesn't have this VLAN
    
    def discover_all_infrastructure_links(self):
        """
        Discover ALL links between DNaaS infrastructure devices (Leaf ↔ Spine ↔ SuperSpine).
        Also discovers ALL termination DUTs connected to each BD.
        
        IMPORTANT: Infrastructure links are now assigned BD colors based on which BDs traverse them.
        No more gray infrastructure links - each link shows which BDs it carries.
        """
        print(f"\n{Colors.HEADER}=== Discovering ALL Infrastructure Links ==={Colors.ENDC}")
        
        # Collect all DNaaS devices we've seen and which BDs they're part of
        dnaas_devices = {}  # device_name -> set of BD names
        for bd_name, bd_info in self.bd_paths.items():
            for hop in bd_info.path:
                device_type = hop['device_type']
                if device_type in ['LEAF', 'SPINE', 'SUPERSPINE']:
                    device_name = hop['device']
                    if device_name not in dnaas_devices:
                        dnaas_devices[device_name] = {'type': device_type, 'bds': set()}
                    dnaas_devices[device_name]['bds'].add(bd_name)
        
        print(f"  Found {len(dnaas_devices)} DNaaS devices to query")
        
        # Track infrastructure links and which BDs they carry
        infra_links = {}  # (dev1, dev2) -> {'from_if', 'to_if', 'bds': set()}
        
        # Query each DNaaS device for LLDP neighbors
        # IMPORTANT: Use list() to avoid "dictionary changed size during iteration" error
        for device_name, device_info in list(dnaas_devices.items()):
            device_type = device_info['type']
            device_bds = device_info['bds']
            
            print(f"\n  {Colors.CYAN}Querying {device_name} ({device_type}, {len(device_bds)} BDs)...{Colors.ENDC}")
            
            # Get IP from DNAAS table
            dnaas_info = self.dnaas_table.lookup(device_name) if self.dnaas_table else None
            device_ip = dnaas_info.get('ip', '') if dnaas_info else ''
            
            if not device_ip:
                print(f"    {Colors.YELLOW}No IP for {device_name}, skipping{Colors.ENDC}")
                continue
            
            ssh = SSHConnection(device_ip, DNAAS_USER, DNAAS_PASS)
            if not ssh.connect():
                print(f"    {Colors.RED}Failed to connect{Colors.ENDC}")
                continue
            
            try:
                bd_discovery = BridgeDomainDiscovery(ssh, device_name)
                lldp_neighbors = bd_discovery.get_lldp_neighbors()
                
                for n in lldp_neighbors:
                    neighbor = n['neighbor_name']
                    local_if = n['local_interface']
                    remote_if = n['neighbor_interface']
                    
                    # Classify neighbor
                    neighbor_type = self._classify_device(neighbor)
                    
                    # Check if this is a DNaaS infrastructure link
                    is_infra = neighbor_type in ['LEAF', 'SPINE', 'SUPERSPINE']
                    
                    if is_infra:
                        # Create sorted key to avoid duplicates
                        link_key = tuple(sorted([device_name, neighbor]))
                        
                        # Find which BDs this link carries (intersection of BDs from both devices)
                        neighbor_bds = set()
                        if neighbor in dnaas_devices:
                            neighbor_bds = dnaas_devices[neighbor]['bds']
                        
                        # BDs that BOTH devices are part of = BDs that flow through this link
                        shared_bds = device_bds & neighbor_bds if neighbor_bds else device_bds
                        
                        # ONLY add link if BOTH devices are already in the BD path
                        # This prevents adding infrastructure devices that aren't part of the traced BDs
                        if neighbor not in dnaas_devices:
                            # Neighbor wasn't discovered in BD path - skip this link
                            print(f"    {Colors.DIM}Skipping {neighbor} - not in traced BD path{Colors.ENDC}")
                            continue
                        
                        if link_key not in infra_links:
                            infra_links[link_key] = {
                                'from': link_key[0],
                                'to': link_key[1],
                                'from_if': local_if if device_name == link_key[0] else remote_if,
                                'to_if': remote_if if device_name == link_key[0] else local_if,
                                'bds': shared_bds
                            }
                            print(f"    {Colors.GREEN}+ Infra link: {link_key[0]} ↔ {link_key[1]} ({len(shared_bds)} BDs){Colors.ENDC}")
                        else:
                            # Update with more BDs
                            infra_links[link_key]['bds'].update(shared_bds)
                    
                    # Check for termination DUTs (non-DNaaS devices)
                    elif neighbor_type in ['PE', 'TEST_EQUIPMENT', 'SERVER'] or neighbor_type == 'UNKNOWN':
                        # This is a potential termination DUT - check if connected to any BD
                        print(f"    {Colors.BLUE}? Potential termination: {neighbor} via {local_if}{Colors.ENDC}")
                        
                        # Find BDs on this interface
                        bd_names = bd_discovery.find_bd_from_config(local_if.split('.')[0])
                        
                        # ========================================
                        # VLAN VALIDATION: Query DUT for its sub-interfaces
                        # Only add termination if DUT has matching VLAN
                        # ========================================
                        dut_subif_info = None
                        if bd_names:
                            dut_subif_info = self._get_dut_subinterfaces(neighbor, remote_if)
                        
                        for bd_name in bd_names:
                            # Check if this BD is one we're tracking
                            if bd_name in self.bd_paths:
                                bd_info = self.bd_paths[bd_name]
                                bd_vlan = bd_info.global_vlan
                                
                                # VLAN validation: Check if DUT has sub-interface with this VLAN
                                # Default: construct expected sub-if from physical interface + BD VLAN
                                base_remote_if = remote_if.split('.')[0]
                                dut_subif = f"{base_remote_if}.{bd_vlan}" if bd_vlan else remote_if
                                vlan_validated = False
                                
                                if bd_vlan and dut_subif_info:
                                    vlan_map = dut_subif_info.get('vlan_map', {})
                                    if bd_vlan in vlan_map:
                                        dut_subif = vlan_map[bd_vlan]
                                        vlan_validated = True
                                        print(f"      {Colors.GREEN}✓ DUT {neighbor} has sub-if {dut_subif} (VLAN {bd_vlan}){Colors.ENDC}")
                                    else:
                                        # DUT queried but doesn't have this VLAN - skip this BD
                                        print(f"      {Colors.YELLOW}✗ DUT {neighbor} has NO sub-if with VLAN {bd_vlan} - skipping BD {bd_name}{Colors.ENDC}")
                                        continue
                                elif bd_vlan and not dut_subif_info:
                                    # Couldn't query DUT - accept with constructed sub-if name
                                    print(f"      {Colors.YELLOW}⚠ Cannot query DUT {neighbor} - assuming sub-if {dut_subif}{Colors.ENDC}")
                                    vlan_validated = False  # Not validated but accepted
                                
                                # Check if already have this termination
                                exists = any(
                                    c['from'] == device_name and c['to'] == neighbor and c.get('bd_name') == bd_name
                                    for c in self.discovered_connections
                                )
                                if not exists:
                                    print(f"      {Colors.GREEN}+ Termination on BD {bd_name}: {neighbor} (sub-if: {dut_subif}){Colors.ENDC}")
                                    self.discovered_connections.append({
                                        'from': device_name,
                                        'to': neighbor,
                                        'from_if': local_if,
                                        'to_if': dut_subif,
                                        'bd_name': bd_name,
                                        'is_termination': True,
                                        'vlan_validated': vlan_validated
                                    })
                                    
                                    # Add to path
                                    found = any(h['device'] == neighbor for h in bd_info.path)
                                    if not found:
                                        bd_info.path.append({
                                            'device': neighbor,
                                            'device_type': 'PE',
                                            'interface': dut_subif,
                                            'is_termination': True,
                                            'vlan_validated': vlan_validated
                                        })
            except Exception as e:
                print(f"    {Colors.RED}Error: {e}{Colors.ENDC}")
            finally:
                ssh.close()
        
        # Now add infrastructure links as BD-colored connections
        # Create one connection per BD for each infrastructure link
        print(f"\n  {Colors.CYAN}Creating BD-colored infrastructure links...{Colors.ENDC}")
        for link_key, link_info in infra_links.items():
            bds = link_info['bds']
            if not bds:
                # No BDs found, assign to first BD as fallback
                if self.bd_paths:
                    bds = {list(self.bd_paths.keys())[0]}
            
            for bd_name in bds:
                # Check if connection already exists for this BD
                exists = any(
                    c['from'] == link_info['from'] and c['to'] == link_info['to'] and c.get('bd_name') == bd_name
                    for c in self.discovered_connections
                )
                if not exists:
                    self.discovered_connections.append({
                        'from': link_info['from'],
                        'to': link_info['to'],
                        'from_if': link_info['from_if'],
                        'to_if': link_info['to_if'],
                        'bd_name': bd_name,
                        'is_infra': True  # Mark as infra but with BD color
                    })
        
        print(f"\n{Colors.GREEN}Infrastructure discovery complete - {len(self.discovered_connections)} total connections{Colors.ENDC}")
    
    def generate_multi_bd_topology(self, source_info: DeviceInfo) -> Dict[str, Any]:
        """
        Generate topology JSON with multi-BD support.
        Creates colored links and text boxes for each BD.
        """
        # First, discover ALL infrastructure links and termination DUTs
        self.discover_all_infrastructure_links()
        topology = {
            "version": "1.0",
            "objects": [],
            "metadata": {
                "name": "Multi-BD DNAAS Discovery",
                "created": datetime.now().isoformat(),
                "source": source_info.hostname or source_info.serial,
                "bridge_domains": [],  # Will be populated with ALL discovered BDs
                "device_bd_mapping": {},  # Device hostname -> list of BD names
                "view_mode": "combined"
            }
        }
        
        # Collect unique devices AND their BDs
        devices_seen = {}
        device_ids = {}
        device_idx = 0
        
        # First pass: collect all devices and track ALL their BDs
        for bd_name, bd_info in self.bd_paths.items():
            # Add BD to metadata list
            topology["metadata"]["bridge_domains"].append({
                "name": bd_name,
                "type": bd_info.bd_type,
                "vlan": bd_info.global_vlan,
                "color": bd_info.color
            })
            
            for hop in bd_info.path:
                device_name = hop['device']
                if device_name not in devices_seen:
                    devices_seen[device_name] = {
                        'type': hop['device_type'],
                        'bds': []
                    }
                # Add BD to device's list (avoid duplicates)
                if bd_name not in devices_seen[device_name]['bds']:
                    devices_seen[device_name]['bds'].append(bd_name)
        
        # Build device-to-BD mapping in metadata
        for device_name, device_data in devices_seen.items():
            topology["metadata"]["device_bd_mapping"][device_name] = {
                "bridge_domains": device_data['bds'],
                "bd_count": len(device_data['bds']),
                "device_type": device_data['type']
            }
        
        # Layout devices - WIDER spacing for clean visuals
        tier_heights = {
            "SUPERSPINE": 200,
            "SPINE": 500,
            "LEAF": 800,
            "PE": 1100,
            "TEST_EQUIPMENT": 1100,  # Same level as PE (Spirent, Ixia)
            "SERVER": 1100           # Same level as PE (hosts/servers)
        }
        
        tier_devices = {"SUPERSPINE": [], "SPINE": [], "LEAF": [], "PE": [], "TEST_EQUIPMENT": [], "SERVER": []}
        for device_name, device_data in devices_seen.items():
            tier = device_data['type']
            if tier not in tier_devices:
                tier = "PE"
            tier_devices[tier].append(device_name)
        
        # Calculate positions with ADAPTIVE spacing
        for tier, tier_devs in tier_devices.items():
            if not tier_devs:
                continue
            y = tier_heights.get(tier, 500)
            
            # ADAPTIVE: Wide spacing for full device names
            num_devs = len(tier_devs)
            if num_devs <= 3:
                spacing = 550
            elif num_devs <= 6:
                spacing = 480
            elif num_devs <= 10:
                spacing = 400
            else:
                spacing = 320  # Still wide for many devices
            
            # Center around a wider canvas (x=1200)
            total_width = (num_devs - 1) * spacing
            start_x = 1200 - total_width / 2
            
            for i, device_name in enumerate(sorted(tier_devs)):
                device_id = f"device_{device_idx}"
                device_ids[device_name] = device_id
                
                color = {
                    "SUPERSPINE": "#c0392b",
                    "SPINE": "#9b59b6",
                    "LEAF": "#e67e22",
                    "PE": "#3498db",
                    "TEST_EQUIPMENT": "#16a085",  # Teal for Spirent/Ixia
                    "SERVER": "#2ecc71"           # Green for servers
                }.get(tier, "#3498db")
                
                # Server style for DNAAS fabric AND test equipment/servers
                visual_style = "server" if tier in ["SUPERSPINE", "SPINE", "LEAF", "TEST_EQUIPMENT", "SERVER"] else "classic"
                
                device_obj = {
                    "id": device_id,
                    "type": "device",
                    "deviceType": "router",
                    "x": start_x + i * spacing,
                    "y": y,
                    "radius": 50,
                    "rotation": 0,
                    "color": color,
                    "label": device_name,
                    "locked": False,
                    "visualStyle": visual_style
                }
                
                # Add SSH credentials for DNAAS devices
                is_dnaas = tier in ["SUPERSPINE", "SPINE", "LEAF"]
                if is_dnaas:
                    # DNAAS devices use different credentials
                    # Look up IP from DNAAS table
                    dnaas_info = self.dnaas_table.lookup(device_name) if self.dnaas_table else None
                    dnaas_ip = dnaas_info.get('ip', '') if dnaas_info else ''
                    device_obj["sshConfig"] = {
                        "host": dnaas_ip or device_name,  # IP preferred, fallback to hostname
                        "hostBackup": device_name,  # Hostname as backup
                        "user": DNAAS_USER,
                        "password": DNAAS_PASS
                    }
                else:
                    # PE/termination devices use dnroot credentials
                    # Look up from inventory to get IP or serial
                    pe_info = self.inventory.get_device_by_hostname(device_name) if hasattr(self.inventory, 'get_device_by_hostname') else None
                    if not pe_info:
                        pe_info = self.inventory.get_device(device_name)
                    
                    pe_ip = ''
                    pe_serial = ''
                    if pe_info:
                        pe_ip = pe_info.mgmt_ip or ''
                        pe_serial = pe_info.serial or ''
                        # For cluster devices, resolve to active NCC
                        active = _find_active_ncc(pe_serial, self.inventory.data.get("devices", {}))
                        if active:
                            active_info = self.inventory.data["devices"].get(active, {})
                            pe_serial = active
                            if not pe_ip:
                                pe_ip = active_info.get('mgmt_ip', '') or ''
                    
                    # ENHANCED: Extract serial from device name if it contains SN pattern
                    # Patterns like: R2-WK31D7VW00016, R1-CL48, YOR_PE-1, R3_WDY1B77700018
                    if not pe_serial:
                        # Check if name contains a serial number pattern
                        import re
                        sn_patterns = [
                            r'([A-Z]{2}\d{2}[A-Z0-9]{6,})',  # e.g., WK31D7VW00016
                            r'([A-Z]{3}\d[A-Z]\d{2}[A-Z0-9]+)',  # e.g., YBW1F8VJ00002P1
                            r'-([A-Z0-9]{8,})',  # After dash, 8+ chars
                            r'_([A-Z]{2,3}\d+[A-Z0-9]+\d{4,})',  # After underscore with numbers
                        ]
                        for pattern in sn_patterns:
                            match = re.search(pattern, device_name, re.IGNORECASE)
                            if match:
                                pe_serial = match.group(1) if match.lastindex else match.group(0)
                                print(f"      Extracted serial '{pe_serial}' from device name '{device_name}'")
                                break
                    
                    device_obj["sshConfig"] = {
                        "host": pe_ip or pe_serial or device_name,  # IP > Serial > Hostname
                        "hostBackup": pe_serial or device_name,  # Serial or hostname as backup
                        "user": "dnroot",
                        "password": "dnroot"
                    }
                    # Also store serial for discovery
                    if pe_serial:
                        device_obj["deviceSerial"] = pe_serial
                
                topology["objects"].append(device_obj)
                device_idx += 1
        
        # Create links grouped by BD
        link_idx = 0
        connections_by_pair = {}  # (dev1, dev2) -> list of connections
        
        for conn in self.discovered_connections:
            from_dev = conn['from']
            to_dev = conn['to']
            pair = (min(from_dev, to_dev), max(from_dev, to_dev))
            
            if pair not in connections_by_pair:
                connections_by_pair[pair] = []
            connections_by_pair[pair].append(conn)
        
        for (dev1, dev2), conns in connections_by_pair.items():
            dev1_id = device_ids.get(dev1)
            dev2_id = device_ids.get(dev2)
            
            if not dev1_id or not dev2_id:
                continue
            
            # ENHANCED: Create SEPARATE links for each BD with distinct colors
            # This allows visualizing all BDs between the same device pair
            bd_links_for_pair = {}  # Group connections by BD name
            
            for conn in conns:
                bd_name = conn.get('bd_name', '')
                if bd_name not in bd_links_for_pair:
                    bd_links_for_pair[bd_name] = []
                bd_links_for_pair[bd_name].append(conn)
            
            # Create a separate link for EACH BD
            for bd_idx, (bd_name, bd_conns) in enumerate(bd_links_for_pair.items()):
                bd_info = self.bd_paths.get(bd_name)
                
                # ALL links use BD color (no more gray infrastructure links)
                link_color = bd_info.color if bd_info else "#2c3e50"
                
                # Use first connection for interface info
                first_conn = bd_conns[0]
                
                vlan_val = str(bd_info.global_vlan) if bd_info and bd_info.global_vlan else ""
                from_if = first_conn.get('from_if', '')
                to_if = first_conn.get('to_if', '')
                
                # Determine interface types (bundle, sub-interface, physical)
                bundleA = ""
                bundleB = ""
                subIfA = ""
                subIfB = ""
                physicalA = from_if
                physicalB = to_if
                
                # Parse interface names to extract bundle/sub-interface info
                if 'bundle' in from_if.lower():
                    bundleA = from_if.split('.')[0]
                    if '.' in from_if:
                        subIfA = from_if
                if 'bundle' in to_if.lower():
                    bundleB = to_if.split('.')[0]
                    if '.' in to_if:
                        subIfB = to_if
                elif '.' in from_if:
                    subIfA = from_if
                    physicalA = from_if.split('.')[0]
                if '.' in to_if:
                    subIfB = to_if
                    physicalB = to_if.split('.')[0]
                
                # Get cached interface details from discovery if available
                if_details_a = first_conn.get('interface_details_a', {})
                if_details_b = first_conn.get('interface_details_b', {})
                
                # Description includes BD name
                link_description = f"{dev1} ↔ {dev2} ({bd_name})"
                
                # Check if this is a termination link (from connection data)
                is_termination = first_conn.get('is_termination', False)
                vlan_validated = first_conn.get('vlan_validated', False)
                
                link_details = {
                    "interfaceA": from_if,
                    "interfaceB": to_if,
                    "physicalInterfaceA": physicalA,
                    "physicalInterfaceB": physicalB,
                    "bundleA": bundleA,
                    "bundleB": bundleB,
                    "description": link_description,
                    "bd_name": bd_name,
                    "bd_type": bd_info.bd_type if bd_info else "",
                    "vlan_id": bd_info.global_vlan if bd_info else None,
                    "is_termination": is_termination,
                    "vlan_validated": vlan_validated,
                    # Fields for Link Table auto-fill (frontend expects these)
                    "vlanIdA": vlan_val,
                    "vlanIdB": vlan_val,
                    "subInterfaceA": subIfA or (f"{physicalA}.{vlan_val}" if vlan_val else ""),
                    "subInterfaceB": subIfB or (f"{physicalB}.{vlan_val}" if vlan_val else ""),
                    # VLAN manipulation (from BD config)
                    "encapsulationA": f"dot1q {vlan_val}" if vlan_val else "",
                    "encapsulationB": f"dot1q {vlan_val}" if vlan_val else "",
                    # Interface state/transceiver from discovery
                    "adminStateA": if_details_a.get('admin_state', ''),
                    "adminStateB": if_details_b.get('admin_state', ''),
                    "operStateA": if_details_a.get('oper_state', ''),
                    "operStateB": if_details_b.get('oper_state', ''),
                    "speedA": if_details_a.get('speed', ''),
                    "speedB": if_details_b.get('speed', ''),
                    "mtuA": if_details_a.get('mtu', ''),
                    "mtuB": if_details_b.get('mtu', ''),
                    "transceiverA": if_details_a.get('transceiver', ''),
                    "transceiverB": if_details_b.get('transceiver', ''),
                    "fecA": if_details_a.get('fec', ''),
                    "fecB": if_details_b.get('fec', ''),
                    "l2ServiceA": if_details_a.get('l2_service', False),
                    "l2ServiceB": if_details_b.get('l2_service', False),
                    "bridgeDomains": [{
                        "bd_name": bd_name,
                        "bd_type": bd_info.bd_type if bd_info else "",
                        "vlan_id": bd_info.global_vlan if bd_info else None
                    }]
                }
                
                # Calculate curve offset for multiple links between same pair
                # This prevents links from overlapping visually
                curve_offset = 0
                if len(bd_links_for_pair) > 1:
                    curve_offset = (bd_idx - (len(bd_links_for_pair) - 1) / 2) * 40
                
                # Check if this is an infrastructure link (between DNaaS devices only)
                first_conn = bd_conns[0]
                is_infra_link = first_conn.get('is_infra', False)
                
                # Infrastructure links show BD name, PE links are solid
                link_label = bd_name
                link_style = "dashed" if is_infra_link else "solid"
                link_width = 2 if is_infra_link else 3
                
                topology["objects"].append({
                    "id": f"link_{link_idx}",
                    "type": "link",
                    "originType": "MULTI-BD",
                    "device1": dev1_id,
                    "device2": dev2_id,
                    "color": link_color,
                    "style": link_style,
                    "width": link_width,
                    "curveOverride": len(bd_links_for_pair) > 1,  # Enable curve for multiple links
                    "curveMagnitude": abs(curve_offset) if curve_offset != 0 else 0,
                    "curveDirection": 1 if curve_offset >= 0 else -1,
                    "linkDetails": link_details,
                    "label": link_label
                })
                
                link_idx += 1
        
        # NOTE: Text boxes (TBs) are NO LONGER generated here
        # BD info is displayed in the Bridge Domains panel in the UI
        # This avoids overlap with devices and keeps the canvas clean
        
        # Enhance BD metadata with path details
        for bd_name, bd_info in self.bd_paths.items():
            for bd_entry in topology["metadata"]["bridge_domains"]:
                if bd_entry.get("name") == bd_name:
                    bd_entry["path_devices"] = [h['device'] for h in bd_info.path]
                    bd_entry["interfaces"] = bd_info.interfaces
                    bd_entry["dut_subinterface"] = bd_info.dut_subinterface
                    break
        
        return topology


# ============================================================================
# OUTPUT GENERATORS
# ============================================================================

class OutputGenerator:
    """Generates output in various formats with Bridge Domain support"""
    
    def __init__(self, output_dir: Path = OUTPUT_DIR):
        self.output_dir = output_dir
        self.output_dir.mkdir(exist_ok=True)
    
    def generate_all(self, path: DNAASPath, devices: List[DeviceInfo], 
                     bd_summary: Dict[str, Any] = None,
                     discovered_connections: List[Dict[str, str]] = None):
        """Generate all output formats including BD info"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"dnaas_path_{timestamp}"
        
        # Generate JSON with connection data
        json_path = self.generate_json(path, devices, base_name, bd_summary, discovered_connections)
        
        # Generate text report
        txt_path = self.generate_text_report(path, devices, base_name, bd_summary)
        
        # Generate Excel (if available)
        xlsx_path = None
        if EXCEL_AVAILABLE:
            xlsx_path = self.generate_excel(path, devices, base_name, bd_summary)
        
        # Print terminal summary
        self.print_terminal_summary(path, devices, bd_summary)
        
        return json_path, txt_path, xlsx_path
    
    def generate_json(self, path: DNAASPath, devices: List[DeviceInfo], 
                      base_name: str, bd_summary: Dict[str, Any] = None,
                      discovered_connections: List[Dict[str, str]] = None) -> Path:
        """Generate Topology Creator compatible JSON with BD info"""
        filepath = self.output_dir / f"{base_name}.json"
        
        # Build topology structure
        topology = {
            "version": "1.0",
            "objects": [],
            "metadata": {
                "name": f"DNAAS Path Discovery",
                "created": datetime.now().isoformat(),
                "source": path.source_device,
                "destination": path.destination_device
            }
        }
        
        # Add BD summary to metadata if available
        normalized_bds = []
        if bd_summary:
            # Transform bd_path to normalized format expected by topology.js showBDLegend
            raw_bds = bd_summary.get("bd_path", [])
            normalized_bds = []
            seen_bd_names = set()
            for bd_item in raw_bds:
                bd_name = bd_item.get('bd_name', '')
                if not bd_name or bd_name in seen_bd_names:
                    continue
                seen_bd_names.add(bd_name)
                
                # Classify BD type from name
                bd_type, global_vlan = classify_bd_type(bd_name)
                
                # Use vlan_id if present, else use extracted global_vlan
                vlan = bd_item.get('vlan_id') or global_vlan
                
                normalized_bds.append({
                    "name": bd_name,
                    "type": bd_type,
                    "vlan": vlan,
                    "color": BD_COLOR_PALETTE[len(normalized_bds) % len(BD_COLOR_PALETTE)],
                    # Keep original data for link details
                    "device": bd_item.get('device'),
                    "interfaces": bd_item.get('interfaces', [])
                })
            
            topology["metadata"]["bridge_domains"] = normalized_bds
            topology["metadata"]["bundles"] = bd_summary.get("bundles_discovered", {})
            topology["metadata"]["config_changes"] = bd_summary.get("config_changes", [])
        
        # Classify device type including SuperSpine, test equipment, servers
        def classify_device_tier(name: str, dev_type: str) -> str:
            name_upper = name.upper()
            
            # Use device_type if it's a known special type
            if dev_type in ['TEST_EQUIPMENT', 'SERVER']:
                return dev_type
            
            # Test equipment detection (Spirent, Ixia, etc.)
            if any(test in name_upper for test in ['SPIRENT', 'IXIA', 'KEYSIGHT', 'TESTER', 'TRAFFIC']):
                return 'TEST_EQUIPMENT'
            
            # Server/Host detection
            if any(srv in name_upper for srv in ['SERVER', 'HOST', 'VM', 'WORKSTATION']):
                return 'SERVER'
            
            # DNAAS fabric tiers
            if 'SUPERSPINE' in name_upper:
                return 'SUPERSPINE'
            elif 'SPINE' in name_upper:
                return 'SPINE'
            elif 'LEAF' in name_upper:
                return 'LEAF'
            # Check if it's a DNAAS device we missed
            if any(kw in name_upper for kw in ['DNAAS', 'TOR', 'AGGREGATION']):
                return 'LEAF'  # Default DNAAS to LEAF tier
            return 'PE'  # Everything else is PE/edge
        
        # ============================================================
        # SPINE-LEAF TOPOLOGY LAYOUT
        # ============================================================
        # Visual layout (Y positions - smaller = higher on screen):
        #   
        #              [SuperSpine]           <- Y=80
        #            /      |      \
        #     [Spine]   [Spine]   [Spine]     <- Y=200  
        #        |  \   /  |  \   /  |
        #     [Leaf] [Leaf] [Leaf] [Leaf]     <- Y=350
        #       |      |      |      |
        #     [PE]   [PE]   [PE]   [PE]       <- Y=500
        #
        # ============================================================
        
        # CLEAN HIERARCHICAL HEIGHTS - comfortable vertical spacing
        tier_heights = {
            "SUPERSPINE": 250,      # Very top
            "SPINE": 600,           # Below SuperSpine (350px spacing)
            "LEAF": 950,            # Below Spine (350px spacing)
            "PE": 1300,             # Bottom (edge devices, 350px spacing)
            "TEST_EQUIPMENT": 1300, # Same level as PE (Spirent, Ixia)
            "SERVER": 1300          # Same level as PE (hosts/servers)
        }
        
        # First pass: group devices by tier (including test equipment and servers)
        tier_devices = {"SUPERSPINE": [], "SPINE": [], "LEAF": [], "PE": [], "TEST_EQUIPMENT": [], "SERVER": []}
        seen_devices = set()  # Stores device names
        seen_base_names = set()  # Stores base names (without prefix) for duplicate detection
        
        def get_base_name(name: str) -> str:
            """Extract base hostname, stripping common prefixes like YOR_, LAB_, etc."""
            return re.sub(r'^[A-Z]+_', '', name.upper())
        
        for hop in path.hops:
            # Check for exact match
            if hop.device_name in seen_devices:
                continue
            
            # ENHANCED: Check for base name match (handles YOR_PE-1 vs PE-1 duplicates)
            base_name = get_base_name(hop.device_name)
            if base_name in seen_base_names:
                print(f"  [OutputGenerator] Skipping duplicate device: {hop.device_name} (base: {base_name})")
                continue
            
            seen_devices.add(hop.device_name)
            seen_base_names.add(base_name)
            tier = classify_device_tier(hop.device_name, hop.device_type)
            tier_devices[tier].append(hop)
        
        # ════════════════════════════════════════════════════════════════════
        # RE-CATEGORIZE devices based on connections (interop routers to SS tier)
        # ════════════════════════════════════════════════════════════════════
        if discovered_connections:
            # Build sets of device names by tier
            ss_names = {d.device_name for d in tier_devices["SUPERSPINE"]}
            spine_names = {d.device_name for d in tier_devices["SPINE"]}
            leaf_names = {d.device_name for d in tier_devices["LEAF"]}
            
            # Helper to get connection profile
            def get_connections(device_name):
                connects_to_ss = False
                connects_to_spine = False
                connects_to_leaf = False
                for conn in discovered_connections:
                    from_dev = conn.get("from", "")
                    to_dev = conn.get("to", "")
                    if device_name in (from_dev, to_dev):
                        other = to_dev if from_dev == device_name else from_dev
                        if other in ss_names:
                            connects_to_ss = True
                        if other in spine_names:
                            connects_to_spine = True
                        if other in leaf_names:
                            connects_to_leaf = True
                return connects_to_ss, connects_to_spine, connects_to_leaf
            
            # Reclassify PE devices that connect to SuperSpine
            pe_to_reclassify = tier_devices["PE"][:]
            tier_devices["PE"] = []
            
            for hop in pe_to_reclassify:
                connects_to_ss, connects_to_spine, connects_to_leaf = get_connections(hop.device_name)
                
                if connects_to_ss and not connects_to_leaf and not connects_to_spine:
                    # Device connects ONLY to SuperSpine = place at SS tier
                    tier_devices["SUPERSPINE"].append(hop)
                    print(f"  [Layout] Promoted {hop.device_name} to SuperSpine tier (connects only to SS)")
                elif connects_to_ss and not connects_to_leaf:
                    # Device connects to SS but not LEAF = SPINE tier
                    tier_devices["SPINE"].append(hop)
                    print(f"  [Layout] Promoted {hop.device_name} to Spine tier (connects to SS, not LEAF)")
                else:
                    tier_devices["PE"].append(hop)
        
        # Count devices per tier for SYMMETRICAL layout
        tier_counts = {tier: len(devs) for tier, devs in tier_devices.items()}
        max_tier_count = max(tier_counts.values()) if tier_counts.values() else 1
        
        # BASE SPACING - comfortable for device names
        base_spacing = 320
        
        # Calculate GLOBAL width based on widest tier (for symmetry)
        global_width = max_tier_count * base_spacing
        canvas_width = max(2000, global_width + 600)
        canvas_center_x = canvas_width / 2
        
        # ADAPTIVE spacing function - ensures symmetry across tiers
        def get_tier_spacing(tier_count):
            if tier_count <= 1:
                return 0
            elif tier_count <= 3:
                return base_spacing * 1.2  # Extra wide for small groups
            elif tier_count < max_tier_count:
                # Scale so smaller tiers span similar width to largest
                return min(base_spacing * 1.1, global_width / (tier_count - 1))
            else:
                return base_spacing
        
        device_ids = {}
        device_positions = {}
        device_idx = 0
        
        # Build adjacency map from connections for intelligent placement
        adjacency = {}  # device -> list of connected devices
        if discovered_connections:
            for conn in discovered_connections:
                from_dev = conn.get("from", "")
                to_dev = conn.get("to", "")
                if from_dev:
                    adjacency.setdefault(from_dev, []).append(to_dev)
                if to_dev:
                    adjacency.setdefault(to_dev, []).append(from_dev)
        
        # Position devices tier by tier - SYMMETRICAL layout
        for tier in ["SUPERSPINE", "SPINE", "LEAF", "PE", "TEST_EQUIPMENT", "SERVER"]:
            devices_in_tier = tier_devices[tier]
            if not devices_in_tier:
                continue
            
            tier_count = len(devices_in_tier)
            spacing = get_tier_spacing(tier_count)
            y_pos = tier_heights.get(tier, 300)
            
            # For PE tier - sort by connected LEAF position for cleaner links
            if tier == "PE":
                # Sort PEs by their connected LEAF x position (for logical ordering)
                leaf_positions = {dev.device_name: device_positions.get(dev.device_name, {}).get("x", canvas_center_x) 
                                  for dev in tier_devices.get("LEAF", [])}
                
                pe_with_leaf_x = []
                for hop in devices_in_tier:
                    # Find which LEAF this PE connects to
                    connected_leafs = [dev for dev in adjacency.get(hop.device_name, []) 
                                       if classify_device_tier(dev, "") == "LEAF"]
                    if connected_leafs:
                        avg_leaf_x = sum(leaf_positions.get(l, canvas_center_x) for l in connected_leafs) / len(connected_leafs)
                    else:
                        avg_leaf_x = canvas_center_x
                    pe_with_leaf_x.append((hop, avg_leaf_x))
                
                # Sort by connected LEAF position
                pe_with_leaf_x.sort(key=lambda x: x[1])
                
                # SYMMETRICAL: Place PEs centered with adaptive spacing
                total_pes = len(pe_with_leaf_x)
                pe_tier_width = (total_pes - 1) * spacing
                pe_start_x = canvas_center_x - pe_tier_width / 2
                
                for i, (hop, _) in enumerate(pe_with_leaf_x):
                    x_pos = pe_start_x + i * spacing
                    
                    device_id = f"device_{device_idx}"
                    device_ids[hop.device_name] = device_id
                    device_positions[hop.device_name] = {"x": x_pos, "y": y_pos}
                    
                    # Build device object with SSH config for terminal access
                    device_obj = {
                        "id": device_id,
                        "type": "device",
                        "deviceType": "router",
                        "x": x_pos,
                        "y": y_pos,
                        "radius": 45,
                        "rotation": 0,
                        "color": "#3498db",
                        "label": hop.device_name,
                        "locked": False,
                        "visualStyle": "classic",
                        "deviceSerial": hop.device_serial or ""
                    }
                    # Add SSH config for PE/termination devices
                    # PE devices: use mgmt IP as primary, SN as backup
                    if hop.device_ip or hop.device_serial:
                        device_obj["sshConfig"] = {
                            "host": hop.device_ip or "",
                            "hostBackup": hop.device_serial or "",
                            "user": "dnroot",
                            "password": "dnroot"
                        }
                    topology["objects"].append(device_obj)
                    device_idx += 1
                continue
            
            # For other tiers, center and distribute evenly
            total_width = (len(devices_in_tier) - 1) * spacing
            start_x = canvas_center_x - total_width / 2
            
            for i, hop in enumerate(devices_in_tier):
                device_id = f"device_{device_idx}"
                device_ids[hop.device_name] = device_id
                
                x_pos = start_x + i * spacing
                device_positions[hop.device_name] = {"x": x_pos, "y": y_pos}
                
                # Determine visual style and color based on tier/device type
                if tier == "SUPERSPINE":
                    color = "#c0392b"  # RED for SuperSpine - stands out at top
                    style = "server"
                    radius = 55
                elif tier == "SPINE":
                    color = "#9b59b6"  # Purple
                    style = "server"
                    radius = 50
                elif tier == "LEAF":
                    color = "#e67e22"  # Orange
                    style = "server"
                    radius = 50
                elif tier == "TEST_EQUIPMENT":
                    color = "#16a085"  # Teal for Spirent/Ixia
                    style = "server"
                    radius = 50
                elif tier == "SERVER":
                    color = "#2ecc71"  # Green for servers/hosts
                    style = "server"
                    radius = 45
                else:
                    color = "#3498db"  # Blue for PE/other
                    style = "classic"
                    radius = 45
                
                # Build device object with SSH config for terminal access
                device_obj = {
                    "id": device_id,
                    "type": "device",
                    "deviceType": "router",
                    "x": x_pos,
                    "y": y_pos,
                    "radius": radius,
                    "rotation": 0,
                    "color": color,
                    "label": hop.device_name,
                    "locked": False,
                    "visualStyle": style,
                    "deviceSerial": hop.device_serial or ""
                }
                # Add SSH config based on device type
                # Determine if it's a DNAAS fabric device or interop/PE device
                device_name_upper = hop.device_name.upper()
                is_dnaas_fabric = any(kw in device_name_upper for kw in ['DNAAS', 'LEAF', 'SPINE'])
                is_interop_at_ss = (tier == "SUPERSPINE" and not is_dnaas_fabric)
                
                if is_dnaas_fabric:
                    # DNAAS devices: use hostname as primary (resolvable via DNS)
                    # Use DNAAS infrastructure credentials
                    device_obj["sshConfig"] = {
                        "host": hop.device_name,  # Hostname for DNAAS devices
                        "hostBackup": hop.device_ip or "",  # IP as backup
                        "user": "sisaev",
                        "password": "Drive1234!"
                    }
                elif is_interop_at_ss:
                    # Interop devices at SuperSpine tier (non-DNAAS routers connecting to SS)
                    # Use interop credentials: dn/drive1234!
                    device_obj["sshConfig"] = {
                        "host": hop.device_name,  # Try hostname first
                        "hostBackup": hop.device_ip or "",  # IP as backup
                        "user": "dn",
                        "password": "drive1234!"
                    }
                elif hop.device_ip or hop.device_serial:
                    # PE/DUT devices: use mgmt IP, dnroot credentials
                    device_obj["sshConfig"] = {
                        "host": hop.device_ip or "",
                        "hostBackup": hop.device_serial or "",
                        "user": "dnroot",
                        "password": "dnroot"
                    }
                topology["objects"].append(device_obj)
                device_idx += 1
        
        # Add links using discovered connections (or fall back to consecutive hops)
        connections_to_use = discovered_connections or []
        
        # If no connections provided, fall back to consecutive hops
        if not connections_to_use:
            for i in range(len(path.hops) - 1):
                hop1 = path.hops[i]
                hop2 = path.hops[i + 1]
                connections_to_use.append({
                    "from": hop1.device_name,
                    "to": hop2.device_name,
                    "from_if": hop1.egress_interface or "",
                    "to_if": hop2.ingress_interface or ""
                })
        
        # Deduplicate connections (avoid duplicate links)
        seen_links = set()
        unique_connections = []
        for conn in connections_to_use:
            link_key = tuple(sorted([conn["from"], conn["to"]]))
            if link_key not in seen_links:
                seen_links.add(link_key)
                unique_connections.append(conn)
        
        # Build BD color map directly from the normalized metadata BDs
        # so link colors are guaranteed to match the BD panel colors
        bd_color_map = {}
        for bd_meta in normalized_bds:
            bd_color_map[bd_meta["name"]] = bd_meta["color"]
        
        for i, conn in enumerate(unique_connections):
            from_device = conn["from"]
            to_device = conn["to"]
            from_if = conn.get("from_if", "")
            to_if = conn.get("to_if", "")
            
            # Skip if either device doesn't have an ID (not in topology)
            if from_device not in device_ids or to_device not in device_ids:
                continue
            
            link_id = f"link_{i}"
            
            # Get interface details for the link
            link_details = {
                "interfaceA": from_if,
                "interfaceB": to_if,
                "description": f"{from_device} to {to_device}"
            }
            
            # Add sub-interface and VLAN info if available from devices
            for device in devices:
                if device.hostname == from_device:
                    iface_details = device.interface_details.get(from_if, {})
                    if iface_details:
                        link_details["speedA"] = iface_details.get('speed', '')
                        link_details["bundleA"] = iface_details.get('bundle', '')
                        sub_ifaces = iface_details.get('sub_interfaces', [])
                        if sub_ifaces:
                            link_details["subInterfacesA"] = [s['name'] for s in sub_ifaces[:5]]
                            link_details["vlansA"] = [s['vlan'] for s in sub_ifaces if s.get('vlan')][:5]
            
            # Add Bridge Domain info and interface configs from bd_summary
            if bd_summary and bd_summary.get("bd_path"):
                bd_info_for_link = []
                interface_configs_a = {}
                interface_configs_b = {}
                
                for bd_entry in bd_summary["bd_path"]:
                    # Check if this BD is relevant to either device
                    device_name = bd_entry.get("device")
                    if device_name in [from_device, to_device]:
                        # Extract interface configs for this link
                        if bd_entry.get("interface_configs"):
                            for iface_name, iface_config in bd_entry["interface_configs"].items():
                                # Match interface to from_if or to_if
                                if from_if and (iface_name == from_if or iface_name.startswith(from_if.split('.')[0])):
                                    interface_configs_a[iface_name] = iface_config
                                if to_if and (iface_name == to_if or iface_name.startswith(to_if.split('.')[0])):
                                    interface_configs_b[iface_name] = iface_config
                    
                    if device_name in [from_device, to_device]:
                        bd_info_for_link.append({
                            "bd_id": bd_entry.get("bd_id"),
                            "bd_name": bd_entry.get("bd_name", ""),
                            "vlan_id": bd_entry.get("vlan_id"),
                            "state": bd_entry.get("state", "")
                        })
                if bd_info_for_link:
                    link_details["bridgeDomains"] = bd_info_for_link
                
                # Add interface configuration details (VLAN manipulation, state, etc.)
                if interface_configs_a:
                    # Get the main interface config
                    main_config = interface_configs_a.get(from_if) or next(iter(interface_configs_a.values()), {})
                    if main_config:
                        if main_config.get('vlan_id'):
                            link_details["vlanIdA"] = main_config['vlan_id']
                        if main_config.get('inner_vlan'):
                            link_details["innerVlanA"] = main_config['inner_vlan']
                        if main_config.get('encapsulation'):
                            link_details["encapsulationA"] = main_config['encapsulation']
                        if main_config.get('rewrite_ingress'):
                            link_details["rewriteIngressA"] = main_config['rewrite_ingress']
                        if main_config.get('rewrite_egress'):
                            link_details["rewriteEgressA"] = main_config['rewrite_egress']
                        if main_config.get('admin_state'):
                            link_details["adminStateA"] = main_config['admin_state']
                        if main_config.get('oper_state'):
                            link_details["operStateA"] = main_config['oper_state']
                        if main_config.get('l2_service'):
                            link_details["l2ServiceA"] = True
                    link_details["interfaceConfigsA"] = interface_configs_a
                
                if interface_configs_b:
                    main_config = interface_configs_b.get(to_if) or next(iter(interface_configs_b.values()), {})
                    if main_config:
                        if main_config.get('vlan_id'):
                            link_details["vlanIdB"] = main_config['vlan_id']
                        if main_config.get('inner_vlan'):
                            link_details["innerVlanB"] = main_config['inner_vlan']
                        if main_config.get('encapsulation'):
                            link_details["encapsulationB"] = main_config['encapsulation']
                        if main_config.get('rewrite_ingress'):
                            link_details["rewriteIngressB"] = main_config['rewrite_ingress']
                        if main_config.get('rewrite_egress'):
                            link_details["rewriteEgressB"] = main_config['rewrite_egress']
                        if main_config.get('admin_state'):
                            link_details["adminStateB"] = main_config['admin_state']
                        if main_config.get('oper_state'):
                            link_details["operStateB"] = main_config['oper_state']
                        if main_config.get('l2_service'):
                            link_details["l2ServiceB"] = True
                    link_details["interfaceConfigsB"] = interface_configs_b
            
            # Get device positions for link start/end
            dev1 = next((o for o in topology["objects"] if o["id"] == device_ids.get(from_device)), None)
            dev2 = next((o for o in topology["objects"] if o["id"] == device_ids.get(to_device)), None)
            
            if dev1 and dev2:
                # Calculate better link endpoints based on relative positions
                dx = dev2["x"] - dev1["x"]
                dy = dev2["y"] - dev1["y"]
                
                # Adjust start/end based on direction
                if abs(dx) > abs(dy):
                    # More horizontal link
                    start = {"x": dev1["x"] + 55, "y": dev1["y"]}
                    end = {"x": dev2["x"] - 55, "y": dev2["y"]}
                else:
                    # More vertical link
                    if dy > 0:  # dev2 is below dev1
                        start = {"x": dev1["x"], "y": dev1["y"] + 50}
                        end = {"x": dev2["x"], "y": dev2["y"] - 50}
                    else:  # dev2 is above dev1
                        start = {"x": dev1["x"], "y": dev1["y"] - 50}
                        end = {"x": dev2["x"], "y": dev2["y"] + 50}
            else:
                start = {"x": 0, "y": 0}
                end = {"x": 0, "y": 0}
            
            # Determine link color based on BD
            link_color = "#2c3e50"  # Default
            if "bridgeDomains" in link_details and link_details["bridgeDomains"]:
                first_bd = link_details["bridgeDomains"][0]
                bd_name = first_bd.get("bd_name", "")
                if bd_name in bd_color_map:
                    link_color = bd_color_map[bd_name]
            
            topology["objects"].append({
                "id": link_id,
                "type": "link",
                "originType": "QL",
                "device1": device_ids.get(from_device, ""),
                "device2": device_ids.get(to_device, ""),
                "color": link_color,
                "start": start,
                "end": end,
                "style": "solid",
                "width": 3,
                "linkDetails": link_details
            })
        
        # NOTE: Text boxes (TBs) are NO LONGER generated here
        # BD info is displayed in the Bridge Domains panel in the UI
        # This avoids overlap with devices and keeps the canvas clean
        
        # Add BD info to metadata for legend display (grouped by VLAN)
        if vlan_color_map:
            topology["metadata"]["bridge_domains_colors"] = [
                {
                    "vlan": vlan,
                    "color": color,
                    "bd_names": vlan_bds.get(vlan, [])
                } for vlan, color in vlan_color_map.items()
            ]
            topology["metadata"]["view_mode"] = "combined"
        
        with open(filepath, 'w') as f:
            json.dump(topology, f, indent=2)
        
        print(f"{Colors.GREEN}JSON saved: {filepath}{Colors.ENDC}")
        return filepath
    
    def generate_text_report(self, path: DNAASPath, devices: List[DeviceInfo],
                             base_name: str, bd_summary: Dict[str, Any] = None) -> Path:
        """Generate detailed text report with BD info"""
        filepath = self.output_dir / f"{base_name}.txt"
        
        lines = []
        lines.append("=" * 70)
        lines.append("DNAAS PATH DISCOVERY REPORT (BD-AWARE)")
        lines.append("=" * 70)
        lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Source: {path.source_device}")
        lines.append(f"Destination: {path.destination_device or 'N/A'}")
        lines.append(f"Status: {'SUCCESS' if path.success else 'PARTIAL'}")
        lines.append("")
        
        # Device Details
        lines.append("-" * 70)
        lines.append("DEVICE INVENTORY")
        lines.append("-" * 70)
        
        for device in devices:
            lines.append(f"\n[{device.hostname or device.serial}]")
            lines.append(f"  Serial: {device.serial}")
            lines.append(f"  Management IP: {device.mgmt_ip}")
            lines.append(f"  Management IPv6: {device.mgmt_ipv6 or 'N/A'}")
            lines.append(f"  System Type: {device.system_type}")
            lines.append(f"  DNOS Version: {device.dnos_version}")
            lines.append(f"  DNAAS Interfaces: {', '.join(device.dnaas_interfaces) or 'None detected'}")
            lines.append(f"  Last Seen: {device.last_seen}")
        
        # Path Details
        lines.append("\n" + "-" * 70)
        lines.append("PATH TRACE")
        lines.append("-" * 70)
        
        for i, hop in enumerate(path.hops):
            arrow = ">>>" if i < len(path.hops) - 1 else ""
            lines.append(f"\n[Hop {i + 1}] {hop.device_name} ({hop.device_type})")
            lines.append(f"  Serial: {hop.device_serial or 'N/A'}")
            lines.append(f"  IP: {hop.device_ip or 'N/A'}")
            if hop.ingress_interface:
                lines.append(f"  Ingress: {hop.ingress_interface}")
            if hop.egress_interface:
                lines.append(f"  Egress: {hop.egress_interface}")
            if hop.vlan_id:
                lines.append(f"  VLAN: {hop.vlan_id}")
            if arrow:
                lines.append(f"    {arrow}")
        
        # LLDP Details
        lines.append("\n" + "-" * 70)
        lines.append("LLDP NEIGHBOR DETAILS")
        lines.append("-" * 70)
        
        for device in devices:
            lines.append(f"\n[{device.hostname or device.serial}] LLDP Neighbors:")
            for neighbor in device.lldp_neighbors:
                dnaas_tag = " [DNAAS]" if neighbor.get('is_dnaas', False) else ""
                lines.append(f"  {neighbor.get('local_interface', 'N/A'):20} -> "
                           f"{neighbor.get('neighbor_name', 'N/A')}:{neighbor.get('neighbor_interface', 'N/A')}"
                           f"{dnaas_tag}")
        
        # Bridge Domain Details (if available)
        if bd_summary and bd_summary.get("bd_path"):
            lines.append("\n" + "-" * 70)
            lines.append("BRIDGE DOMAINS DISCOVERED")
            lines.append("-" * 70)
            
            for bd_entry in bd_summary["bd_path"]:
                lines.append(f"\n[{bd_entry.get('device', 'Unknown')}] BD {bd_entry.get('bd_id', 'N/A')}")
                if bd_entry.get('bd_name'):
                    lines.append(f"  Name: {bd_entry['bd_name']}")
                if bd_entry.get('vlan_id'):
                    lines.append(f"  VLAN ID: {bd_entry['vlan_id']}")
                lines.append(f"  State: {bd_entry.get('state', 'unknown')}")
                if bd_entry.get('interfaces'):
                    lines.append(f"  Interfaces:")
                    for iface in bd_entry['interfaces']:
                        lines.append(f"    - {iface}")
            
            # LACP Bundles
            if bd_summary.get("bundles_discovered"):
                lines.append("\n" + "-" * 70)
                lines.append("LACP BUNDLES DISCOVERED")
                lines.append("-" * 70)
                
                for bundle_id, bundle_info in bd_summary["bundles_discovered"].items():
                    lines.append(f"\n[{bundle_id}]")
                    lines.append(f"  State: {bundle_info.get('state', 'unknown')}")
                    if bundle_info.get('members'):
                        lines.append(f"  Members:")
                        for member in bundle_info['members']:
                            lines.append(f"    - {member}")
            
            # Configuration Changes Made
            if bd_summary.get("config_changes"):
                lines.append("\n" + "-" * 70)
                lines.append("CONFIGURATION CHANGES MADE")
                lines.append("-" * 70)
                
                for change in bd_summary["config_changes"]:
                    lines.append(f"\n[{change.get('device', 'Unknown')}] at {change.get('timestamp', 'N/A')}")
                    if change.get('interfaces'):
                        lines.append(f"  Interfaces configured:")
                        for iface in change['interfaces']:
                            lines.append(f"    - {iface}")
        
        lines.append("\n" + "=" * 70)
        lines.append("END OF REPORT")
        lines.append("=" * 70)
        
        with open(filepath, 'w') as f:
            f.write('\n'.join(lines))
        
        print(f"{Colors.GREEN}Text report saved: {filepath}{Colors.ENDC}")
        return filepath
    
    def generate_excel(self, path: DNAASPath, devices: List[DeviceInfo],
                       base_name: str, bd_summary: Dict[str, Any] = None) -> Optional[Path]:
        """Generate Excel spreadsheet with BD info"""
        if not EXCEL_AVAILABLE:
            return None
        
        filepath = self.output_dir / f"{base_name}.xlsx"
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sheet 1: Path Overview
        ws1 = wb.active
        ws1.title = "Path Overview"
        
        headers = ["Hop", "Device", "Type", "Serial", "IP", "Ingress IF", "Egress IF", "VLAN"]
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row, hop in enumerate(path.hops, 2):
            ws1.cell(row=row, column=1, value=row - 1).border = border
            ws1.cell(row=row, column=2, value=hop.device_name).border = border
            ws1.cell(row=row, column=3, value=hop.device_type).border = border
            ws1.cell(row=row, column=4, value=hop.device_serial).border = border
            ws1.cell(row=row, column=5, value=hop.device_ip).border = border
            ws1.cell(row=row, column=6, value=hop.ingress_interface).border = border
            ws1.cell(row=row, column=7, value=hop.egress_interface).border = border
            ws1.cell(row=row, column=8, value=hop.vlan_id or "").border = border
        
        # Sheet 2: Device Details
        ws2 = wb.create_sheet("Device Details")
        
        headers2 = ["Serial", "Hostname", "Mgmt IP", "Mgmt IPv6", "System Type", 
                   "DNOS Version", "DNAAS Interfaces", "Last Seen"]
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row, device in enumerate(devices, 2):
            ws2.cell(row=row, column=1, value=device.serial).border = border
            ws2.cell(row=row, column=2, value=device.hostname).border = border
            ws2.cell(row=row, column=3, value=device.mgmt_ip).border = border
            ws2.cell(row=row, column=4, value=device.mgmt_ipv6).border = border
            ws2.cell(row=row, column=5, value=device.system_type).border = border
            ws2.cell(row=row, column=6, value=device.dnos_version).border = border
            ws2.cell(row=row, column=7, value=', '.join(device.dnaas_interfaces)).border = border
            ws2.cell(row=row, column=8, value=device.last_seen).border = border
        
        # Sheet 3: LLDP Neighbors
        ws3 = wb.create_sheet("LLDP Neighbors")
        
        headers3 = ["Device", "Local Interface", "Neighbor", "Neighbor Interface", "TTL", "Is DNAAS"]
        for col, header in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row = 2
        for device in devices:
            for neighbor in device.lldp_neighbors:
                ws3.cell(row=row, column=1, value=device.hostname or device.serial).border = border
                ws3.cell(row=row, column=2, value=neighbor.get('local_interface', '')).border = border
                ws3.cell(row=row, column=3, value=neighbor.get('neighbor_name', '')).border = border
                ws3.cell(row=row, column=4, value=neighbor.get('neighbor_interface', '')).border = border
                ws3.cell(row=row, column=5, value=neighbor.get('ttl', 0)).border = border
                ws3.cell(row=row, column=6, value="Yes" if neighbor.get('is_dnaas', False) else "No").border = border
                row += 1
        
        # Sheet 4: Interface Details (Sub-interfaces, VLANs)
        ws4 = wb.create_sheet("Interface Details")
        
        headers4 = ["Device", "Interface", "Sub-Interface", "VLAN", "IPv4", "Service/VRF", "Admin", "Oper", "Speed", "Bundle"]
        for col, header in enumerate(headers4, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row = 2
        for device in devices:
            iface_details = getattr(device, 'interface_details', {})
            for iface_name, details in iface_details.items():
                # Main interface row
                ws4.cell(row=row, column=1, value=device.hostname or device.serial).border = border
                ws4.cell(row=row, column=2, value=iface_name).border = border
                ws4.cell(row=row, column=3, value="(physical)").border = border
                ws4.cell(row=row, column=4, value="").border = border
                ws4.cell(row=row, column=5, value="").border = border
                ws4.cell(row=row, column=6, value="").border = border
                ws4.cell(row=row, column=7, value="").border = border
                ws4.cell(row=row, column=8, value="").border = border
                ws4.cell(row=row, column=9, value=details.get('speed', '')).border = border
                ws4.cell(row=row, column=10, value=details.get('bundle', '')).border = border
                row += 1
                
                # Sub-interface rows
                for sub in details.get('sub_interfaces', []):
                    ws4.cell(row=row, column=1, value=device.hostname or device.serial).border = border
                    ws4.cell(row=row, column=2, value=iface_name).border = border
                    ws4.cell(row=row, column=3, value=sub.get('name', '')).border = border
                    ws4.cell(row=row, column=4, value=sub.get('vlan', '')).border = border
                    ws4.cell(row=row, column=5, value=sub.get('ipv4', '')).border = border
                    ws4.cell(row=row, column=6, value=sub.get('service', '')).border = border
                    ws4.cell(row=row, column=7, value=sub.get('admin', '')).border = border
                    ws4.cell(row=row, column=8, value=sub.get('oper', '')).border = border
                    ws4.cell(row=row, column=9, value="").border = border
                    ws4.cell(row=row, column=10, value="").border = border
                    row += 1
        
        # Adjust column widths
        for ws in [ws1, ws2, ws3, ws4]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        wb.save(filepath)
        print(f"{Colors.GREEN}Excel saved: {filepath}{Colors.ENDC}")
        return filepath
    
    def print_terminal_summary(self, path: DNAASPath, devices: List[DeviceInfo],
                               bd_summary: Dict[str, Any] = None):
        """Print color-coded terminal summary with BD info"""
        print("\n" + "=" * 70)
        print(f"{Colors.BOLD}{Colors.HEADER}DNAAS PATH DISCOVERY SUMMARY (BD-AWARE){Colors.ENDC}")
        print("=" * 70)
        
        status_color = Colors.GREEN if path.success else Colors.YELLOW
        status_text = 'SUCCESS' if path.success and not path.error_message else ('PARTIAL' if path.success else 'INCOMPLETE')
        print(f"\n{Colors.BOLD}Status:{Colors.ENDC} {status_color}{status_text}{Colors.ENDC}")
        print(f"{Colors.BOLD}Source:{Colors.ENDC} {path.source_device}")
        print(f"{Colors.BOLD}Destination:{Colors.ENDC} {path.destination_device or 'N/A'}")
        if path.error_message:
            print(f"{Colors.YELLOW}Note: {path.error_message}{Colors.ENDC}")
        
        print(f"\n{Colors.BOLD}{Colors.UNDERLINE}Discovered Devices:{Colors.ENDC}")
        for device in devices:
            dnaas_count = len(device.dnaas_interfaces)
            print(f"  {Colors.CYAN}{device.hostname or device.serial}{Colors.ENDC}")
            print(f"    IP: {device.mgmt_ip}")
            print(f"    DNAAS Interfaces: {Colors.GREEN}{dnaas_count}{Colors.ENDC}")
        
        print(f"\n{Colors.BOLD}{Colors.UNDERLINE}Path Trace:{Colors.ENDC}")
        for i, hop in enumerate(path.hops):
            type_colors = {"PE": Colors.BLUE, "LEAF": Colors.YELLOW, "SPINE": Colors.CYAN}
            type_color = type_colors.get(hop.device_type, Colors.ENDC)
            
            connector = "  |" if i < len(path.hops) - 1 else ""
            arrow = f"  |--- {hop.egress_interface} --->" if hop.egress_interface else ""
            
            print(f"  [{type_color}{hop.device_type}{Colors.ENDC}] {Colors.BOLD}{hop.device_name}{Colors.ENDC}")
            if hop.ingress_interface:
                print(f"        Ingress: {hop.ingress_interface}")
            if hop.egress_interface:
                print(f"        Egress: {hop.egress_interface}")
            if connector:
                print(f"        |")
                print(f"        v")
        
        # Bridge Domains summary
        if bd_summary and bd_summary.get("bd_path"):
            print(f"\n{Colors.BOLD}{Colors.UNDERLINE}Bridge Domains Discovered:{Colors.ENDC}")
            bd_ids = set()
            for bd_entry in bd_summary["bd_path"]:
                bd_id = bd_entry.get("bd_id")
                if bd_id not in bd_ids:
                    bd_ids.add(bd_id)
                    bd_name = bd_entry.get("bd_name", "")
                    device = bd_entry.get("device", "")
                    print(f"  {Colors.GREEN}BD {bd_id}{Colors.ENDC} {bd_name} on {device}")
            
            # LACP bundles
            if bd_summary.get("bundles_discovered"):
                print(f"\n{Colors.BOLD}{Colors.UNDERLINE}LACP Bundles:{Colors.ENDC}")
                for bundle_id, bundle_info in bd_summary["bundles_discovered"].items():
                    members = bundle_info.get("members", [])
                    print(f"  {Colors.CYAN}{bundle_id}{Colors.ENDC}: {len(members)} members")
        
        print("\n" + "=" * 70)

# ============================================================================
# MAIN DISCOVERY FLOW
# ============================================================================

def _try_db_lldp_lookup(serial: str) -> list:
    """Try to get LLDP neighbors from local DB sources (no SSH)."""
    neighbors = []
    search_lower = serial.lower()
    
    # Source 1: Scaler DB (operational.json)
    db_configs = Path('/home/dn/SCALER/db/configs')
    if db_configs.exists():
        for device_dir in db_configs.iterdir():
            if device_dir.is_dir():
                op_file = device_dir / 'operational.json'
                if op_file.exists():
                    try:
                        with open(op_file, 'r') as f:
                            op_data = json.load(f)
                        dev_hostname = op_data.get('hostname', device_dir.name)
                        dev_serial = op_data.get('serial_number', '')
                        dev_ip = op_data.get('connection_ip', '')
                        if dev_hostname.lower() == search_lower or \
                           search_lower in dev_hostname.lower() or \
                           dev_hostname.lower() in search_lower or \
                           (dev_serial and dev_serial.lower() == search_lower) or \
                           (dev_ip and dev_ip == serial):
                            lldp_data = op_data.get('lldp_neighbors', [])
                            if lldp_data:
                                for n in lldp_data:
                                    neigh = n.get('neighbor_name') or n.get('neighbor_device') or n.get('neighbor', '')
                                    if 'is_dnaas' not in n:
                                        n['is_dnaas'] = any(kw in neigh.upper() for kw in DNAAS_KEYWORDS)
                                print(f"  {Colors.CYAN}[DB] Found {len(lldp_data)} LLDP neighbors in Scaler DB for {dev_hostname}{Colors.ENDC}")
                                return lldp_data
                    except Exception:
                        pass
    
    # Source 2: NetworkMapper MCP
    try:
        sys.path.insert(0, '/home/dn/SCALER') if '/home/dn/SCALER' not in sys.path else None
        from scaler.network_mapper_client import NetworkMapperClient
        nm = NetworkMapperClient()
        nm_neighbors = nm.get_device_lldp(serial)
        if nm_neighbors:
            neighbors = [{
                'local_interface': n.local_interface,
                'neighbor_name': n.neighbor_name,
                'neighbor_interface': n.neighbor_interface,
                'is_dnaas': any(kw in n.neighbor_name.upper() for kw in DNAAS_KEYWORDS)
            } for n in nm_neighbors]
            print(f"  {Colors.CYAN}[DB] Found {len(neighbors)} LLDP neighbors via NetworkMapper{Colors.ENDC}")
            return neighbors
    except Exception as e:
        print(f"  {Colors.YELLOW}[DB] NetworkMapper lookup skipped: {e}{Colors.ENDC}")
    
    return neighbors


def _fuzzy_serial_match(input_serial: str, known_serial: str, threshold: float = 0.75) -> bool:
    """Check if input_serial fuzzy-matches known_serial (handles l/1, case, typos)."""
    from difflib import SequenceMatcher
    a = input_serial.upper().replace('L', '1')
    b = known_serial.upper().replace('L', '1')
    if a == b:
        return True
    ratio = SequenceMatcher(None, a, b).ratio()
    return ratio >= threshold


def _find_active_ncc(device_name: str, devices: dict) -> str:
    """Find the active NCC for a cluster device.
    
    Strategy 1: NCC serial prefix matching (e.g. kvm108-cl408d-ncc0 -> ncc1)
    Strategy 2: Hostname matching (e.g. YOR_CL_PE-4 -> best NCC serial)
    
    Prefers: clean hostname (no date suffix) > most recent last_seen.
    """
    import re as _re
    name_norm = device_name.split('(')[0].strip().lower()
    _date_re = _re.compile(r'\(\d{2}-\w{3}-\d{4}')
    
    def _sort_key(c):
        """Sort: no-date-suffix first, then most recent last_seen."""
        return (c[2], '')  # has_date=False sorts before True
    
    def _pick_best(candidates, current_entry):
        """Pick best candidate, comparing against current entry."""
        no_date = sorted([c for c in candidates if not c[2]], key=lambda c: c[1], reverse=True)
        with_date = sorted([c for c in candidates if c[2]], key=lambda c: c[1], reverse=True)
        ordered = no_date + with_date
        best = ordered[0]
        
        if not best[2]:
            print(f"  {Colors.CYAN}[Resolve] {device_name} -> {best[0]} (active NCC, clean hostname){Colors.ENDC}")
            return best[0]
        
        # No current entry = hostname lookup (Strategy 2), just pick most recent
        if not current_entry:
            print(f"  {Colors.CYAN}[Resolve] {device_name} -> {best[0]} (most recent NCC, last_seen={best[1]}){Colors.ENDC}")
            return best[0]
        
        # Both have date suffixes: pick more recently seen
        current_has_date = bool(_date_re.search(current_entry.get('hostname', '')))
        if current_has_date and best[1] > (current_entry.get('last_seen', '') or ''):
            print(f"  {Colors.CYAN}[Resolve] {device_name} -> {best[0]} (more recent NCC, last_seen={best[1]}){Colors.ENDC}")
            return best[0]
        return ''
    
    # Strategy 1: NCC serial prefix matching
    ncc_match = _re.match(r'^(.+)-ncc\d+$', name_norm, _re.IGNORECASE)
    if ncc_match:
        prefix = ncc_match.group(1)
        candidates = []
        for key, d in devices.items():
            if key.lower().startswith(prefix + '-ncc') and key.lower() != name_norm:
                h = d.get('hostname') or ''
                candidates.append((key, d.get('last_seen', ''), bool(_date_re.search(h))))
        if candidates:
            current = devices.get(device_name) or devices.get(device_name.upper()) or {}
            return _pick_best(candidates, current)
    
    # Strategy 2: hostname matching
    candidates = []
    for key, d in devices.items():
        if key.lower() == name_norm:
            continue
        h = (d.get('hostname') or '').split('(')[0].strip().lower()
        if h == name_norm:
            candidates.append((key, d.get('last_seen', ''), bool(_date_re.search(d.get('hostname', '')))))
    if not candidates:
        return ''
    return _pick_best(candidates, {})


def _resolve_ssh_target(serial: str, inventory: DeviceInventory) -> str:
    """Resolve a serial/hostname to an SSH-connectable target (IP or hostname).
    
    Checks: inventory -> Scaler DB -> credentials, with fuzzy serial matching.
    """
    search_lower = serial.lower()
    search_upper = serial.upper()
    
    # If already an IP address, return as-is
    import re
    if re.match(r'^\d+\.\d+\.\d+\.\d+$', serial):
        return serial
    
    # 1. Check device inventory (exact + fuzzy)
    devices = inventory.data.get("devices", {})
    matched_entry = None
    for key, d in devices.items():
        dev_hostname = (d.get('hostname') or '').lower()
        dev_serial = d.get('serial', key)
        mgmt_ip = d.get('mgmt_ip', '')
        
        exact = key.lower() == search_lower or dev_hostname == search_lower
        fuzzy = _fuzzy_serial_match(serial, dev_serial)
        
        if exact or fuzzy:
            matched_entry = (key, d)
            if mgmt_ip:
                print(f"  {Colors.CYAN}[Resolve] {serial} -> {mgmt_ip} (inventory){Colors.ENDC}")
                return mgmt_ip
            if re.match(r'^\d+\.\d+\.\d+\.\d+$', key):
                print(f"  {Colors.CYAN}[Resolve] {serial} -> {key} (inventory key=IP){Colors.ENDC}")
                return key
            break
    
    # 1b. If primary entry had no mgmt_ip, try NCC lookup (cluster devices)
    if matched_entry:
        ncc = _find_active_ncc(serial, devices)
        if ncc:
            return ncc
    
    # 2. Check Scaler DB operational.json (has mgmt_ip and ssh_host)
    db_configs = Path('/home/dn/SCALER/db/configs')
    if db_configs.exists():
        for device_dir in db_configs.iterdir():
            if device_dir.is_dir():
                op_file = device_dir / 'operational.json'
                if op_file.exists():
                    try:
                        with open(op_file, 'r') as f:
                            op_data = json.load(f)
                        dev_hostname = op_data.get('hostname', device_dir.name) or device_dir.name
                        dev_serial = (op_data.get('serial_number', '') or '')
                        dev_ip = op_data.get('connection_ip', '') or ''
                        dev_mgmt = (op_data.get('mgmt_ip', '') or '').split('/')[0]
                        
                        exact = (dev_serial.lower() == search_lower or
                                 device_dir.name.lower() == search_lower or
                                 search_lower in dev_hostname.lower() or
                                 dev_hostname.lower() in search_lower)
                        fuzzy = dev_serial and _fuzzy_serial_match(serial, dev_serial)
                        
                        if exact or fuzzy:
                            if dev_ip:
                                print(f"  {Colors.CYAN}[Resolve] {serial} -> {dev_ip} (Scaler DB connection_ip){Colors.ENDC}")
                                return dev_ip
                            if dev_mgmt and re.match(r'^\d+\.\d+\.\d+\.\d+$', dev_mgmt):
                                print(f"  {Colors.CYAN}[Resolve] {serial} -> {dev_mgmt} (Scaler DB mgmt_ip){Colors.ENDC}")
                                return dev_mgmt
                            ssh_host = (op_data.get('ssh_host', '') or '')
                            if ssh_host:
                                print(f"  {Colors.CYAN}[Resolve] {serial} -> {ssh_host} (Scaler DB ssh_host){Colors.ENDC}")
                                return ssh_host
                    except Exception:
                        pass
    
    # 3. Check credentials store
    creds = CredentialManager()
    cred_entry = creds.get(serial)
    if cred_entry:
        mgmt_ip = cred_entry.get('mgmt_ip', '')
        if mgmt_ip and mgmt_ip != serial:
            print(f"  {Colors.CYAN}[Resolve] {serial} -> {mgmt_ip} (credentials){Colors.ENDC}")
            return mgmt_ip
    
    # 4. NCC hostname fallback (picks most recently seen NCC)
    ncc = _find_active_ncc(serial, devices)
    if ncc:
        return ncc
    
    # 5. Try uppercase (DNOS serials resolve via DNS in uppercase)
    if search_upper != serial:
        print(f"  {Colors.YELLOW}[Resolve] Trying uppercase: {search_upper}{Colors.ENDC}")
        return search_upper
    
    return serial


def discover_device(serial: str, inventory: DeviceInventory, 
                    use_cache: bool = False,
                    ssh_user: str = DEFAULT_USER,
                    ssh_pass: str = DEFAULT_PASS) -> Optional[DeviceInfo]:
    """Discover a device by serial number. Tries local DB before SSH."""
    
    # Check cache first if requested
    if use_cache:
        cached = inventory.get_device(serial)
        if cached:
            print(f"{Colors.CYAN}Using cached data for {serial}{Colors.ENDC}")
            return cached
    
    print(f"\n{Colors.HEADER}[DISCOVERY] {serial}{Colors.ENDC}")
    print("-" * 50)
    
    # Try DB-first for LLDP data (avoids SSH when data is already cached)
    db_lldp = _try_db_lldp_lookup(serial)
    
    # Resolve serial to an SSH-connectable target (IP or resolvable hostname)
    ssh_target = _resolve_ssh_target(serial, inventory)
    
    # Connect to device
    ssh = SSHConnection(ssh_target, ssh_user, ssh_pass)
    if not ssh.connect():
        # If SSH fails but we have DB data, build a partial DeviceInfo
        if db_lldp:
            print(f"{Colors.YELLOW}SSH failed but using DB LLDP data ({len(db_lldp)} neighbors){Colors.ENDC}")
            device = DeviceInfo(
                serial=serial,
                hostname=serial,
                lldp_neighbors=db_lldp
            )
            return device
        print(f"{Colors.RED}Failed to connect to {serial}{Colors.ENDC}")
        return None
    
    try:
        discovery = DeviceDiscovery(ssh)
        
        # Get hostname
        print("  Getting hostname...", flush=True)
        hostname = discovery.get_hostname()
        print(f"    Hostname: {Colors.GREEN}{hostname or 'Unknown'}{Colors.ENDC}")
        
        # Get system info
        print("  Getting system info...", flush=True)
        sys_info = discovery.get_system_info()
        print(f"    System Type: {sys_info.get('system_type', 'Unknown')}")
        print(f"    DNOS Version: {sys_info.get('dnos_version', 'Unknown')}")
        
        # Get management interfaces
        print("  Getting management interfaces...", flush=True)
        mgmt_interfaces = discovery.get_management_interfaces()
        print(f"    Found {len(mgmt_interfaces)} management interfaces")
        
        mgmt_ip, mgmt_ipv6 = discovery.get_primary_mgmt_ip(mgmt_interfaces)
        print(f"    Primary IPv4: {Colors.GREEN}{mgmt_ip or 'None'}{Colors.ENDC}")
        print(f"    Primary IPv6: {mgmt_ipv6 or 'None'}")
        
        # Get LLDP neighbors (use DB data if already fetched, otherwise SSH)
        if db_lldp:
            print(f"  Using {len(db_lldp)} LLDP neighbors from DB (skipping SSH fetch)", flush=True)
            neighbors = []
            for n in db_lldp:
                local_if = n.get('local_interface') or n.get('interface', '')
                neigh_name = n.get('neighbor_name') or n.get('neighbor_device') or n.get('neighbor', '')
                neigh_if = n.get('neighbor_interface') or n.get('neighbor_port') or n.get('remote_port', '')
                is_dnaas = n.get('is_dnaas', any(kw in neigh_name.upper() for kw in DNAAS_KEYWORDS))
                neighbors.append(LLDPNeighbor(
                    local_interface=local_if,
                    neighbor_name=neigh_name,
                    neighbor_interface=neigh_if,
                    ttl=n.get('ttl', 120),
                    is_dnaas=is_dnaas
                ))
        else:
            print("  Getting LLDP neighbors...", flush=True)
            neighbors = discovery.get_lldp_neighbors()
        print(f"    Found {len(neighbors)} LLDP neighbors")
        
        # Get interface descriptions
        print("  Getting interface descriptions...", flush=True)
        descriptions = discovery.get_interface_descriptions()
        
        # Detect DNAAS interfaces
        print("  Detecting DNAAS interfaces...", flush=True)
        dnaas_interfaces = discovery.detect_dnaas_interfaces(neighbors, descriptions)
        print(f"    Found {Colors.GREEN}{len(dnaas_interfaces)}{Colors.ENDC} DNAAS interfaces")
        
        # Get detailed interface info for each DNAAS interface
        interface_details = {}
        for iface in dnaas_interfaces:
            print(f"    Getting details for {iface}...", flush=True)
            details = discovery.get_interface_details(iface)
            interface_details[iface] = details
            if details['sub_interfaces']:
                print(f"      Sub-interfaces: {len(details['sub_interfaces'])}")
                for sub in details['sub_interfaces'][:3]:  # Show first 3
                    vlan_info = f"VLAN {sub['vlan']}" if sub['vlan'] else ""
                    svc_info = sub['service'] if sub['service'] else ""
                    print(f"        {sub['name']}: {vlan_info} {svc_info}")
                if len(details['sub_interfaces']) > 3:
                    print(f"        ... and {len(details['sub_interfaces']) - 3} more")
        
        # Build device info
        device = DeviceInfo(
            serial=serial,
            hostname=hostname,
            mgmt_ip=mgmt_ip,
            mgmt_ipv6=mgmt_ipv6,
            system_type=sys_info.get('system_type', ''),
            dnos_version=sys_info.get('dnos_version', ''),
            dnaas_interfaces=dnaas_interfaces,
            lldp_neighbors=[asdict(n) for n in neighbors],
            all_mgmt_interfaces=[asdict(m) for m in mgmt_interfaces],
            interface_details=interface_details
        )
        
        # Save to inventory
        inventory.update_device(device)
        
        return device
        
    finally:
        ssh.close()

def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description="DNAAS Path Auto-Discovery with Bridge Domain Support",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 dnaas_path_discovery.py wk31d7vv00023
  python3 dnaas_path_discovery.py wk31d7vv00023 100.64.0.220
  python3 dnaas_path_discovery.py --use-cache wk31d7vv00023
  python3 dnaas_path_discovery.py --bd-aware wk31d7vv00023
  python3 dnaas_path_discovery.py --list-inventory
        """
    )
    
    parser.add_argument('serials', nargs='*', help='Device serial number(s) to discover')
    parser.add_argument('--use-cache', action='store_true', 
                        help='Use cached data if available')
    parser.add_argument('--list-inventory', action='store_true',
                        help='List all devices in inventory')
    parser.add_argument('--list-credentials', action='store_true',
                        help='List all stored device credentials')
    parser.add_argument('--bd-aware', action='store_true',
                        help='Use BD-aware tracing (auto-configures DUT interfaces)')
    parser.add_argument('--multi-bd', action='store_true',
                        help='Discover ALL Bridge Domains and create color-coded topology')
    parser.add_argument('--no-auto-config', action='store_true',
                        help='Skip auto-configuration of DUT interfaces (use with --bd-aware)')
    parser.add_argument('--user', default=DEFAULT_USER, help='SSH username')
    parser.add_argument('--password', default=DEFAULT_PASS, help='SSH password')
    
    args = parser.parse_args()
    
    # Initialize inventory
    inventory = DeviceInventory()
    
    # List inventory if requested
    if args.list_inventory:
        devices = inventory.list_devices()
        if devices:
            print(f"\n{Colors.HEADER}Device Inventory ({len(devices)} devices){Colors.ENDC}")
            print("-" * 50)
            for serial in devices:
                device = inventory.get_device(serial)
                if device:
                    print(f"  {Colors.CYAN}{serial}{Colors.ENDC}: {device.hostname or 'Unknown'} ({device.mgmt_ip})")
        else:
            print("No devices in inventory.")
        return
    
    # List credentials if requested
    if args.list_credentials:
        creds = credential_manager.list_devices()
        if creds:
            print(f"\n{Colors.HEADER}Stored Credentials ({len(creds)} devices){Colors.ENDC}")
            print("-" * 50)
            for cred in creds:
                device_type = cred.get('device_type', 'DUT')
                type_color = Colors.BLUE if device_type == 'DUT' else Colors.YELLOW
                print(f"  {type_color}[{device_type}]{Colors.ENDC} {cred['hostname']} ({cred['mgmt_ip']})")
                print(f"         Last used: {cred.get('last_used', 'Never')[:19]}")
        else:
            print("No stored credentials.")
        return
    
    # Check for serial numbers
    if not args.serials:
        parser.print_help()
        print(f"\n{Colors.RED}Error: At least one serial number is required{Colors.ENDC}")
        sys.exit(1)
    
    # Use provided credentials
    ssh_user = args.user
    ssh_pass = args.password
    
    # Handle Multi-BD discovery mode
    if args.multi_bd:
        print(f"\n{Colors.BOLD}{Colors.HEADER}MULTI-BD DNAAS DISCOVERY{Colors.ENDC}")
        print("=" * 70)
        print(f"Source device: {args.serials[0] if args.serials else 'None'}")
        print("=" * 70)
        
        if not args.serials:
            print(f"{Colors.RED}Error: Device serial/hostname required for multi-BD discovery{Colors.ENDC}")
            sys.exit(1)
        
        # Discover source device first
        source_serial = args.serials[0]
        source_device = discover_device(source_serial, inventory, args.use_cache, ssh_user, ssh_pass)
        
        if not source_device:
            print(f"\n{Colors.RED}Failed to connect to source device {source_serial}{Colors.ENDC}")
            sys.exit(1)
        
        # Check if we have LLDP neighbors - if not, fail early with clear message for UI
        if not source_device.lldp_neighbors:
            print(f"\n{Colors.RED}✗ Found 0 LLDP neighbors on {source_device.hostname or source_serial}{Colors.ENDC}")
            print(f"{Colors.RED}No LLDP neighbors detected - cannot discover DNAAS fabric{Colors.ENDC}")
            print(f"\n{Colors.YELLOW}Tip: LLDP must be enabled on the device interfaces{Colors.ENDC}")
            print(f"{Colors.YELLOW}Run: python3 dnaas_path_discovery.py --enable-lldp {source_serial}{Colors.ENDC}")
            sys.exit(1)  # Exit with error to trigger UI LLDP dialog
        
        # Check if any neighbors are DNAAS devices
        dnaas_neighbors = [n for n in source_device.lldp_neighbors if n.get('is_dnaas', False)]
        if not dnaas_neighbors:
            print(f"\n{Colors.RED}✗ Found 0 DNAAS neighbors on {source_device.hostname or source_serial}{Colors.ENDC}")
            print(f"{Colors.RED}Found {len(source_device.lldp_neighbors)} LLDP neighbors but none are DNAAS infrastructure{Colors.ENDC}")
            print(f"\n{Colors.YELLOW}LLDP neighbors found:{Colors.ENDC}")
            for n in source_device.lldp_neighbors[:5]:
                print(f"  {n.get('local_interface', 'N/A')} -> {n.get('neighbor_name', 'Unknown')}")
            print(f"\n{Colors.YELLOW}Tip: Device may not be connected to DNAAS fabric{Colors.ENDC}")
            sys.exit(1)
        
        print(f"\n{Colors.GREEN}Found {len(dnaas_neighbors)} DNAAS neighbors:{Colors.ENDC}")
        for n in dnaas_neighbors:
            print(f"  {n.get('local_interface', 'N/A')} -> {Colors.CYAN}{n.get('neighbor_name', 'Unknown')}{Colors.ENDC}")
        
        # Initialize Multi-BD tracer
        dnaas_table = DNAASTable()
        multi_bd_tracer = MultiBDPathTracer(inventory, dnaas_table)
        
        # Discover all Bridge Domains
        bd_paths = multi_bd_tracer.discover_all_bds(source_serial)
    
        if not bd_paths:
            print(f"\n{Colors.RED}No Bridge Domains discovered{Colors.ENDC}")
            sys.exit(1)
        
        # Generate Multi-BD topology
        topology = multi_bd_tracer.generate_multi_bd_topology(source_device)
        
        # Save output
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = OUTPUT_DIR / f"multi_bd_{timestamp}.json"
        
        with open(output_file, 'w') as f:
            json.dump(topology, f, indent=2, default=str)
        
        print(f"\n{Colors.GREEN}Multi-BD topology saved: {output_file}{Colors.ENDC}")
        print(f"Discovered {len(bd_paths)} Bridge Domains:")
        for bd_name, bd_info in bd_paths.items():
            print(f"  • {bd_name}: {bd_info.bd_type}, VLAN {bd_info.global_vlan or 'N/A'}, Color {bd_info.color}")


if __name__ == "__main__":
    main()
