"""
DNAAS Lookup Module

Looks up DNAAS device information from the Excel inventory file.
"""

import os
from typing import Optional
from openpyxl import load_workbook


class DNAASLookup:
    """Looks up DNAAS devices from Excel inventory"""
    
    DEFAULT_XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "dnaas_table.xlsx")
    
    def __init__(self, xlsx_path: str = None):
        self.xlsx_path = xlsx_path or self.DEFAULT_XLSX_PATH
        self.devices: list[dict] = []
        self._load_devices()
    
    def _load_devices(self) -> None:
        """Load devices from Excel file"""
        try:
            wb = load_workbook(self.xlsx_path)
            sheet = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            
            # Parse each row into a device dict
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None:  # Skip empty rows
                    continue
                    
                device = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        # Normalize header names
                        key = header.lower().replace(" ", "_").replace("-", "_")
                        device[key] = row[i]
                
                if device.get("name"):
                    self.devices.append(device)
            
            wb.close()
            
        except Exception as e:
            print(f"Warning: Could not load DNAAS table: {e}")
            self.devices = []
    
    def get_device_ip(self, device_name: str) -> Optional[str]:
        """Get IP address for a device by name (case-insensitive partial match)"""
        device_name_lower = device_name.lower().strip()
        
        for device in self.devices:
            name = str(device.get("name", "")).lower()
            if device_name_lower in name or name in device_name_lower:
                ip = device.get("ip_adress") or device.get("ip_address")
                if ip and ip != "TBD":
                    return str(ip)
        
        return None
    
    def get_device_by_name(self, device_name: str) -> Optional[dict]:
        """Get full device info by name"""
        device_name_lower = device_name.lower().strip()
        
        for device in self.devices:
            name = str(device.get("name", "")).lower()
            if device_name_lower in name or name in device_name_lower:
                return device
        
        return None
    
    def get_device_by_serial(self, serial: str) -> Optional[dict]:
        """Get device info by serial number"""
        serial_lower = serial.lower().strip()
        
        for device in self.devices:
            dev_serial = str(device.get("serial", "")).lower()
            if dev_serial == serial_lower:
                return device
        
        return None
    
    def get_all_devices(self) -> list[dict]:
        """Get all devices from inventory"""
        return self.devices.copy()
    
    def is_accessible(self, device_name: str) -> bool:
        """Check if device is marked as accessible"""
        device = self.get_device_by_name(device_name)
        if device:
            accessible = str(device.get("accessible", "")).lower()
            return accessible == "yes"
        return False
    
    def is_dnaas_device(self, device_name: str) -> bool:
        """Check if a device name matches a DNAAS device pattern"""
        name_lower = device_name.lower()
        return "dnaas" in name_lower or self.get_device_by_name(device_name) is not None
    
    def get_device_type(self, device_name: str) -> str:
        """Determine device type from name"""
        name_lower = device_name.lower()
        
        if "superspine" in name_lower:
            return "superspine"
        elif "spine" in name_lower:
            return "spine"
        elif "leaf" in name_lower:
            return "leaf"
        else:
            return "unknown"
    
    def __len__(self):
        return len(self.devices)
    
    def __repr__(self):
        return f"DNAASLookup({len(self.devices)} devices)"











